from dataclasses import make_dataclass
import datetime
from fcntl import F_GETFD

from django.contrib import messages
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render,redirect
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.core.mail import send_mail
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.db.models import Sum
from ehrms .models import taskdata,list,HR, Employs,employ_drop,CustomUser,payslip_request,documents_setup,empdocs,ad_salary,duration_year,duration_months,Reimbursement,adminnav,employnav,types,typeofd,year,LeaveReportEmploy,employ_add_form,employ_payslip,NotificationEmploy,checkin,checkout,employ_tax_form,reimbursementsetup1,documents_setup1,company_details
from io import BytesIO
from django.template.loader import get_template
from django.core.paginator import Paginator
from notifications.signals import *
from notifications.models import Notification

from worktride.settings import BASE_DIR
from .models import *
from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned  # Import MultipleObjectsReturned

# from xhtml2pdf import pisa  

from django import forms
from ehrms.forms import ScreenshotsForm

class HomeForm(forms.Form):
    name = forms.CharField(max_length=100)
    email = forms.EmailField()



from.models import project_drop
def Employ_home(request):
    

    # projects_drops=project_drop.objects.filter(parent_category=None).order_by('id')
    employ_obj=Employs.objects.get(admin=request.user.id)
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname

    findcompid=Companys.objects.filter(id=data.companyid.id).first()
    palnid=findcompid.plantype
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if palnid == "1":
        if data2:
            s=employnav.objects.filter(is_name_exist=1,hr_options=1,show1=1)
        elif project_manager:
            s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1,show1=1)
        elif teamleadop:
            s=employnav.objects.filter(is_name_exist=1,is_tl_option=1,show1=1)
        else:
            s=employnav.objects.filter(is_name_exist=1,employ_options=1,show1=1)
        projects_drops=project_drop.objects.filter(parent_category=None,show1=1).order_by('id')
        admin_drops=employ_drop.objects.filter(parent_category=None,show1=1).order_by('id')
    elif palnid == "2":
        if data2:
            s=employnav.objects.filter(is_name_exist=1,hr_options=1,show2=1)

        elif project_manager:
            s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1,show2=1)


        elif teamleadop:
            s=employnav.objects.filter(is_name_exist=1,is_tl_option=1,show2=1)
     
        else:
            s=employnav.objects.filter(is_name_exist=1,employ_options=1,show2=1)
        projects_drops=project_drop.objects.filter(parent_category=None,show2=1).order_by('id')
        admin_drops=employ_drop.objects.filter(parent_category=None,show2=1).order_by('id')
        
    elif palnid == "3":
        if data2:
            s=employnav.objects.filter(is_name_exist=1,hr_options=1,show3=1)

        elif project_manager:
            s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1,show3=1)

        elif teamleadop:
            s=employnav.objects.filter(is_name_exist=1,is_tl_option=1,show3=1)

        else:
            s=employnav.objects.filter(is_name_exist=1,employ_options=1,show3=1)
        projects_drops=project_drop.objects.filter(parent_category=None,show3=1).order_by('id')
        admin_drops=employ_drop.objects.filter(parent_category=None,show3=1).order_by('id')
    else:
        if data2:
            s=employnav.objects.filter(is_name_exist=1,hr_options=1)
        elif project_manager:
            s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)

        elif teamleadop:
            s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)

        else:
            s=employnav.objects.filter(is_name_exist=1,employ_options=1)
            
        projects_drops=project_drop.objects.filter(parent_category=None,show3=1).order_by('id')
        admin_drops=employ_drop.objects.filter(parent_category=None,show3=1).order_by('id')


    notification1=NotificationEmploy.objects.filter(employ_id=employ_obj.id)
    notifications1=NotificationEmploy.objects.filter(employ_id=employ_obj.id).count()
    notifications2=admin_project_create.objects.filter(is_status=0,read=0,admin_id=employ_obj.id).count() 
    notifications3=admin_project_create.objects.filter(read=0,admin_id=employ_obj.id)
    

    user=Employs.objects.filter(admin=request.user.id).first()
    userid=user.id
    projectm=admin_project_create.objects.filter(admin_id = userid,read=1)
    projectm1=admin_project_create.objects.filter(admin_id = userid)

    da1=Employs.objects.filter(admin=request.user.id).first()
    da2=da1.id
    project_details = Project.objects.filter(o_id=da2).prefetch_related('teammember_set') # Adjust this query to filter projects as needed

    emp_details=Employs.objects.filter(admin=request.user.id) 
    emp_detail=Employs.objects.get(admin=request.user.id) 
    today = date.today()
    next_week = today + timedelta(days=6)
    tlop=TeamMember.objects.filter(employee=data1,is_team_lead=1)
    tloptions=employnav.objects.filter(is_name_exist=1,is_tl_option=1)

    # Query employees whose birthdays fall within the next week
    employees2 = employ_add_form.objects.filter(dob__day__gte=today.day, dob__day__lte=next_week.day, dob__month=today.month).order_by('dob')
    
    for employee in employees2:
        if employee.dob.day == today.day:
            subject =  'Heartfelt Birthday Greetings for a Valued Team Member 🎉'  
            message = f"Dear {employee.firstname1} {employee.lastname1},\n\n “Wishing you a day filled with joy, laughter, and unforgettable moments as you celebrate another year of wonderful experiences and accomplishments. 🎂 Happy Birthday!”\n\nYour presence and contributions have greatly enriched our team, bringing forth creativity, dedication, and a positive spirit. On this special day, we recognize and appreciate all that you do to make our workplace a better and more enjoyable environment for everyone.\n\nBest wishes,\n HR,\n DevelopTrees Refinding IT solutions\n"
            send_mail(subject, message, 'saipathivada1234@gmail.com', [employee.email2])

    team_members = TeamMember.objects.filter(is_team_lead=1,employee=data1).first()
    task = tlassigntask.objects.filter(employid=data1)
    k=taskdata.objects.filter(team_lead=data1).last()
    if k:
        saved_line = k.p_url
        if ' ' in saved_line:
            description, url = saved_line.rsplit(' ', 1)
        else:
            description = saved_line
            url = ""
    else:
        description = ""
        url = ""

    if k:
        saved_line = k.p_url
        if ' ' in saved_line:
            description, url = saved_line.rsplit(' ', 1)
            request.session['meeting_description'] = description
            request.session['meeting_url'] = url
        else:
            request.session['meeting_description'] = saved_line
            request.session['meeting_url'] = ""
    else:
            request.session['meeting_description'] = ""
            request.session['meeting_url'] = ""
    teammemberid=CustomUser.objects.filter(id=request.user.id).first()
    teammemberlink=teammemberid.id
    k1=Meeting.objects.filter(team_member=teammemberlink).last()
 
    if k1:
        saved_line1 = k1.meeting_url
        if ' ' in saved_line1:
            description1, url1 = saved_line1.rsplit(' ', 1)
        else:
            description1 = saved_line1
            url1 = ""
    else:
        description1 = ""
        url1 = ""

    if team_members:
       team1 = team_members.project
       employee_ids = TeamMember.objects.filter(project=team1).values_list('employee_id', flat=True)
       project_ids = TeamMember.objects.filter(project=team1).values_list('project_id', flat=True)
       employees = Employs.objects.filter(id__in=employee_ids)
       projects = Project.objects.filter(id__in=project_ids)
    else:
    # No team members found
      team1 = None
      employees = None
      projects = None
    emp_detail=Employs.objects.get(admin=request.user.id)
    teaml=TeamMember.objects.filter(employee=emp_detail).first()
    if teaml:
        teamlead=teaml.is_team_lead
    else:
        teamlead=None        

    missing=employ_add_form.objects.filter(student_id_id=employ_obj)
    
    missing_doc=empdocs.objects.filter(employ_id=employ_obj)
    
    try:
        user_data=employ_add_form.objects.get(student_id_id=employ_obj)
        is_form_complete=all(getattr(user_data,field) for field in ['pan','ifsecode','acno','beneficiaryname','phno','gender1','dob','address1','heq','aadharno','bloodgroup','profile_pic'])
    except employ_add_form.DoesNotExist:
        is_form_complete = False 
    # try:
    #     user_data_doc = empdocs.objects.filter(employ_id=employ_obj).first()
    #     if user_data_doc:
    #      is_form_complete_doc = all(getattr(user_data_doc, field) for field in ['documenttype1', 'imagefile', 'description'])
    #     else:
    #      is_form_complete_doc = False
    # except empdocs.DoesNotExist:
    #     is_form_complete_doc = False   

    try:
        user_data_doc = empdocs.objects.filter(employ_id=employ_obj).first()
        if user_data_doc:
         is_form_complete_doc = all(getattr(user_data_doc, field) for field in ['documenttype1', 'imagefile', 'description'])
        else:
         is_form_complete_doc = False
    except empdocs.DoesNotExist:
        is_form_complete_doc = False 
    try:
        user_data_doc = empdocs.objects.filter(employ_id=employ_obj).first()
        is_form_complete_doc = user_data_doc is not None and all(getattr(user_data_doc, field) for field in ['documenttype1', 'imagefile', 'description'])
    except empdocs.DoesNotExist:
        is_form_complete_doc = False

    try:
        user_profile_pic = Employs.objects.filter(admin=request.user.id).first()
        if user_profile_pic:
         is_profile_pic_complete = all(getattr(user_profile_pic, field) for field in ['profile_pic'])
        else:
         is_profile_pic_complete = False
    except empdocs.DoesNotExist:
        is_profile_pic_complete = False 
    items_per_page = 10
    project_details = Project.objects.filter(o_id=da2).prefetch_related('teammember_set') # Adjust this query to filter projects as needed
    paginator = Paginator(project_details, items_per_page)
    page = request.GET.get('page')

    try:
        project_details = paginator.page(page)
    except PageNotAnInteger:
        project_details = paginator.page(1)
    except EmptyPage:
        project_details = paginator.page(paginator.num_pages) 
    return render(request,"employ-template/employ_home_template.html",{'description': description, 'url': url,'description1': description1, 'url1': url1, 'palnid':palnid,'teamlead':teamlead,'organization_name':organization_name,'project_details':project_details,'projectm':projectm,'data2':data2,'tloptions':tloptions,'tlop':tlop,'k':k,'k1':k1,'task':task,'team_members':team_members,'projects':projects,'employees':employees,'data':data,'employees2': employees2,'is_form_complete_doc':is_form_complete_doc,'user_data_doc':user_data_doc if is_form_complete_doc else None,'is_form_complete':is_form_complete,'user_data':user_data if is_form_complete else None,'emp_detail':emp_detail,'emp_details':emp_details,"employ_obj":employ_obj,"notification1":notification1,'missing':missing,'notifications2':notifications2,'notifications3':notifications3,'projectm1':projectm1,
           'is_profile_pic_complete':is_profile_pic_complete,'user_profile_pic':user_profile_pic if is_profile_pic_complete else None  })

############### project views ################

from .models import employperformance, Project, employ_nav  # Add this line to import employ_nav

# The rest of your code remains unchanged...

# def employper1(request):

#     data=Employs.objects.filter(admin=request.user.id).first()
#     data1=data.id
#     data2=data.hroptions
#     project_manager=data.projectmanagerop
#     compid=data.companyid
#     teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
#     regname=Companys.objects.filter(id=data.companyid.id).first()
#     organization_name=regname.organizationname
 
 

#     user=Employs.objects.filter(admin=request.user.id).first()
#     userid=user.id
#     projectm=admin_project_create.objects.filter(admin_id = userid)
#     da1 = Employs.objects.filter(admin=request.user.id).first()
#     da2 = da1.id
#     # admin = AdminHod.objects.get(id=1)
#     h = HR.objects.all()
#     employs_all = Employs.objects.all()
#     # admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
#     emm=Employs.objects.filter(admin=request.user.id).first()
#     em1=emm.companyid
#     pr = Employs.objects.filter(admin=request.user.id).first()
#     pr1 = pr.id
#     projects = Project.objects.filter(o_id=pr1)
#     employees = Employs.objects.all()
#     tsk1_data = employperformance.objects.filter(employ_name__companyid=em1,project_name__o_id=pr1,employ_name__hroptions=0).select_related('project_task1')

#     for task in tsk1_data:
#         try:
#             project = Project.objects.get(id=task.project_name.id)
#             task.project_name = project  # Assign the Project instance directly
#         except Project.DoesNotExist:
#             task.project_name = None  # Handle the case when the project is not found

#     if request.method == "POST":
#         project_id = request.POST.get('project_name')
#         employ_id = request.POST.get('employ_name')
#         task_name = request.POST.get('task_name')
#         performance = request.POST.get('performance')

#         try:
#             selected_project = Project.objects.get(id=project_id)
#         except Project.DoesNotExist:
#             selected_project = None  

#         try:
#             selected_employee = Employs.objects.get(id=employ_id)
#         except Employs.DoesNotExist:
#             selected_employee = None 
#         try:
#             selected_task=tlassigntask.objects.get(task=task_name)
#         except tlassigntask.DoesNotExist:
#             selected_task=None




#         k = employperformance(
#             project_name=selected_project,
#             employ_name=selected_employee,
#             performance=performance,
#             task_name1=task_name
            
#         )
#         k.save()
#         messages.success(request,"Sucessfully Submitted Employee Performance")


#     return render(request, "employ-template/employ-employper.html", {
        
        
#         'tsk1_data': tsk1_data,
#         'projects': projects,
#         'employees': employees,
#         'user': user,
#         'da2': da2,
#         'h': h,
#         'organization_name':organization_name,
        
#         'projectm': projectm,
#         'employs_all': employs_all,
#         'data': data,
        
#     })


# def employper1(request):
#     # Fetch necessary data from the database
#     data = Employs.objects.filter(admin=request.user.id).first()
#     em1 = data.companyid
#     pr1 = data.id
#     projects = Project.objects.filter(o_id=pr1)
#     employees = Employs.objects.all()
#     tsk1_data = employperformance.objects.filter(employ_name__companyid=em1, project_name__o_id=pr1, employ_name__hroptions=0).select_related('project_task1')

#     # Pagination logic
#     items_per_page = 10
#     paginator = Paginator(tsk1_data, items_per_page)
#     page = request.GET.get('page')

#     try:
#         tsk1_data = paginator.page(page)
#     except PageNotAnInteger:
#         tsk1_data = paginator.page(1)
#     except EmptyPage:
#         tsk1_data = paginator.page(paginator.num_pages)

#     if request.method == "POST":
#         project_id = request.POST.get('project_name')
#         employ_id = request.POST.get('employ_name')
#         task_name = request.POST.get('task_name')
#         performance = request.POST.get('performance')

#         # Check if task exists for the selected employee and project
#         task_exists = employperformance.objects.filter(project_name=project_id, employ_name=employ_id).exists()

#         if task_name:
#             try:
#                 selected_project = Project.objects.get(id=project_id)
#                 selected_employee = Employs.objects.get(id=employ_id)

#                 # Save performance data only if the task exists
#                 if task_exists:
#                     k = employperformance(
#                         project_name=selected_project,
#                         employ_name=selected_employee,
#                         performance=performance,
#                         task_name1=task_name
#                     )
#                     k.save()
#                     messages.success(request, "Successfully submitted employee performance")
#                 else:
#                     messages.warning(request, "No task assigned for the selected employee and project.")
#             except (Project.DoesNotExist, Employs.DoesNotExist) as e:
#                 messages.error(request, "Error: Project or Employee not found.")
#         else:
#             messages.warning(request, "No task assigned for the selected employee and project.")

#     return render(request, "employ-template/employ-employper.html", {
#         'tsk1_data': tsk1_data,
#         'projects': projects,
#         'employees': employees,
#         'user': data,
#     })


def employper1(request):
    pr = Employs.objects.filter(admin=request.user.id).first()
    pr1 = pr.id
    all_projects = Project.objects.filter(o_id=pr1)
    unique_projects = {}
    for project in all_projects:
        if project.p_name not in unique_projects:
            unique_projects[project.p_name] = project

    projects = unique_projects.values()
    if request.method == "POST":
        project_name = request.POST.get('project_name')
        tl_name = request.POST.get('tl_name')
        employ_name = request.POST.get('employ_name')
        task_name1 = request.POST.get('task_name1')  
        performance = request.POST.get('performance')
        status = request.POST.get('status')
        
        employ_performance = employperformance(
                project_name=project_name,
                tl_name_id=tl_name,
                employ_name_id=employ_name,
                # project_task1=project_task1,
                performance=performance,
                task_name1=task_name1,
                status="4",
            

            )
        employ_performance.save()
        related_tasks = tlassigntask.objects.filter(employid=employ_name,description=task_name1)
        for task in related_tasks:
            task.status ="4"
            task.save()
       
        messages.success(request, "Employee performance submitted successfully.")
    return render(request, "employ-template/employ-employper.html",{'projects': projects})

def get_project_employees(request, project_name):
    team_leads = TeamMember.objects.filter(project_id__p_name=project_name, is_team_lead=True)
    team_lead_data = [{'id': lead.employee.id, 'first_name': lead.employee.first_name, 'last_name': lead.employee.last_name} for lead in team_leads]
    return JsonResponse({'team_lead': team_lead_data})

def get_project_employees1(request, project_name, team_lead_id):
    team_leads = TeamMember.objects.filter(employee_id=team_lead_id, project_id__p_name=project_name, is_team_lead=True)
    if team_leads.exists():
        team_lead = team_leads.first()
        team_members = TeamMember.objects.filter(project_id=team_lead.project_id, is_team_lead=False)
        team_members_data = [{'id': member.employee.id, 'first_name': member.employee.first_name, 'last_name': member.employee.last_name} for member in team_members]
        return JsonResponse({'team_members': team_members_data})
    else:
        return JsonResponse({'team_members': []})

from django.http import JsonResponse

def get_employee_tasks(request, project_name, employee_id):
    tasks = tlassigntask.objects.filter(project_id__p_name=project_name, employid=employee_id).exclude(status="4")
    unique_descriptions = set(task.description for task in tasks)
    tasks_data = [{'description': description} for description in unique_descriptions]
    return JsonResponse({'tasks': tasks_data})





from django.db.utils import IntegrityError

from django.db.models import Q

# def create_proj1(request):

#     data=Employs.objects.filter(admin=request.user.id).first()
#     data1=data.id
#     data2=data.hroptions
#     project_manager=data.projectmanagerop
#     compid=data.companyid
#     regname=Companys.objects.filter(id=data.companyid.id).first()
#     organization_name=regname.organizationname
#     teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
 
 

#     em=Employs.objects.filter(admin=request.user.id).first()
#     em1=em.id
#     em2=em.companyid
#     pname=admin_project_create.objects.filter(admin_id=em1)
#     team_lead_employee = None
#     selected_team_lead_id = None
#     # s=adminnav.objects.filter(parent_category=None).order_by('id')
#     team_members = Employs.objects.filter(companyid=em2,hroptions=0,projectmanagerop=0)

#     if request.method == 'POST':
#         p_name = request.POST['p_name']
#         p_desc = request.POST['p_desc']
#         pr_deadline = request.POST['project_deadline_date']
#         # project_manager = request.POST['manager_name']
#         # status = request.POST['status']
#         o_id = Employs.objects.get(admin=request.user.id)

#         selected_employee_ids = request.POST.getlist('selected_employees[]')
#         pro=admin_project_create.objects.filter(project_name=p_name).first()
#         proid=pro.id 
#         prom=admin_project_create.objects.get(id=proid)

#         project = Project.objects.create(
#             p_name=p_name,
#             p_desc=p_desc,
#             o_id=o_id,
#             # project_manager=project_manager,
#             project_deadline=pr_deadline,
#             # status=status
#             proj=prom
            
#         )

#         selected_team_lead_name = request.POST.get('team_lead')
#         existing_team_member_ids = TeamMember.objects.filter(project=project).values_list('employee_id', flat=True)
#         available_team_members = Employs.objects.exclude(id__in=existing_team_member_ids)
#         team_lead_employee = Employs.objects.get(id=selected_team_lead_name)
        
#         if team_lead_employee:
#             selected_team_lead_id = team_lead_employee.id
#         # Check if a TeamMember entry with the same project and employee already exists
#         existing_team_member = TeamMember.objects.filter(project=project, employee=team_lead_employee.id, is_team_lead=True).first()
#         if not existing_team_member:
#             # Create the team lead only if it doesn't exist
#             team_lead_member = TeamMember(
#                 project=project,
#                 employee=team_lead_employee,
#                 is_team_lead=True,
#             )
#             team_lead_member.save()

#         # Mark the selected team lead as a team lead in the database
        

#         for emp_id in selected_employee_ids:
#             try:
#                 employee = Employs.objects.get(id=emp_id)
#                 TeamMember.objects.create(project=project, employee=employee, is_team_lead=False)
#                 # available_team_members = available_team_members.exclude(id=emp_id)
#             except IntegrityError:
#                 existing_entry = TeamMember.objects.get(project=project, employee=employee, is_team_lead=False)
#                 existing_entry.save()
        
#         # Fetch available team members (exclude team leads)
#         # team_members = Employ.objects.filter(
#         #         Q(teammember__isnull=True) | Q(teammember__project__status='completed')
#         #         ).distinct()

#         messages.success(request, 'Project created successfully and employees assigned.')
#         return redirect('/Employ_home')
#     return render(request, 'employ-template/OrgCreateProject.html', {'organization_name':organization_name,'data':data,'pname':pname,"team_members": team_members, "selected_team_lead_id": selected_team_lead_id})

def get_teamb(request, designation_employ__designation_name):
    em=Employs.objects.filter(admin=request.user.id).first()
    em1=em.id
    em2=em.companyid
    alreadyteam=TeamMember.objects.values('employee')
    team_members =  Employs.objects.filter(companyid=em2,hroptions=0,projectmanagerop=0,designation_employ__designation_name=designation_employ__designation_name).exclude(id__in=alreadyteam)
    data = [{'id': member.id, 'name': f"{member.first_name} {member.last_name}"}
            for member in team_members]
    return JsonResponse({'team_members': data})

from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import IntegrityError
from .models import Project, TeamMember, Employs, admin_project_create, Companys

def create_proj1(request):
    try:
        # Fetching necessary data and initializations
        employ = Employs.objects.filter(admin=request.user.id).first()
        company_id = employ.companyid
        organization_name = Companys.objects.filter(id=company_id.id).first().organizationname
        team_lead_op = TeamMember.objects.filter(employee=employ.id, is_team_lead=1)
        creater = Employs.objects.filter(projectmanagerop=1, companyid=employ.id)
        project_name = admin_project_create.objects.filter(admin_id=employ.id)

        # Fetching available team members and distinct department designations
        team_members = Employs.objects.filter(companyid=company_id, hroptions=0, projectmanagerop=0)
        distinct_department_designations = team_members.exclude(designation_employ__designation_name__isnull=True).exclude(designation_employ__designation_name__exact='').values('designation_employ__designation_name').distinct()
        prom = project_name.first()    
        team_lead_employee = None
        selected_team_lead_id = None

        if request.method == 'POST':
            # Handling form data from POST request
            p_name = request.POST['p_name']
            p_desc = request.POST['p_desc']
            pr_deadline = request.POST['project_deadline_date']
            team_lead_id = request.POST['team_lead']

            # Creating the project instance
            project = Project.objects.create(
                p_name=p_name,
                p_desc=p_desc,
                o_id=employ,
                project_deadline=pr_deadline,
                project_manager=employ.first_name,
                companyid=company_id,
                proj=prom,
            )

            # Updating the status for the project in the admin_project_create table
            update_status = admin_project_create.objects.filter(project_name=p_name)
            for status in update_status:
                status.status = "On Going..."
                status.save()

            # Handling team lead assignment
            team_lead_employee = Employs.objects.get(id=team_lead_id)
            if team_lead_employee:
                selected_team_lead_id = team_lead_employee.id

            existing_team_member = TeamMember.objects.filter(project=project, employee=team_lead_employee.id, is_team_lead=True).first()
            if not existing_team_member:
                team_lead_member = TeamMember(
                    project=project,
                    employee=team_lead_employee,
                    is_team_lead=True,
                )
                team_lead_member.save()

            # Handling selected team members
            selected_employee_ids = request.POST.getlist('selected_employees[]')
            for emp_id in selected_employee_ids:
                try:
                    employee = Employs.objects.get(id=emp_id)
                    TeamMember.objects.create(project=project, employee=employee, is_team_lead=False)
                except IntegrityError:
                    existing_entry = TeamMember.objects.get(project=project, employee=employee, is_team_lead=False)
                    existing_entry.save()

            # Redirecting to home page after successful project creation
            messages.success(request, 'Project created successfully and employees assigned.')
            return redirect('/Employ_home')
        
        # Rendering the template with necessary context data
        return render(request, 'employ-template/OrgCreateProject.html', {
            'departments': distinct_department_designations,
            'organization_name': organization_name,
            'data': employ,
            'pname': project_name,
            "selected_team_lead_id": selected_team_lead_id
        })

    except Exception as e:
        # Handling any unexpected errors
        messages.error(request, f'An error occurred: {str(e)}')
        return redirect("/Employ_home")    





from .models import Project, projecttask
from django.shortcuts import render
from django.utils import timezone

def projecttask2(request):
    pr = Employs.objects.filter(admin=request.user.id).first()
    pr1 = pr.id
    all_projects = Project.objects.filter(o_id=pr1)
    unique_projects = {}
    for project in all_projects:
        if project.p_name not in unique_projects:
            unique_projects[project.p_name] = project

    projects = unique_projects.values()


    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = Employs.objects.filter(admin=request.user.id).first()
    da2 = da1.id

    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
 
 

    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    # admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')



    if request.method == "POST":
        p_id = request.POST['p_name']  # Assuming 'p_name' contains the ID of the project
        l_name = request.POST['l_name']
        task_names = request.POST.getlist('task_name[]')
        # deadlines = request.POST.getlist('deadline[]')
        # Fetch the Project instance based on the ID obtained
        project_instance = Project.objects.get(pk=p_id)
        for task_name in task_names:
            # Assign the project instance (not just the ID) to the projecttask's p_name field
            task = projecttask(p_name=project_instance, tasks=task_name, l_name=l_name)
            task.save()
        team_leads = TeamMember.objects.filter(project=project_instance, is_team_lead=True)
        for lead in team_leads:
            notify.send(sender=request.user, recipient=lead.pr,
                            verb='Task Assigned',
                            description=f'A new task "{task_name}" has been assigned to your team.')
         
            messages.success(request, "New Task created successfully")
            return redirect('/Employ_home')
    return render(request, 'employ-template/form123.html', {'organization_name':organization_name,'projects': projects,'user':user,'da1':da1,'da2':da2,'h':h,'projectm':projectm,'data':data,'employs_all':employs_all,})

def get_team_members1(request, project_id):
    team_members = TeamMember.objects.filter(project_id=project_id)
    data = [{'id': member.employee.id, 'name': f"{member.employee.first_name} {member.employee.last_name}"}
            for member in team_members]
    return JsonResponse({'team_members': data})


def get_team_lead1(request, project_id):
    pr = Employs.objects.filter(admin=request.user.id).first()
    pr1 = pr.id
    all_projects = Project.objects.filter(o_id=pr1)
    unique_projects = {}
    for project in all_projects:
        if project.p_name not in unique_projects:
            unique_projects[project.p_name] = project

    projects = unique_projects.values()

    all_projects = Project.objects.filter(p_name=project_id)
    unique_project_names = set()
    unique_projects = []
    team_lead = TeamMember.objects.filter(project_id__p_name=project_id,is_team_lead=1)

    for project in all_projects:
        if project.p_name not in unique_project_names:
            unique_project_names.add(project.p_name)
            unique_projects.append(project)
    if request.method == "POST":
        p_id = request.POST['p_name'] 
        l_name = request.POST['l_name']
        task_names = request.POST.getlist('task_name[]')
        description=request.POST['description']

        project_instance = Project.objects.get(pk=p_id)

        for task_name in task_names:
        
            task = projecttask(p_name=project_instance, tasks=task_name, l_name=l_name,description=description)
            task.save()
        messages.success(request, "New Task created successfully") 
        team_leads = TeamMember.objects.filter(project=project_instance, is_team_lead=True)

            # Send notification to each team lead about the new task assignment
        for lead in team_leads:
            notify.send(
                sender=request.user,
                recipient=lead.employee.admin,
                verb='Task Assigned',
                    description=f'A new task "{task_name}" has been assigned to your team for project "{project_instance.p_name}".'
                )


    team_lead = TeamMember.objects.filter(project_id__p_name=project_id, is_team_lead=True) if project_id else []
 

    return render(request,"employ-template/form123.html",{'team_lead':team_lead,'unique_projects':unique_projects,'projects':projects})


from django.shortcuts import render, redirect
from .models import admin_project_create, taskdata, TeamMember

def display2(request):
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    pr = Employs.objects.filter(admin=request.user.id).first()
    pr1 = pr.id
    all_projects = Project.objects.filter(o_id=pr1)
    unique_projects = {}
    for project in all_projects:
        if project.p_name not in unique_projects:
            unique_projects[project.p_name] = project
    projects = unique_projects.values()
    tl=TeamMember.objects.filter(is_team_lead=1)
    if request.method == 'POST':
        p_name = request.POST.get('p_name')
        p_url = request.POST.get('p_url')
        team_lead = request.POST.get('team_lead')
        # date = request.POST['date']
        # time = request.POST['time']

        # Create a new project
        p1 = taskdata(
            p_name=p_name,
            p_url=p_url,
            team_lead=team_lead,
            # date=date,
            # time=time,
        )
        p1.save()
        messages.success(request,"Review Link sent successfully")
        return redirect('/display2')
    return render(request, 'employ-template/data.html', {'organization_name':organization_name, 'tl':tl,'projects': projects,'data':data})


  
def performancetask1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.id

    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
 
 
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = Employs.objects.filter(admin=request.user.id).first()
    da2 = da1.id
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    projects = Project.objects.all()  
    items_per_page = 10
    performance_data = employperformance.objects.filter(employ_name_id__companyid=em1)
    paginator = Paginator(performance_data, items_per_page)
    page = request.GET.get('page')

    try:
        performance_data = paginator.page(page)
    except PageNotAnInteger:
        performance_data = paginator.page(1)
    except EmptyPage:
        performance_data = paginator.page(paginator.num_pages)
    return render(request, 'employ-template/emptaskper.html', {'organization_name':organization_name,'projects':projects,'data':data,'performance_data': performance_data,'user':user,'da1':da1,'da2':da2,'h':h,'projectm':projectm,'employs_all':employs_all})
import openpyxl
from django.http import HttpResponse
from django.db.models import Avg

def download_all_performance_data_excel1(request):
    performance_data = employperformance.objects.all()

    # Create an Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Employee ID", "Employee Name", "Project Name", "Date", "Performance"])  # Include "Date" in the header

    for performance in performance_data:
        try:
            employee = Employs.objects.get(id=performance.employ_name.id)
            project_name = performance.project_name.p_name
            worksheet.append([employee.empid, employee.first_name, project_name, performance.date, performance.performance])  # Include date in the row
        except Employs.DoesNotExist:
            pass  # Handle the case where the employee doesn't exist

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="all_performance_data.xlsx"'
    workbook.save(response)

    return response

def performanceallproject1(request):
    emm = Employs.objects.filter(admin=request.user.id).first()
    em1 = emm.companyid
    data = Employs.objects.filter(admin=request.user.id).first()
    data1 = data.id
    data2 = data.hroptions
    project_manager = data.projectmanagerop
    compid = data.companyid
    regname = Companys.objects.filter(id=data.companyid.id).first()
    organization_name = regname.organizationname
    teamleadop = TeamMember.objects.filter(employee=data1, is_team_lead=1)
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = Employs.objects.filter(admin=request.user.id).first()
    da2 = da1.id
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    projects = Project.objects.all()
    items_per_page = 10
    tsk1_data = employperformance.objects.filter(employ_name__companyid=em1).values('employ_name').annotate(avg_performance=Avg('performance'))
    
    employees_data = []
    for task in tsk1_data:
        employ_id = task['employ_name']
        try:
            employee = Employs.objects.get(id=employ_id)
            task['employee_name'] = employee.first_name
            task['employee_id'] = employee.empid
            project_names = employperformance.objects.filter(employ_name=employee).values_list('project_name', flat=True).distinct()
            task['project_names'] = project_names
        except Employs.DoesNotExist:
            task['employee_name'] = "N/A"
            task['employee_id'] = "N/A"
            task['project_names'] = "N/A"
        employees_data.append(task)

    try:
        paginator = Paginator(employees_data, items_per_page)
        page = request.GET.get('page')
        employees_data = paginator.page(page)
    except PageNotAnInteger:
        employees_data = paginator.page(1)
    except EmptyPage:
        employees_data = paginator.page(paginator.num_pages)

    return render(request, 'employ-template/empavgper.html', {'organization_name': organization_name, 'projects': projects, 'data': data, 'tsk1_data': employees_data, 'user': user, 'da1': da1, 'da2': da2, 'h': h, 'projectm': projectm, 'employs_all': employs_all})


import openpyxl
from openpyxl.styles import Alignment

def download_excel1(request):
    tsk1_data = employperformance.objects.values('employ_name').annotate(avg_performance=Avg('performance'))

    # Create an Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Employee ID", "Employee Name", "Project Names", "Average Performance"])

    for task in tsk1_data:
        employ_id = task['employ_name']
        try:
            employee = Employs.objects.get(id=employ_id)
            task['employee_name'] = employee.first_name
            task['employee_id'] = employee.empid

            # Aggregate unique project names for the employee
            project_names = set(employperformance.objects.filter(employ_name=employee).values_list('project_name__p_name', flat=True))
            # Join unique project names with line breaks
            task['project_names'] = "\n".join(project_names)

        except Employs.DoesNotExist:
            task['employee_name'] = "N/A"
            task['employee_id'] = "N/A"
            task['project_names'] = "N/A"

        # Append the data to the Excel sheet
        worksheet.append([task['employee_id'], task['employee_name'], task['project_names'], task['avg_performance']])

    # Apply alignment style to the cell containing project names for line breaks
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Create an HttpResponse with the Excel file
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="tsk1_data.xlsx"'
    workbook.save(response)

    return response

from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import admin_project_create

def projectstatus1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id

    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
 

   
    da1 = Employs.objects.filter(admin=request.user.id).first()
    da2 = da1.id
    # projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()


    if request.method == "POST":

       


        projectm = admin_project_create.objects.filter(admin_id=request.user.id)


        
        for project in projectm:
            project_id = project.id
            status = request.POST.get(f'status-{project_id}')
            
            if status in ('Complete', 'Incomplete','Ongoing', 'On Hold'):
                project.status = status 
                project.read=False
                project.save()

        return redirect('admin_home')  # You can replace 'admin_home' with the actual URL you want to redirect to

    # Only get the projects assigned to the manager
    projectm = admin_project_create.objects.filter(companyid=em1,admin_id=da2)

    return render(request, "employ-template/projectstatus.html", {'organization_name':organization_name ,'projectm': projectm,'user':user,'da1':da1,'da2':da2,'h':h,'projectm':projectm,'data':data,'employs_all':employs_all})

def update_project_status(request):
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        project_id = data.get('projectId')
        status = data.get('status')
        # Update the project status in the database
        project = admin_project_create.objects.get(pk=project_id)
        project.status = status
        project.read=False
        project.save()
        same_project = Project.objects.get(proj=project)
        same_project.status = status

        same_project.save()

        return JsonResponse({'success': True})

    return JsonResponse({'success': False})

import pandas as pd
import plotly.express as px
from plotly.offline import plot
from datetime import timedelta
import json
from.models import AdminHod

def projectwise_task(request, pid):

    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    # if data2:
    #    s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    # elif project_manager:
    #     s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    
    # elif teamleadop:
    #     s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    # else:
    #     s=employnav.objects.filter(is_name_exist=1,employ_options=1)


    o_id = Employs.objects.get(admin=request.user.id)
    team_members = TeamMember.objects.filter(project=pid)
    team_member_scores = [(team_member, team_member.calculate_total_score_in_project()) for team_member in team_members]
    team_member_scores = sorted(team_member_scores, key=lambda x: x[1], reverse=True)
    project_details=Project.objects.filter(id=pid)
    task_details = Task.objects.filter(o_id_id=o_id, id=pid).all()
    # s = adminnav.objects.filter(parent_category=None).order_by('id')

    if request.method == 'GET':
        t_status = request.GET.get('t_status', 'total')  # Default to 'total' if not provided

        if t_status == 'total':
            task_details = Task.objects.filter(o_id_id=o_id, p_id=pid).all()
        elif t_status == 'completed':
            task_details = Task.objects.filter(o_id_id=o_id, p_id=pid, t_status='completed')
        elif t_status == 'in-progress':
            # task_details = Task.objects.filter(o_id_id=o_id, p_id=pid, t_status='in-progress').all()
            task_details = Task.objects.filter(o_id_id=o_id, p_id=pid).exclude(t_status='completed')

        

        tasks = task_details
        count_no_of_total_tasks = Task.objects.filter(o_id_id=o_id, p_id_id=pid).count()
        count_no_of_completed_tasks = Task.objects.filter(o_id_id=o_id, p_id_id=pid, t_status="completed").count()
        count_no_of_pending_tasks = count_no_of_total_tasks - count_no_of_completed_tasks
        task_data = []

        for task in tasks:
            now = timezone.now()
            task_score = task.calculate_task_score()
            remaining_time = task.calculate_remaining_time()
            progress_percentage = min(task_score / 10 * 100, 100)
            excess_points = max(task_score - 10, 0)  # Calculate the excess points (if any)
            neg_points = max(-task_score, 0)

            task_data.append({
                'task': task,
                'progress_percentage': progress_percentage,
                'excess_points': excess_points,
                'neg_points': neg_points,
                'task_score': task_score,
                'remaining_time': remaining_time,
            })
        
        mn=taskdata.objects.filter(p_name=pid).last()
        if mn:
            saved_line = mn.p_url
            if ' ' in saved_line:
                description, url = saved_line.rsplit(' ', 1)
            else:
                description = saved_line
                url = ""
        else:
            description = ""
            url = ""
            
        
        context = {
        
        "project_details": project_details,
        'task_total': count_no_of_total_tasks,
        'task_completed': count_no_of_completed_tasks,
        'task_pending': count_no_of_pending_tasks,
        'team_members': team_members,
        # 's': s,
        't_status':t_status,
        "task_details": task_details,
        'task_data': task_data,
        'data':data,
        'team_member_scores': team_member_scores,
        'mn':mn,
        'description':description,
        'url':url
        }
    
    return render(request, 'employ-template/ViewProjectwiseTasks1.html', context)
from django.shortcuts import render, redirect
from django.shortcuts import get_object_or_404
from django.contrib import messages
from django.db.models import Q
from .models import Project, Employs, TeamMember, admin_drop  # Make sure to import your models

# def update_data1(request, pid):
#     emm=Employs.objects.filter(admin=request.user.id).first()
#     em1=emm.companyid


#     # Get the project to be updated
#     project = get_object_or_404(Project, id=pid)
#     team_leader_info = project.get_team_leader()

#     # Get available team members excluding team leads and those already selected for the project
#     team_members = TeamMember.objects.filter(project=project)
#     selected_team_member_ids = team_members.values_list('employee_id', flat=True)
#     available_team_memb = Employs.objects.filter(companyid=em1,hroptions=0,projectmanagerop=0)
#     available_team_members=available_team_memb.exclude(id__in=selected_team_member_ids)

#     if request.method == 'POST':
#         # Retrieve updated project details from the form
#         p_name = request.POST['p_name']
#         p_desc = request.POST['p_desc']
#         pr_deadline = request.POST['project_deadline_date']
#         project_manager = request.POST['manager_name']
#         status = request.POST.get('status')

#         # Update the project
#         project.p_name = p_name
#         project.p_desc = p_desc
#         project.project_deadline = pr_deadline
#         project.project_manager = project_manager
#         project.status = status
#         project.save()

#         # Update team lead if changed
#         selected_team_lead_id = request.POST.get('team_lead')
#         if selected_team_lead_id:
#             try:
#                 team_lead = Employs.objects.get(id=selected_team_lead_id)
#                 project.teammember_set.update(is_team_lead=False)  # Remove team leader status from existing team leader

#                 # Create a new TeamMember if it doesn't exist
#                 team_member, created = TeamMember.objects.get_or_create(project=project, employee=team_lead)
#                 team_member.is_team_lead = True
#                 team_member.save()
#             except Employs.DoesNotExist:
#                 # Handle the case where the selected team leader doesn't exist
#                 messages.error(request, 'Selected team leader does not exist.')
#                 return redirect('/Employ_home')

#         # Update selected team members
#         selected_employee_ids = request.POST.getlist('selected_employees[]')
#         existing_team_member_ids = TeamMember.objects.filter(project=project).values_list('employee_id', flat=True)
#         for emp_id in selected_employee_ids:
#             emp_id = int(emp_id)
#             if emp_id not in existing_team_member_ids:
#                 employee = get_object_or_404(Employs, id=emp_id)
#                 TeamMember.objects.create(project=project, employee=employee, is_team_lead=False)

#         messages.success(request, 'Project updated successfully and employees assigned.')
#         return redirect('/Employ_home')

#     # Render the template with the updated available team members
#     return render(request, 'employ-template/editproject.html', {
  
#         "project": project,
#         "team_members": team_members,
#         "available_team_members": available_team_members,
#         'team_leader_info': team_leader_info,
        
#     })

# def update_data1(request, pid):
#     emm=Employs.objects.filter(admin=request.user.id).first()
#     em1=emm.companyid


#     # Get the project to be updated
#     project = get_object_or_404(Project, id=pid)
#     team_leader_info = project.get_team_leader()

#     # Get available team members excluding team leads and those already selected for the project
#     team_members = TeamMember.objects.filter(project=project)
#     team_mm=TeamMember.objects.filter(project=project,is_team_lead=1).first()
#     if team_mm:
#       empde=team_mm.employee.designation
#     else:
#         empde=None

#     selected_team_member_ids = team_members.values_list('employee_id', flat=True)
#     if empde:
#         available_team_members = Employs.objects.filter(companyid=em1,hroptions=0,projectmanagerop=0,designation=empde)

#         available_team_memb = Employs.objects.filter(companyid=em1,designation=empde)
   
#         available_team_memberss = available_team_memb.exclude(id__in=selected_team_member_ids)

#     #     available_team_memb = Employs.objects.filter(companyid=em1,hroptions=0,projectmanagerop=0,designation=empde)

#     # else:
#     #     available_team_memb = Employs.objects.filter(companyid=em1,hroptions=0,projectmanagerop=0)
#     # available_team_members=available_team_memb.exclude(id__in=selected_team_member_ids)
   


#     if request.method == 'POST':
#         # Retrieve updated project details from the form
#         p_name = request.POST['p_name']
#         p_desc = request.POST['p_desc']
#         pr_deadline = request.POST['project_deadline_date']
#         status = request.POST.get('status')
#         project_manager = request.POST['manager_name']


#         # Update the project
#         project.p_name = p_name
#         project.p_desc = p_desc
#         project.project_deadline = pr_deadline
#         project.status = status
#         project.project_manager = project_manager

#         project.save()

#         # Update team lead if changed
#         selected_team_lead_id = request.POST.get('team_lead')
#         if selected_team_lead_id:
#             try:
#                 team_lead = Employs.objects.get(id=selected_team_lead_id)
#                 project.teammember_set.update(is_team_lead=False)  # Remove team leader status from existing team leader

#                 # Create a new TeamMember if it doesn't exist
#                 team_member, created = TeamMember.objects.get_or_create(project=project, employee=team_lead)
#                 team_member.is_team_lead = True
#                 team_member.save()
#             except Employs.DoesNotExist:
#                 # Handle the case where the selected team leader doesn't exist
#                 messages.error(request, 'Selected team leader does not exist.')
#                 return redirect('/Employ_home')
#         selected_team_lead_name = request.POST.get('team_lead')
        
#         team_lead_employee = Employs.objects.get(id=selected_team_lead_name)
        
#         if team_lead_employee:
#             selected_team_lead_id = team_lead_employee.id
#         # Check if a TeamMember entry with the same project and employee already exists
#         existing_team_member = TeamMember.objects.filter(project=project, employee=team_lead_employee.id, is_team_lead=True).first()
#         if not existing_team_member:
#             # Create the team lead only if it doesn't exist
#             team_lead_member = TeamMember(
#                 project=project,
                
#                 employee=team_lead_employee,
#                 is_team_lead=True,
#             )
#             team_lead_member.save()

#         # Update selected team members
#         selected_employee_ids = request.POST.getlist('selected_employees[]')
#         existing_team_member_ids = TeamMember.objects.filter(project=project).values_list('employee_id', flat=True)
#         for emp_id in selected_employee_ids:
#             emp_id = int(emp_id)
#             if emp_id not in existing_team_member_ids:
#                 employee = get_object_or_404(Employs, id=emp_id)
#                 TeamMember.objects.create(project=project, employee=employee, is_team_lead=False)

#         messages.success(request, 'Project updated successfully and employees assigned.')
#         return redirect('/Employ_home')

#     # Render the template with the updated available team members
#     return render(request, 'employ-template/editproject.html', {
#         "project": project,
#         "team_members": team_members,
#         "available_team_members": available_team_members,
#         'team_leader_info': team_leader_info,
#         "available_team_memberss": available_team_memberss,

        
#     })

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import IntegrityError
from .models import Project, TeamMember, Employs,employ_designation3

def update_data1(request, pid):
    employ = Employs.objects.get(admin=request.user)
    company_id = employ.companyid
    project = get_object_or_404(Project, id=pid)
    team_leader_info = project.get_team_leader()

    team_lead_member = TeamMember.objects.filter(project=project).first()
    team_lead_designation = None
    if team_lead_member:
        team_lead_designation = team_lead_member.employee.designation

    team_members = TeamMember.objects.filter(project=project)
    selected_team_member_ids = team_members.values_list('employee_id', flat=True)
    available_team_members = Employs.objects.filter(companyid=company_id, hroptions=0, projectmanagerop=0, designation=team_lead_designation)
    available_team_memb = Employs.objects.filter(companyid=company_id, designation=team_lead_designation)
    available_team_memberss = available_team_memb.exclude(id__in=selected_team_member_ids)

    roles = employ_designation3.objects.filter(designation_name__in=Employs.objects.filter(companyid=company_id).values('designation_employ__designation_name')).distinct()

    if request.method == 'POST':
        p_name = request.POST.get('p_name')
        p_desc = request.POST.get('p_desc')
        pr_deadline = request.POST.get('project_deadline_date')
        project_manager = request.POST.get('manager_name')

        project.p_name = p_name
        project.p_desc = p_desc
        project.project_deadline = pr_deadline
        project.project_manager = project_manager
        project.save()

        selected_team_lead_id = request.POST.get('team_lead')
        if selected_team_lead_id:
            team_lead = Employs.objects.filter(id=selected_team_lead_id).first()
            if team_lead:
                # If team lead exists, update or create TeamMember record
                project.teammember_set.update(is_team_lead=False)
                team_member, created = TeamMember.objects.get_or_create(project=project, employee=team_lead)
                team_member.is_team_lead = True
                team_member.save()

        # Process selected team members
        selected_employee_ids = request.POST.getlist('selected_employees[]')
        existing_team_member_ids = TeamMember.objects.filter(project=project).values_list('employee_id', flat=True)
        for emp_id in selected_employee_ids:
            emp_id = int(emp_id)
            if emp_id not in existing_team_member_ids:
                employee = Employs.objects.filter(id=emp_id).first()
                if employee:
                    try:
                        TeamMember.objects.create(project=project, employee=employee, is_team_lead=False)
                    except IntegrityError:
                        continue  # Skip if the record already exists

        messages.success(request, 'Project updated successfully and employees assigned.')
        return redirect('/Employ_home')

    return render(request, 'employ-template/editproject.html', {
        "project": project,
        "team_members": team_members,
        "available_team_members": available_team_members,
        'team_leader_info': team_lead_member,
        'available_team_memberss': available_team_memberss,
        'team_leader_info': team_leader_info,
        'roles': roles,
    })

def delete_user1(request,employee_id , project_id):
        # Assuming your model has a field 'id' representing the user's ID
        employee = TeamMember.objects.get(employee_id=employee_id,project_id=project_id)
        employee.delete()
        referer = request.META.get('HTTP_REFERER')
        if referer:
          return redirect(referer)
        else:
          return redirect(reverse('home'))



############# project views end ##################
def employ_people_count(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid

    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    compid=data.companyid
    project_manager=data.projectmanagerop

    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
 
    
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    employs = Employs.objects.filter(companyid=em1)
    hr=HR.objects.all()
    emp = CustomUser.objects.filter(is_active=True,employs__companyid=em1).count()  # Count active employees
    dis = CustomUser.objects.filter(is_active=False,employs__companyid=em1).count()  # Count inactive employees
    total_employees_count = emp + dis 
    
    items_per_page = 10  # Adjust the number of items per page as needed

    # Retrieve all employees for pagination
    employees = Employs.objects.filter(companyid=em1)
    paginator = Paginator(employees,items_per_page)
    page_number = request.GET.get('page')

    try:
        employees = paginator.page(page_number)
    except PageNotAnInteger:
        employees = paginator.page(1)
    except EmptyPage:
        employees = paginator.page(paginator.num_pages)
    
    return render(request, "employ-template/employpeoples.html", {
    
     
        
        'employs': employs,
        'emp': emp,
        'dis': dis,
        'organization_name':organization_name,
        'total_employees_count': total_employees_count,
        
        'hr':hr,
        
        'data':data,
        
        
        'employees':employees
        
    })


def employ_findactive(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    compid=data.companyid
    project_manager=data.projectmanagerop

    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
 
    
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    employs = Employs.objects.filter(companyid=em1)
    hr=HR.objects.all()
    emp = CustomUser.objects.filter(is_active=True,employs__companyid=em1).count()  # Count active employees
    dis = CustomUser.objects.filter(is_active=False,employs__companyid=em1).count()  # Count inactive employees
    total_employees_count = emp + dis 
    
    items_per_page = 10  # Adjust the number of items per page as needed

    # Retrieve all employees for pagination
    employees = Employs.objects.filter(companyid=em1)

    paginator = Paginator(employees, items_per_page)
    page_number = request.GET.get('page')

    try:
        employees = paginator.page(page_number)
    except PageNotAnInteger:
        employees = paginator.page(1)
    except EmptyPage:
        employees = paginator.page(paginator.num_pages)
    
    return render(request, "employ-template/employ_findactive.html", {
    
     
        
        'employs': employs,
        'emp': emp,
        'dis': dis,
        'organization_name':organization_name,
        'total_employees_count': total_employees_count,
        
        'hr':hr,
        
        'data':data,
        
        
        'employees':employees
        
    })

def employ_findinactive(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    compid=data.companyid
    project_manager=data.projectmanagerop

    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
 
    
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    employs = Employs.objects.filter(companyid=em1)
    hr=HR.objects.all()
    emp = CustomUser.objects.filter(is_active=True,employs__companyid=em1).count()  # Count active employees
    dis = CustomUser.objects.filter(is_active=False,employs__companyid=em1).count()  # Count inactive employees
    total_employees_count = emp + dis 
    
    items_per_page = 10  # Adjust the number of items per page as needed

    # Retrieve all employees for pagination
    employees = Employs.objects.filter(companyid=data1,admin__is_active=False)

    paginator = Paginator(employees, items_per_page)
    page_number = request.GET.get('page')

    try:
        employees = paginator.page(page_number)
    except PageNotAnInteger:
        employees = paginator.page(1)
    except EmptyPage:
        employees = paginator.page(paginator.num_pages)
    
    return render(request, "employ-template/employ_findinactive.html", {
    
     
        
        'employs': employs,
        'emp': emp,
        'dis': dis,
        'organization_name':organization_name,
        'total_employees_count': total_employees_count,
        
        'hr':hr,
        
        'data':data,
        
        
        'employees':employees
        
    })






from .models import admin_home_drop
def edit_people_admin1(request,std):
    request.session['employ_id']=std
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')

    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
    da1 = AdminHod.objects.filter(admin=request.user.id).first()
    
    s = adminnav.objects.all()
    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
    data = Companys.objects.filter(usernumber=request.user.id).first()
    data1=data.id
    a=company_details.objects.filter(companyid=data1).first()



    # a=companylogo.objects.all()
    employ=Employs.objects.get(admin=std)
    return render(request,"admin-template/example.html",{'employ':employ,'user':user,'da1':da1,'admin_home_drops':admin_home_drops,'h':h,'projectm':projectm,'data':data,'employs_all':employs_all})

def edit_people(request,std):
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    comp_id=data.companyid
    a=company_details.objects.filter(companyid=comp_id).first()
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)



    request.session['employ_id']=std
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    # li=list.objects.all()
    employ=Employs.objects.get(admin=std)
    objss=employ.id 
    users=CustomUser.objects.get(id=std)
    datas=employ_add_form.objects.filter(student_id=objss)
    return render(request,"employ-template/people_employ.html",{'organization_name':organization_name,'data':data,'employ':employ,'datas':datas,'objss':objss,'users':users})

def enable_button1(request, std):

    request.session['employ_id']=std
    instance = CustomUser.objects.get(id=std)
    instance.is_active = True
    instance.save()
    return redirect('edit_people', std=std)  # Redirect to the detail view

def disable_button1(request, std):
    request.session['employ_id']=std
    instance = CustomUser.objects.get(id=std)
    instance.is_active = False
    instance.save()
    return redirect('edit_people',std=std)

def individual_tax_report(request,std):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')


    data=Employs.objects.filter(admin=request.user.id).first()
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)

    
    request.session['employ_id']=std
    employ=Employs.objects.all()
    tax=employ_tax_form.objects.filter(employ=std)
    
    leave_date=LeaveReportEmploy.objects.filter()
    return render(request,"employ-template/individual_taxReport.html",{'employ':employ,'tax':tax,'leave_date':leave_date,'data':data})

def individual_employ_documentreport(request,std):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    comp_id=data.companyid
    a=company_details.objects.filter(companyid=comp_id).first()
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)

   
    request.session['employ_id']=std
    employ_id=Employs.objects.all()
    doc=empdocs.objects.filter(employ_id=std)
    items_per_page = 10
    paginator = Paginator(doc, items_per_page)

    page = request.GET.get('page')
    try:
        doc = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        doc = paginator.page(1)
    except EmptyPage:
        # If the page is out of range (e.g. 9999), deliver the last page of results.
        doc = paginator.page(paginator.num_pages)

 
    return render (request,"employ-template/individual_employ_docreport.html",{'organization_name':organization_name,  'doc':doc,'paginator':paginator, 'employ_id':employ_id,'data':data})

# def individual_employ_attendance(request,std):
#     admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
#     request.session['employ_id']=std
#     employ_id=Employs.objects.all()
#     attend=checkin.objects.filter(employ_id=std)
 
#     return render (request,"admin-template/individual_attendance_report.html",{'attend':attend,'employ_id':employ_id,'admin_drops':admin_drops})


def individual_reimbursement_view1(request,std):
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid


    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    elif project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    
    elif teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    else:
        s=employnav.objects.filter(is_name_exist=1,employ_options=1)


    request.session['employ_id']=std
    # s=adminnav.objects.all()
    employ_id=Employs.objects.all()
    leaves=Reimbursement.objects.filter(employ_id=std)
    items_per_page = 10
    paginator = Paginator(leaves, items_per_page)

    page = request.GET.get('page')
    try:
        leaves = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        leaves = paginator.page(1)
    except EmptyPage:
        # If the page is out of range (e.g. 9999), deliver the last page of results.
        leaves = paginator.page(paginator.num_pages)

    return render(request,"employ-template/individual_employ_reimbursement.html",{'organization_name':organization_name,"leaves":leaves,"employ_id":employ_id,'data':data})

from django.views.generic import TemplateView
from .utils import Calendar1,caltable
# class CalendarView1(TemplateView):
#     template_name = "employ-template/calendarbasic.html"
    
#     def get_context_data(self, **kwargs,):
#         data=Employs.objects.filter(admin=self.request.user.id).first()
#         data1=data.id
#         data2=data.hroptions
#         project_manager=data.projectmanagerop
#         compid=data.companyid
#         regname=Companys.objects.filter(id=data.companyid.id).first()
#         organization_name=regname.organizationname
#         teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
  
    


    
    

#         tlop=TeamMember.objects.filter(employee=data1,is_team_lead=1)
#         tloptions=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
        


#         context = super(CalendarView1,self).get_context_data(**kwargs)
#         d = get_date(self.request.GET.get('month', None))
#         cal = Calendar(d.year, d.month,d.day)
#         html_cal = cal.formatmonth(self.request,withyear=True)
#         days_list,weekoff_days = caltable(self,d.year,d.month,self.request)
#         employlevsheetcausel = employlev.objects.filter(leave_id=1,companyid=compid).first()
#         employlevsheetmedical=employlev.objects.filter(leave_id=2,companyid=compid).first()
#         employlevsheetearned=employlev.objects.filter(leave_id=3,companyid=compid).first()
#         employs = Employs.objects.get(admin=self.request.user)
#         email = data.email
#         todaysec = datetime.now().date()

#         check_in = checkin.objects.filter(date=todaysec, empid=email).first()
#         check_out = checkout.objects.filter(date=todaysec, empid=email).first()


#         workings = working_shifts.objects.filter(shift_name=employs.working12,companyid=employs.companyid).first()

#         if workings:
#             starting_time = workings.starting_time
#             before_time = workings.befor_time  # Use the correct field name
#             starting_datetime = datetime.combine(todaysec, starting_time)
#             check_in_time = starting_datetime - timedelta(minutes=before_time)
#             check_in_cutoff = starting_datetime + timedelta(minutes=workings.cutoff_time)
#             can_check_in = check_in_time <= datetime.now() <= check_in_cutoff
#             current_month = datetime.today().date().replace(day=1)
#             next_month1 = (current_month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
#         else:
#             starting_time=None
#             before_time=None
#             starting_datetime=None
#             check_in_time=None
#             check_in_cutoff=None
#             can_check_in=None
#             current_month=None
#             next_month1=None
#         weekholiday=editholiday12.objects.first()
#         leave_report_datamedical = LeaveReportEmploy.objects.filter(Q(leave_date__range=[current_month, next_month1]) | Q(to_date__range=[current_month, next_month1]),employ_id=employs.id, leave_status=1,leave_type=2)
#         leave_report_datacausal = LeaveReportEmploy.objects.filter(Q(leave_date__range=[current_month, next_month1]) | Q(to_date__range=[current_month, next_month1]),employ_id=employs.id, leave_status=1,leave_type=1)
#         leave_report_dataearned = LeaveReportEmploy.objects.filter(Q(leave_date__range=[current_month, next_month1]) | Q(to_date__range=[current_month, next_month1]),employ_id=employs.id, leave_status=1,leave_type=3)
#         total_leave_duration = sum([leave.get_leave_duration() for leave in leave_report_datamedical])
#         total_causel=sum([leave.get_leave_duration() for leave in leave_report_datacausal])
#         total_earned=sum([leave.get_leave_duration() for leave in leave_report_dataearned])
#         formatted_data =  f"Medical Leave Total Duration (Days): {total_leave_duration}"
#         if employlevsheetmedical:
#             usedmedical=employlevsheetmedical.defaultleave-total_leave_duration
#         else:
#             usedmedical=None
#         if employlevsheetcausel:
#             usedcausel=employlevsheetcausel.defaultleave-total_causel
#         else:
#            usedcausel=None 
#         if employlevsheetearned:
#               usedearned=employlevsheetearned.defaultleave-total_earned
#         else:
#             usedearned=None

#         today = date.today()
#         is_special_weekend = customholidays.objects.filter(date=today,companyid=compid).exists()
#         is_public_holiday = publicholidays.objects.filter(publicholiday_date=today,companyid=compid).exists()
#         is_weekoff_today =today.strftime('%A') in weekoff_days or is_special_weekend or is_public_holiday
#         context['leave_report_data'] = formatted_data
#         context['calendar'] = mark_safe(html_cal)
#         context['prev_month'] = prev_month(d,self.request)
#         context['next_month'] = next_month(d,self.request)
#         context['employlevsheetcausel'] = employlevsheetcausel
#         context['employlevsheetmedical']=employlevsheetmedical
#         context['employlevsheetearned']=employlevsheetearned
#         context['weekholiday']=weekholiday
#         context['employs']=employs
#         context['formatted_data']=formatted_data
#         context['usedmedical']=usedmedical
#         context['usedcausel']=usedcausel
#         context['usedearned']=usedearned
#         context['weekoff_days']=weekoff_days
#         context['is_weekoff_today']=is_weekoff_today
#         context['is_public_holiday']=is_public_holiday
#         context['is_special_weekend']=is_special_weekend
     
#         context['tlop']=tlop
#         context['tloptions']=tloptions
#         context['data']=data

#         context['check_in']=check_in
#         context['check_out']=check_out
#         context['check_in_time']=check_in_time
#         context['check_in_cutoff']=check_in_cutoff
#         context['can_check_in']=can_check_in
#         context['organization_name']=organization_name



#         inline_context={
#             "days_list":"days_list",
            
            
#         }
#         inline_html_template=Template(f"{days_list}")
#         inline_view=inline_html_template.render(Context(inline_context))
#         context['inline_view']=inline_view
#         return  context
#     def post(self, request, *args, **kwargs):
#         form = HomeForm(request.POST)
#         if form.is_valid():
#             name = form.cleaned_data['name']
#             email = form.cleaned_data['email']
#             # Do something with the name and email
#             return HttpResponseRedirect(reverse('calendar2'))
#         else:
#             context = self.get_context_data(**kwargs)
#             context['form'] = form
           
#             context['check_in'] = check_in(self.request)
           
#             return redirect('/calendar2/')



class CalendarView_2(TemplateView):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')


    
    template_name = "admin-template/admin_calendar.html"
    

   
    def get_context_data(self, **kwargs):
        context = super(CalendarView_2,self).get_context_data(**kwargs)
        d = get_date(self.request.GET.get('month', None))
        cal = Calendar1(d.year, d.month,d.day)
        html_cal = cal.formatmonth(self.request,withyear=True)
        days_list,weekoff_days,weekend_count =caltable(self,d.year,d.month,self.request)
        admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')

        user = CustomUser.objects.filter(id=self.request.user.id).first()
        userid1 = user.id

        s = adminnav.objects.all()
        projectm = admin_project_create.objects.filter(admin_id=userid1)
        h = HR.objects.all()
        employs_all = Employs.objects.all()
        admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
        admin_home_drops=admin_home_drop.objects.filter(parent_category=None).order_by('id')
        data = AdminHod.objects.filter(id=self.request.user.id)
    


       
        
        context['calendar'] = mark_safe(html_cal)
        context['prev_month'] = prev_month(d,self.request)
        context['next_month'] = next_month(d,self.request)
        context['admin_drops'] = self.admin_drops
        context['weekoff_days']=weekoff_days
        context['admin_drops'] = admin_drops
        context['admin_home_drops'] = admin_home_drops
        context['s'] = s
        context['h'] = h
        context['data'] = data
        context['user'] = user
     
        context['projectm'] = projectm
        context['employs_all'] = employs_all


        inline_context={
            "days_list":"days_list"
        }
        inline_html_template=Template(f"{days_list}")
        inline_view=inline_html_template.render(Context(inline_context))
        context['inline_view']=inline_view
        return  context
    
    
    
    def post(self, request, *args, **kwargs):
        form = HomeForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            # Do something with the name and email
            return HttpResponseRedirect(reverse('calendar'))
        else:
            context = self.get_context_data(**kwargs)
            context['form'] = form

            context['check_in'] = check_in(self.request)
            return redirect('/calendar/')
# def reimbursement_approve_status1(request,leave_id):
#     admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
#     leave=Reimbursement.objects.get(id=leave_id)
#     leave.reimbursement_status=1
#     leave.save()
#     user=CustomUser.objects.all()
#     return HttpResponseRedirect(reverse("individual_reimbursement_view1", args=[leave.employ_id_id]))
        
def reimbursement_approve_status1(request,leave_id):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    leave=Reimbursement.objects.get(id=leave_id)
    leave.reimbursement_status=1
    leave.save()
    employsss=leave.employ_id
    emplo=Employs.objects.filter(id=employsss.id).first()
    compid=emplo.companyid.id
    orgname=Companys.objects.filter(id=compid).first()
    orgnam=orgname.organizationname

    admin_obj = request.user.employs.companyid.id
    companys_admin = Companys.objects.get(id=admin_obj)
    notify.send(sender=companys_admin.usernumber, recipient=employsss.admin, verb='Reimbursement Status Approve', description=f"{orgnam}: Your request for Reimbursement  has been approved" )

    return HttpResponseRedirect(reverse("individual_reimbursement_view1", args=[leave.employ_id_id]))
       

# def reimbursement_disapprove_status1(request,leave_id):
#     admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
#     leave=Reimbursement.objects.get(id=leave_id)
#     leave.reimbursement_status=2
#     leave.save()
#     return HttpResponseRedirect(reverse("individual_reimbursement_view1", args=[leave.employ_id_id]))


def reimbursement_disapprove_status1(request,leave_id):
    admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    leave=Reimbursement.objects.get(id=leave_id)
    leave.reimbursement_status=2
    leave.save()
    employsss=leave.employ_id
    emplo=Employs.objects.filter(id=employsss.id).first()
    compid=emplo.companyid.id
    orgname=Companys.objects.filter(id=compid).first()
    orgnam=orgname.organizationname

    admin_obj = request.user.employs.companyid.id
    companys_admin = Companys.objects.get(id=admin_obj)
    notify.send(sender=companys_admin.usernumber, recipient=employsss.admin, verb='Reimbursement Status Reject', description=f"{orgnam}: Your request for Reimbursement  has been rejected" )
    return HttpResponseRedirect(reverse("individual_reimbursement_view1", args=[leave.employ_id_id]))

def individual_employ_leaveReport(request,std):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)

    a=company_details.objects.filter(companyid=data.companyid.id).first()
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    request.session['employ_id']=std
    employ_id=Employs.objects.all()
    leaves=LeaveReportEmploy.objects.filter(employ_id=std)
    items_per_page = 10
    paginator = Paginator(leaves, items_per_page)
    employ_obj=Employs.objects.get(admin=request.user.id)
    page = request.GET.get('page')
    try:
        leaves = paginator.page(page)
    except PageNotAnInteger:
        leaves = paginator.page(1)
    except EmptyPage:
        leaves = paginator.page(paginator.num_pages)
    return render(request,"employ-template/individual_leave_view.html",{'employ_obj':employ_obj, "leaves":leaves,'organization_name':organization_name,'employ_id':employ_id,'data':data,})

# def approve_leave(request,leave_id):
#     admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
#     leave=LeaveReportEmploy.objects.get(id=leave_id)
#     leave.leave_status=1
#     leave.save()
#     return HttpResponseRedirect(reverse("individual_employ_leaveReport", args=[leave.employ_id_id]))

def approve_leave(request,leave_id):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    leave=LeaveReportEmploy.objects.get(id=leave_id)
    leave.leave_status=1
    leave.save()
    employsss=leave.employ_id
    emplo=Employs.objects.filter(id=employsss.id).first()
    compid=emplo.companyid.id
    orgname=Companys.objects.filter(id=compid).first()
    orgnam=orgname.organizationname
    admin_obj = request.user.employs.companyid.id
    companys_admin = Companys.objects.get(id=admin_obj)
    notify.send(sender=companys_admin.usernumber, recipient=employsss.admin, verb='Leave Status Accept', description=f"{orgnam}: Your  request for Leave has been approved" )
    return HttpResponseRedirect(reverse("individual_leaveReport", args=[leave.employ_id_id]))
    
    

from datetime import date
from .models import Companys

# def disapprove_leave(request,leave_id):
#     admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
#     leave=LeaveReportEmploy.objects.get(id=leave_id)
#     leave.leave_status=2
#     leave.save()
#     return HttpResponseRedirect(reverse("individual_employ_leaveReport", args=[leave.employ_id_id]))

def disapprove_leave(request,leave_id):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    leave=LeaveReportEmploy.objects.get(id=leave_id)
    leave.leave_status=2
    leave.save()
    employsss=leave.employ_id
    emplo=Employs.objects.filter(id=employsss.id).first()
    compid=emplo.companyid.id
    orgname=Companys.objects.filter(id=compid).first()
    orgnam=orgname.organizationname
    admin_obj = request.user.employs.companyid.id
    companys_admin = Companys.objects.get(id=admin_obj)
    notify.send(sender=companys_admin.usernumber, recipient=employsss.admin, verb='Leave Status Reject', description=f"{orgnam}: Your  request for Leave has been disapproved" )
    return HttpResponseRedirect(reverse("individual_leaveReport", args=[leave.employ_id_id]))
    

def data_tables1(request):

    employ_obj = Employs.objects.get(admin=request.user.id)
    try:
        company = Companys.objects.get(id=employ_obj.companyid.id)
        organization_name = company.organizationname
    except Companys.DoesNotExist:
        organization_name = "No Organization Name Available"

    employee_projects = TeamMember.objects.filter(employee=employ_obj).values_list('project', flat=True)
    today = date.today()

    team_lead = TeamMember.objects.filter(employee=employ_obj, is_team_lead=True).exists()
    project_manager = employ_obj.projectmanagerop == 1
    hr = employ_obj.hroptions == 1

    if team_lead:
        tsk1_data = tlassigntask.objects.filter(
            task_date=today,
            companyid=employ_obj.companyid,
            project_id__in=employee_projects
        )
    elif project_manager or hr:
        tsk1_data = tlassigntask.objects.filter(
            task_date=today,
            companyid=employ_obj.companyid
        )
    else:
        tsk1_data = tlassigntask.objects.none()

    team_leads = TeamMember.objects.filter(is_team_lead=True, employee__companyid=employ_obj.companyid)

    items_per_page = 10
    paginator = Paginator(tsk1_data, items_per_page)
    page = request.GET.get('page')

    try:
        tsk1_data = paginator.page(page)
    except PageNotAnInteger:
        tsk1_data = paginator.page(1)
    except EmptyPage:
        tsk1_data = paginator.page(paginator.num_pages)

    return render(request, 'employ-template/emptables.html', {
        "data": employ_obj,
        "tsk1_data": tsk1_data,
        "team_leads": team_leads,
        "organization_name": organization_name
    })

    

def individual_advancesalary_Report1(request,std):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop

    data=Employs.objects.filter(admin=request.user.id).first()
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)

    comp_id=data.companyid
    a=company_details.objects.filter(companyid=comp_id).first()
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)

 
    request.session['student_id']=std
    student_id=Employs.objects.all()
    leaves=ad_salary.objects.filter(student_id=std)
    items_per_page = 10
    paginator = Paginator(leaves, items_per_page)

    page = request.GET.get('page')
    try:
        leaves = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        leaves = paginator.page(1)
    except EmptyPage:
        # If the page is out of range (e.g. 9999), deliver the last page of results.
        leaves = paginator.page(paginator.num_pages)
    return render(request,"employ-template/individual_employ_advancesalary_Report.html",{'organization_name':organization_name,"leaves":leaves,'student_id':student_id,'data':data})


def advance_salary_status_approve1(request, leave_id):
    leave = ad_salary.objects.get(id=leave_id)
    leave.request_status = 1
    leave.save()
    return HttpResponseRedirect(reverse("individual_advancesalary_Report1", args=[str(leave.student_id)]))

def advance_salary_status_disapprove1(request, leave_id):
    leave = ad_salary.objects.get(id=leave_id)
    leave.request_status = 2
    leave.save()
    return HttpResponseRedirect(reverse("individual_advancesalary_Report1", args=[str(leave.student_id)]))


def employee_form(request):
    li=list.objects.all()
    if request.method == "POST":
        first_name = request.POST["first_name"]
        last_name = request.POST["last_name"]
        username = request.POST["username"]
        email = request.POST["email"]
        web_mail = request.POST["web_mail"]
        password = request.POST["password"]
        address = request.POST["address"]
        empid = request.POST["empid"]
        Manager = request.POST["Manager"]
        designation = request.POST["designation"]
        location = request.POST["location"]
        package = request.POST["package"]
        pincode = request.POST["pincode"]
        contactno = request.POST["contactno"]
        bloodgroup = request.POST["bloodgroup"]
        dateofjoining = request.POST['dateofjoining']
        sex = request.POST["gender"]
        role = request.POST.get('role') # Added field for role
        
        user = CustomUser.objects.create_user(username=username, password=password, email=email, user_type=2)
        user.employs.first_name = first_name
        user.employs.last_name = last_name
        user.employs.email = email
        user.employs.password = password
        user.employs.address = address
        user.employs.empid = empid
        user.employs.web_mail = web_mail
        user.employs.Manager = Manager
        user.employs.designation = designation
        user.employs.location = location
        user.employs.package = package
        user.employs.pincode = pincode
        user.employs.contactno = contactno
        user.employs.bloodgroup = bloodgroup
        user.employs.dateofjoining = dateofjoining
        user.employs.role = role  # Set the role field for HR

        user.save()
                
       
        html_content = render_to_string("email_template.html", {'title': 'test email', 'first_name': first_name,
                                                                'last_name': last_name, 'empid': empid})
        text_content = strip_tags(html_content)
        subject = "WELCOME TO DEVELOPTRESS"
        email = EmailMultiAlternatives(
            subject,
            text_content,
            settings.EMAIL_HOST_USER,
            [email],
        )
        email.attach_alternative(html_content, "text/html")
        email.fail_silently = True
        email.send()
        
        if user:
            messages.success(request, "Successfully Added employee")
            return redirect("/employee_form")
        else:
            messages.error(request,"Failed to add employ")
    # If the request method is not POST, render the form
    s = adminnav.objects.all()
    form = Employs.objects.all()
    hrform = HR.objects.all()
    role = list.objects.all()
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    return render(request, "employ-template/employ_form_template.html",
                  {'role': role, 'hrform': hrform, "form": form, 's': s, 'li':li})

import pandas as pd

def employdetails_upload_excel(request,message=None, error_message=None):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    company_id=data.companyid
    a=company_details.objects.filter(companyid=company_id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)



    if request.method == 'POST':
        excel_file = request.FILES['excel_file']

        if excel_file.name.endswith('.xlsx'):
            try:
                df = pd.read_excel(excel_file)

                # Assuming the columns in the Excel file match the fields in Employs model
                for index, row in df.iterrows():
                    user=CustomUser.objects.create_user(username=row['username'],password=row['empid'],first_name=row['first_name'],last_name=row['last_name'],email=row['email'],user_type=2)
                    user.employs.first_name=row['first_name']
                    user.employs.last_name=row['last_name']
                    user.employs.email=row['email']
                    user.employs.address=row['address']
                    user.employs.empid=row['empid']
                    user.employs.role=row['role']
                    user.employs.location=row['location']
                    user.employs.package=row['package']
                    user.employs.pincode=row['pincode']
                    user.employs.contactno=row['contactno']
                    user.employs.gender=row['sex']
                    date_string = str(row['dateofjoining'])
                    date_obj = datetime.strptime(date_string, '%Y-%m-%d %H:%M:%S')  # Adjust the format as per your Excel file
                    user.employs.dateofjoining = date_obj.date()
                    user.save()
                    html_content = render_to_string("email_template.html",{'title':'test email','first_name':row['first_name'],'last_name':row['last_name'],'empid':row['empid']})
                    text_content=strip_tags(html_content)
                    subject="WELCOME TO DEVELOPTRESS"
                    email = EmailMultiAlternatives(
                            subject,
                            text_content,
                            settings.EMAIL_HOST_USER,
                            [row['email']],
                            )
                    email.attach_alternative(html_content,"text/html")
                    email.fail_silently = True
                    email.send()
                    
                message = f"{len(df)} records imported successfully."  
                return render(request, 'employ-template/employdetails_upload_bulk_data.html',{'message':message})
            except Exception as e:
                # Handle any exceptions that may occur during processing
                # return HttpResponseRedirect(reverse("upload_excel"))
                error_message = f"{str(e)} An error occurred during processing."
        else:
            # return HttpResponseRedirect(reverse("upload_excel"),{'messagea': 'Please upload a valid Excel file.'}) 
            error_message = 'Please upload a valid Excel file.'
            
    return render(request, 'employ-template/employdetails_upload_bulk_data.html',{'organization_name':organization_name, 'error_message':error_message,'data':data,})

# def edit_role(request):
#     rl=Employs.objects.all()
def edit_role_permission1(request,id):
    admin_drops=admin_drop.objects.filter(parent_category=None).order_by('id')
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    compid=data.companyid
    if request.method == "POST":
        designation=request.POST.get('designation')
        k=Employs.objects.get(id=id)
        k.designation=designation
        if designation == "HR": 
            k.hroptions=1
        else:
             k.hroptions=0
        if designation == "Manager":
             k.projectmanagerop=1
        else:
            k.projectmanagerop=0
        k.save()
        return redirect('/employ_people_count')

    return render(request, "employ-template/rolepermission.html",{"data":data})

def edit_edit_role_permission1(request,std):
    request.session['employ_id']=std
    k=Employs.objects.get(admin=std)
    objss=k.id 
    users=CustomUser.objects.get(id=std)
    datas=employ_add_form.objects.filter(student_id=objss)
    return redirect('/update_role_permission1')
    return render(request,"employ-template/rolepermission.html",{'k':k,'objss':objss,'users':users,'datas':datas})

def update_role_permission1(request,std):
    request.session['employ_id']=std
    k=Employs.objects.get(admin=std)
    objss=k.id 
    users=CustomUser.objects.get(id=std)
    datas=employ_add_form.objects.filter(student_id=objss)
    if request.method == "POST":
        designation = request.POST.get('designation')
        k.designation=designation
        k.save()
        
        return redirect('/edit_people/')

    return render(request, 'admin-template/rolepermission.html',{'k':k,'objss':objss,'users':users,'std':std,'datas':datas})
# from .models import halfldayvreason    
# def employ_apply_leave(request):
#     staff_obj = Employs.objects.get(admin=request.user.id)
#     leave_data=LeaveReportEmploy.objects.filter(employ_id=staff_obj)
#     half=halfldayvreason.objects.first()
#     return render(request,"employ-template/employ_apply_leave.html",{"leave_data":leave_data,'half':half})

# def employ_apply_leave_save(request):
#     if request.method!="POST":
#         return HttpResponseRedirect(reverse("employ_apply_leave"))
#     else:
#         leave_date=request.POST.get("leave_date")
#         leave_type=request.POST.get("leave_type")
#         to_date=request.POST.get("to_date")
#         leave_msg=request.POST.get("leave_msg")
#         email_leave=request.POST.get("email_leave")

#         employ_obj=Employs.objects.get(admin=request.user.id)
#         admin_obj = request.user.employs.companyid.id
#         employsss = Employs.objects.get(id=employ_obj.id)
#         companys_admin = Companys.objects.get(id=admin_obj)
#         try:
#             leave_report=LeaveReportEmploy(email_leave=email_leave,employ_id=employ_obj,leave_type=leave_type,leave_date=leave_date,to_date=to_date,leave_message=leave_msg,leave_status=0)
#             leave_report.save()
#             notify.send(sender=employsss.admin, recipient=companys_admin.usernumber, verb='Leave', description=f"{employsss.first_name}.{employsss.last_name} has applied for Leave" )
#             messages.success(request, "Successfully Applied for Leave")
#             return HttpResponseRedirect(reverse("employ_apply_leave"))
#         except:
#             messages.error(request, "Failed To Apply for Leave")
#             return HttpResponseRedirect(reverse("employ_apply_leave"))

# def employ_apply_leave_save(request):
#     if request.method != "POST":
#         return HttpResponseRedirect(reverse("employ_apply_leave"))
#     else:
#         leave_date = request.POST.get("leave_date")
#         leave_type = request.POST.get("leave_type")
#         to_date = request.POST.get("to_date")
#         leave_msg = request.POST.get("leave_msg")
#         email_leave = request.POST.get("email_leave")

#         employ_obj = Employs.objects.get(admin=request.user.id)
#         admin_obj = request.user.employs.companyid.id
#         employsss = Employs.objects.get(id=employ_obj.id)
#         companys_admin = Companys.objects.get(id=admin_obj)
#         employss = Employs.objects.filter(hroptions=1, companyid=companys_admin)
#         # employss1 = Employs.objects.filter(admin=request.user.id,hroptions=1,companyid=companys_admin)

#         try:
#             leave_report = LeaveReportEmploy(email_leave=email_leave, employ_id=employ_obj, leave_type=leave_type, leave_date=leave_date, to_date=to_date, leave_message=leave_msg, leave_status=0)
#             leave_report.save()

#             # for admin in employss1:
#             #     notify.send(sender=employsss.admin, recipient=admin.admin, verb='Leave', description=f"{employsss.first_name}.{employsss.last_name} has applied for Leave")
            
#             for admin in employss:
#                 notify.send(sender=employsss.admin, recipient=admin.admin, verb='Leave', description=f"{employsss.first_name}.{employsss.last_name} has applied for Leave")
#             notify.send(sender=employsss.admin, recipient=companys_admin.usernumber, verb='Leave', description=f"{employsss.first_name}.{employsss.last_name} has applied for Leave")

#             messages.success(request, "Leave Applied Successfully")
#             return HttpResponseRedirect(reverse("employ_apply_leave"))
#         except:
#             messages.error(request, "Failed To Apply for Leave")
#             return HttpResponseRedirect(reverse("employ_apply_leave"))


# def employ_apply_leave_save(request):
#     if request.method != "POST":
#         return HttpResponseRedirect(reverse("employ_apply_leave"))
    
#     leave_date = request.POST.get("leave_date")
#     leave_type = request.POST.get("leave_type")
#     leave_msg = request.POST.get("leave_msg")
#     email_leave = request.POST.get("email_leave")

#     # Set default values for to_date
#     to_date = None

#     # Check if leave type is Half-Day or Unpaid_halfday
#     if leave_type in ["Half-Day", "Unpaid_halfday"]:
#         # For half-day or unpaid half-day leave, set to_date to the same as leave_date
#         to_date = leave_date

#     employ_obj = Employs.objects.get(admin=request.user.id)
#     admin_obj = request.user.employs.companyid.id
#     employsss = Employs.objects.get(id=employ_obj.id)
#     companys_admin = Companys.objects.get(id=admin_obj)
#     employss = Employs.objects.filter(hroptions=1, companyid=companys_admin)

#     try:
#         # Validate the form data
#         leave_report = LeaveReportEmploy(
#             email_leave=email_leave,
#             employ_id=employ_obj,
#             leave_type=leave_type,
#             leave_date=leave_date,
#             to_date=to_date,
#             leave_message=leave_msg,
#             leave_status=0
#         )
#         leave_report.full_clean()  # Validate the model fields
#         leave_report.save()

#         # Notify admins about the leave application
#         for admin in employss:
#             notify.send(sender=employsss.admin, recipient=admin.admin, verb='Leave', description=f"{employsss.first_name}.{employsss.last_name} has applied for Leave")
#         notify.send(sender=employsss.admin, recipient=companys_admin.usernumber, verb='Leave', description=f"{employsss.first_name}.{employsss.last_name} has applied for Leave")

#         messages.success(request, "Leave Applied Successfully")
#         return HttpResponseRedirect(reverse("employ_apply_leave"))
#     except ValidationError as e:
#         # Handle validation errors
#         error_message = ", ".join(e.messages)
#         messages.error(request, f"Failed To Apply for Leave: {error_message}")
#         return HttpResponseRedirect(reverse("employ_apply_leave"))
#     except Exception as e:
#         # Handle other exceptions
#         print(e)  # Print the error for debugging purposes
#         messages.error(request, "Failed To Apply for Leave")
#         return HttpResponseRedirect(reverse("employ_apply_leave"))





from django.core.exceptions import ValidationError

from django.core.exceptions import ValidationError
from django.http import HttpResponseRedirect
from django.urls import reverse
from .models import LeaveReportEmploy

def employ_apply_leave_save(request):
    if request.method != "POST":
        return HttpResponseRedirect(reverse("employ_apply_leave"))
    
    leave_date = request.POST.get("leave_date")
    leave_type = request.POST.get("leave_type")
    leave_msg = request.POST.get("leave_msg")
    email_leave = request.POST.get("email_leave")
    to_date = request.POST.get("to_date")

    if leave_type == "Half-Day" or leave_type == "Unpaid_halfday":
        # For Half-Day or Unpaid Half-Day, set to_date same as leave_date
        to_date = leave_date

    employ_obj = Employs.objects.get(admin=request.user.id)
    admin_obj = request.user.employs.companyid.id
    employsss = Employs.objects.get(id=employ_obj.id)
    companys_admin = Companys.objects.get(id=admin_obj)
    employss = Employs.objects.filter(hroptions=1, companyid=companys_admin)

    try:
        # Validate the form data
        leave_report = LeaveReportEmploy(
            email_leave=email_leave,
            employ_id=employ_obj,
            leave_type=leave_type,
            leave_date=leave_date,
            to_date=to_date,  # Set to_date based on leave_date for "Half-Day" and "Unpaid Half-Day"
            leave_message=leave_msg,
            leave_status=0
        )
        leave_report.full_clean()  # Validate the model fields
        leave_report.save()

        # Notify admins about the leave application
        for admin in employss:
            notify.send(sender=employsss.admin, recipient=admin.admin, verb='Leave', description=f"{employsss.first_name}.{employsss.last_name} has applied for Leave")
        notify.send(sender=employsss.admin, recipient=companys_admin.usernumber, verb='Leave', description=f"{employsss.first_name}.{employsss.last_name} has applied for Leave")

        messages.success(request, "Leave Applied Successfully")
        return HttpResponseRedirect(reverse("employ_apply_leave"))
    except ValidationError as e:
        # Handle validation errors
        error_message = ", ".join(e.messages)
        messages.error(request, f"Failed To Apply for Leave: {error_message}")
        return HttpResponseRedirect(reverse("employ_apply_leave"))
    except Exception as e:
        # Handle other exceptions
        print(e)  # Print the error for debugging purposes
        messages.error(request, "Failed To Apply for Leave")
        return HttpResponseRedirect(reverse("employ_apply_leave"))




def reimbursement_apply_view(request):

    


    k=types.objects.all()

    staff_obj = Employs.objects.get(admin=request.user.id)
    leave_data=Reimbursement.objects.filter(employ_id=staff_obj)
  
    
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    compid=data.companyid
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    k3=reimbursementsetup1.objects.filter(companyid=compid)
  
    md=reimbursementsetup.objects.filter(companyid=compid)




    return render(request,"employ-template/reimbursement_apply_view.html",{'organization_name':organization_name,"data":data,"data":data,"leave_data":leave_data,'k3':k3,'k':k,'md':md, 

})



# def reimbursement_apply_view_save(request):
#     if request.method =="POST":
#         typea=request.POST.get("typea")
#         date=request.POST.get("date")
#         detail=request.POST.get("detail")
#         amount=request.POST.get("amount")
#         image=request.FILES.get("image")
#         student_obj=request.user.employs.id
#         admin_obj = request.user.employs.companyid.id
#         employsss = Employs.objects.get(id=student_obj)
#         companys_admin = Companys.objects.get(id=admin_obj)
#         try:
#             leave_report=Reimbursement(employ_id_id=student_obj,typea=typea,date=date,detail=detail,amount=amount,image=image,reimbursement_status=0,companyid_id=admin_obj)
#             leave_report.save()
#             notify.send(sender=employsss.admin, recipient=companys_admin.usernumber, verb='Reimbursement', description=f"{employsss.first_name}.{employsss.last_name} has applied for reimbursement of {amount}" )
#             messages.success(request, "Successfully Applied for Reimbursement")
#             return HttpResponseRedirect(reverse("reimbursement_apply_view"))
#         except:
#             messages.error(request, "Failed To Apply for Reimbursement")
#             return HttpResponseRedirect(reverse("reimbursement_apply_view"))
#     return HttpResponseRedirect(reverse("reimbursement_apply_view"))

def reimbursement_apply_view_save(request):
    if request.method =="POST":
        typea=request.POST.get("typea")
        date=request.POST.get("date")
        detail=request.POST.get("detail")
        amount=request.POST.get("amount")
        image=request.FILES.get("image")
        student_obj=request.user.employs.id
        admin_obj = request.user.employs.companyid.id
        employsss = Employs.objects.get(id=student_obj)
        companys_admin = Companys.objects.get(id=admin_obj)
        employss=Employs.objects.filter(hroptions=1,companyid=companys_admin)
        # employss1 = Employs.objects.filter(admin=request.user.id,hroptions=1,companyid=companys_admin)
        try:
            leave_report=Reimbursement(employ_id_id=student_obj,typea=typea,date=date,detail=detail,amount=amount,image=image,reimbursement_status=0,companyid_id=admin_obj)
            leave_report.save()
            # for admin in employss1:
            #     notify.send(sender=employsss.admin, recipient=admin.admin, verb='Reimbursement', description=f"{employsss.first_name}.{employsss.last_name} has applied for reimbursement of {amount}" )
            for admin in employss:
                notify.send(sender=employsss.admin, recipient=admin.admin, verb='Reimbursement', description=f"{employsss.first_name}.{employsss.last_name} has applied for reimbursement of {amount}" )
            notify.send(sender=employsss.admin, recipient=companys_admin.usernumber, verb='Reimbursement', description=f"{employsss.first_name}.{employsss.last_name} has applied for reimbursement of {amount}" )
            messages.success(request, "Successfully applied for Reimbursement")
            return HttpResponseRedirect(reverse("reimbursement_apply_view"))
        except:
            messages.error(request, "Failed To Apply for Reimbursement")
            return HttpResponseRedirect(reverse("reimbursement_apply_view"))
    return HttpResponseRedirect(reverse("reimbursement_apply_view"))



        
from django.db.models import Sum
from .models import Employs, Reimbursement, employnav, types


def reg(request):
    

    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    compid=data.companyid
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    # if data2:
    #    s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    # elif project_manager:
    #     s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    
    # elif teamleadop:
    #     s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    # else:
    #     s=employnav.objects.filter(is_name_exist=1,employ_options=1)
    a=company_details.objects.filter(companyid=data.companyid.id).first()

   
    staff_obj = Employs.objects.get(admin=request.user.id)
 
    k = reimbursementsetup1.objects.filter(companyid=compid)






    st4 = request.POST.get("ss")
    st3 = request.POST.get("vk")
    st = request.POST.get("d1")
    st1 = request.POST.get("d2")

    leave_data = Reimbursement.objects.filter(employ_id=staff_obj).order_by("-date")

    # Apply filters based on user input
    if st4 and st4 != '----Select----':  # Check if a valid status is selected
        leave_data = leave_data.filter(reimbursement_status=st4)
    if st3:  # If "Select Type" is selected
        leave_data = leave_data.filter(typea__icontains=st3)
    if st and st1:
        leave_data = leave_data.filter(date__range=[st, st1])

    total_approved = leave_data.filter(reimbursement_status=1).aggregate(Sum('amount'))['amount__sum'] or 0
    total_pending = leave_data.filter(reimbursement_status=0).aggregate(Sum('amount'))['amount__sum'] or 0
    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator(leave_data, items_per_page)
    page = request.GET.get('page')

    try:
        leave_data = paginator.page(page)
    except PageNotAnInteger:
        leave_data = paginator.page(1)
    except EmptyPage:
        leave_data = paginator.page(paginator.num_pages)


    return render(request, "employ-template/a.html", {
        'leave_data': leave_data,
        # 's': s,
        'k': k,
        'total': total_approved,
        'total1': total_pending,
        'data':data,

       

        
        'organization_name':organization_name,
        
                })







    

def type(request):
    staff_obj = Employs.objects.get(admin=request.user.id)
    leave_data=Reimbursement.objects.filter(employ_id=staff_obj) 
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0).all()
    k=types.objects.all()
    

    if request.method=="POST":
        
        st3=request.POST.get("vk")
      
        if st3!=None:
            leave_data=Reimbursement.objects.filter(typea_icontains=st3).values() 

    return render(request,"employ-template/a.html",{'leave_data':leave_data,'k':k})


def delete(request,id):
    k=Reimbursement.objects.get(id=id)
    k.delete()
    return redirect("/reg")
def edit(request,id):
    
    employ_obj=Employs.objects.get(admin=request.user.id)
    try:
        company = Companys.objects.get(id=employ_obj.companyid.id)  
        organization_name = company.organizationname 
    except Companys.DoesNotExist:
        organization_name = "No Organization Name Available" 



   
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)


 
    compid=data.companyid

    ks=Reimbursement.objects.get(id=id)
    k3=reimbursementsetup1.objects.filter(companyid=compid)
    k=types.objects.all()
    return render(request,"employ-template/updateform.html",{'data':data, 'ks':ks,'k':k, 'k3':k3, 'organization_name':organization_name})
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

def update_reim(request,id):
    if request.method=="POST":
        date=request.POST["date"]
        detail=request.POST["detail"]
        typea=request.POST["typea"]
        amount=request.POST["amount"]
        image=request.FILES.get('image')
        k=Reimbursement.objects.get(id=id);
        if date:
           k.date=date
        k.detail=detail
        k.typea=typea
        k.amount=amount
        if image:
           k.image=image
        k.save();
        return redirect("/reg")
    return render(request,"employ-template/updateform.html")

def documents_uploaded_view(request):
    staff_obj = Employs.objects.get(admin=request.user.id)
    leave_data=empdocs.objects.filter(employ_id=staff_obj)
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)   
    r=documents_setup.objects.all()
    return render(request,"employ-template/documents_uploaded_view.html",{"leave_data":leave_data,'r':r})
from .models import Companys

def trails(request):
    

    user = CustomUser.objects.filter(id=request.user.id).first()

    employ_obj=Employs.objects.get(admin=request.user.id)
    try:
        company = Companys.objects.get(id=employ_obj.companyid.id)  
        organization_name = company.organizationname 
    except Companys.DoesNotExist:
        organization_name = "No Organization Name Available" 


    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
 
 

    staff_obj = Employs.objects.get(admin=request.user.id)
    leave_data=empdocs.objects.filter(employ_id=staff_obj)
    stres1=empdocs.objects.filter(employ_id=staff_obj).values_list('documenttype1')
    # r=documents_setup1.objects.filter(companyid=compid).exclude(document_type__in=stres1)
    remaining_document_types = documents_setup1.objects.filter(companyid=compid)

    r=documents_setup1.objects.filter(companyid=compid)  
    
   
    items_per_page = 5  # Adjust the number of items per page as needed

    paginator = Paginator(leave_data, items_per_page)
    page = request.GET.get('page')

    try:
        leave_data = paginator.page(page)
    except PageNotAnInteger:
        leave_data = paginator.page(1)
    except EmptyPage:
        leave_data = paginator.page(paginator.num_pages)  
    # if request.method=="POST":
    #     documenttype1=request.POST["documenttype1"]
    #     imagefile=request.FILES.get('imagefile')
    #     description=request.POST["description"]
    #     employ_obj=Employs.objects.get(admin=request.user.id)
    #     k=empdocs(employ_id=employ_obj,documenttype1=documenttype1,imagefile=imagefile,description=description,company=compid)
    #     k.save()
    #     messages.success(request,"Your Details Were Successfully Submitted")
    #     # return redirect("/trails")
    #     return redirect("/Employ_home")
    # return render(request,"employ-template/documents_uploaded_view.html",{'organization_name':organization_name,"data":data,'data':data,'leave_data':leave_data,'r':r})

    if request.method == "POST":
        documenttype1 = request.POST["documenttype1"]
        imagefiles = request.FILES.getlist('imagefile')
        description = request.POST["description"]
        employ_obj = Employs.objects.get(admin=request.user.id)
            
        for imagefile in imagefiles:
            k = empdocs(employ_id=employ_obj, documenttype1=documenttype1, imagefile=imagefile, description=description, company=compid)
            k.save()
            
        messages.success(request, "Your Details Were Successfully Submitted")
        return redirect("/trails")
    uploaded_document_types = empdocs.objects.filter(employ_id__admin=request.user.id).values_list('documenttype1', flat=True)
    remaining_document_types = remaining_document_types.exclude(document_type__in=uploaded_document_types)

    return render(request,"employ-template/documents_uploaded_view.html",{'organization_name':organization_name,"data":data,'data':data,'leave_data':leave_data,'r': remaining_document_types,
})



from datetime import datetime, time
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
def task_data(request):
    employ_obj = Employs.objects.get(admin=request.user.id)

    try:
        company = Companys.objects.get(id=employ_obj.companyid.id)
        organization_name = company.organizationname
    except Companys.DoesNotExist:
        organization_name = "No Organization Name Available"

    projects_drops = project_drop.objects.filter(parent_category=None).order_by('id')

    data = Employs.objects.filter(admin=request.user.id).first()
    data1 = data.id
    data2 = data.hroptions
    project_manager = data.projectmanagerop
    teamleadop = TeamMember.objects.filter(employee=data1, is_team_lead=1)
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    # if data2:
    #     s = employnav.objects.filter(is_name_exist=1, hr_options=1)
    # elif project_manager:
    #     s = employnav.objects.filter(is_name_exist=1, projectmanager_options=1)
    # elif teamleadop:
    #     s = employnav.objects.filter(is_name_exist=1, is_tl_option=1)
    # else:
    #     s = employnav.objects.filter(is_name_exist=1, employ_options=1)
    a = company_details.objects.filter(companyid=data.companyid).first()
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    if data2:  
        task_ids = employperformance.objects.values_list('project_task1', flat=True)
        tasks = projecttask.objects.filter(id__in=task_ids)
    elif project_manager:
        task_ids = employperformance.objects.values_list('project_task1', flat=True)
        tasks = projecttask.objects.filter(id__in=task_ids)
    else:  
        task_ids = employperformance.objects.filter(employ_name=employ_obj,status='4').values_list('project_task1', flat=True)
        tasks = tlassigntask.objects.filter(employid=data1).order_by('-task_date')
    tasks_with_status = []
    tsk1_data1 = employperformance.objects.filter(tl_name_id=data1).values('date').annotate(avg_performance=Avg('performance'))
    tsk1_data = employperformance.objects.filter(tl_name_id=data1)

    for task in tasks:
        performance_data = employperformance.objects.filter(project_task1=task,status='2').exists()
       
        deadline = task.deadline
        tasks_with_status.append({'task': task, 'deadline':deadline})

    items_per_page = 10
    paginator = Paginator(tasks_with_status, items_per_page)
    page = request.GET.get('page', 1)

    try:
        tasks_with_status = paginator.page(page)
    except PageNotAnInteger:
        tasks_with_status = paginator.page(1)
    except EmptyPage:
        tasks_with_status = paginator.page(paginator.num_pages)

    return render(request, "employ-template/task.html", {
        'tasks_with_status': tasks_with_status,
        'employ_obj': employ_obj,
        # 's': s,
        'data': data,
        'data1': data1,
        'data2': data2,
        'tasks':tasks,
        'tsk1_data':tsk1_data,
        'tsk1_data1':tsk1_data1,
        'organization_name': organization_name,
        
    })
from django.shortcuts import render, redirect
from .models import tlassigntask             


def task_description_view(request, task_id):
    task = get_object_or_404(tlassigntask, pk=task_id)
    
    if request.method == 'POST':
        if 'mark_started' in request.POST:
            print("POST request received with mark_started button.")  
            print(f"Current task status before any change: {task.status}")  
            
            if task.status == 4:  # Check if the task is already completed
                messages.info(request, "Completed tasks cannot be marked as started again.")
                print("Task is already completed and cannot be started again.")  
            elif task.status == 1:  # Check if the task is already started
                messages.info(request, "Task is already started.")
                print("Task is already started.")  
            else:
                task.status = 1  # Update status to 'Started'
                task.save()
                messages.success(request, "Task status updated to 'Started'.")
                print("Task status updated to 'Started'.")  
            
            # Redirect to the task list page
            return redirect('task_data')  # Ensure 'task_data' is the correct URL name for your task list page
    
    # Render task description for GET request
    task_descriptions = tlassigntask.objects.filter(pk=task_id).values_list('description', flat=True)
    return render(request, 'employ-template/task_description.html', {'task_descriptions': task_descriptions, 'task': task})
    
    

def upload_photo(request):
    user_obj = Employs.objects.get(admin=request.user.id)
    
    try:
            data = employ_add_form.objects.get(student_id=user_obj)
    except employ_add_form.DoesNotExist:
        return redirect('employ_add')

    

    if request.method == "POST":
        profile_pic = request.FILES.get("profile_pic")
        if profile_pic:
           
            data.profile_pic = profile_pic
            
            data,_ = employ_add_form.objects.get_or_create(student_id=user_obj)
            data.profile_pic=profile_pic

            data.save()
            messages.success(request, "Profile picture uploaded successfully.")
        else:
            messages.error(request, "Profile picture upload failed.")

    return render(request, "employ-template/employprofilepic.html", {'data': data})

def delete_profile_pic(request):
    user_obj = Employs.objects.get(admin=request.user.id)
    try:
        data = employ_add_form.objects.get(student_id=user_obj.id)
        data.profile_pic.delete()  # Delete the profile picture file
        data.profile_pic = None  # Clear the profilepic field
        data.save()
        messages.success(request, "Profile picture removed successfully.")
    except employ_add_form.DoesNotExist:
        messages.error(request, "User data not found. Please add user data first.")
    
    return redirect("upload_photo")


from .models import task1
def employ_profile(request):
    
    employ_obj=Employs.objects.get(admin=request.user.id)
    try:
        company = Companys.objects.get(id=employ_obj.companyid.id)  
        organization_name = company.organizationname 
    except Companys.DoesNotExist:
        organization_name = "No Organization Name Available" 

    




    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id if data else None
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)





    compid=data.companyid

    half=task1.objects.first()
    user=CustomUser.objects.get(id=request.user.id)

    

    emp_detail=Employs.objects.get(admin=request.user.id) 
    



    employ=Employs.objects.get(admin=user)
    profile=Employs.objects.all()
    email2=request.session.get('email_1');
    q=employ_add_form.objects.filter(email2=email2).values();
    datas=employ_add_form.objects.filter(student_id=employ)
    return render(request,"employ-template/employ_profile.html",{'emp_detail':emp_detail,'data':data,"profile":profile,"half":half,"datas":datas,"user":user,"employ":employ,"q":q,'data':data,'organization_name':organization_name,
             

})



# Check OpenCV version



# screenshot_project/screenshot_app/views.py

from django.shortcuts import render
from django.views import View
import time
import os
if 'DISPLAY' in os.environ:
    import pyautogui as pg
from django.http import  JsonResponse
from .models import Screenshots , SystemStatus
import threading
import subprocess
import json
from django.views.decorators.csrf import csrf_exempt

class ScreenshotView(View):
    template_name = 'employ-template/calendar.html'
    is_capturing = False
    capture_thread = None

    def get(self, request, *args, **kwargs):
        screenshotform = ScreenshotsForm()
        return render(request, self.template_name, {'screenshotform': screenshotform})

    def post(self, request, *args, **kwargs):
        screenshotform = ScreenshotsForm(request.POST, request.FILES)
        if screenshotform.is_valid():
            employee_id =request.user.employs.id
            company_id = request.user.employs.companyid.id
            screenshot = screenshotform.save(commit=False)
            screenshot.employee_id = employee_id  # Assign the employee field
            screenshot.company_id = company_id
            print(f"Employee ID: {screenshot.employee_id}")
            screenshot.image = self.capture_screenshot(employee_id, company_id)
            screenshot.save()                
            return JsonResponse({'message': 'Image inserted in the database'})
        else:
            print(screenshotform.errors)
            

        return JsonResponse({'error': 'Form is not valid'})

    def capture_screenshot(self, employee_id, company_id):
        random = int(time.time())
        filename = f"screenshot_{random}.png"
        relative_path = os.path.join("screenshots", filename)
        full_path = os.path.join(settings.MEDIA_ROOT, relative_path)
        directory = os.path.dirname(full_path)
        os.makedirs(directory, exist_ok=True)
        
        # Create the directory if it doesn't exist
        screenshot = Screenshots(image=relative_path , employee_id=employee_id, company_id=company_id )
        screenshot.save()
        pg.screenshot(full_path)
        return full_path

    @classmethod
    def start_capture(cls, request):
        cls.is_capturing = True

        # Check if the capture thread is not already running
        if not cls.capture_thread or not cls.capture_thread.is_alive():
            # Create a thread for capturing screenshots at regular intervals
            employee_id = request.user.employs.id
            company_id = request.user.employs.companyid.id
            cls.capture_thread = threading.Thread(target=cls.capture_screenshots_thread, args=(employee_id,company_id))
            cls.capture_thread.start()

        return JsonResponse({'message': 'Capture started'})

    @classmethod
    def stop_capture(cls, request):
        cls.is_capturing = False

        # Wait for the capture thread to finish
        if cls.capture_thread and cls.capture_thread.is_alive():
            cls.capture_thread.join()

        return JsonResponse({'message': 'Capture stopped'})

    @classmethod
    def capture_screenshots_thread(cls, employee_id, company_id):
        while cls.is_capturing:
            screenshot = cls().capture_screenshot(employee_id,  company_id)  # Instantiate the class to call the method
            time.sleep(5)  # Wait for 5 seconds between captures




# Shared variable to track the monitoring status
monitoring_status = False
init_app_process = None  # Store the subprocess object for initApp.py

@csrf_exempt
def update_status(request, status_type, status_message):
    # Save the status information to the database
    employee_id = request.user.employs.id
    company_id = request.user.employs.companyid.id

    SystemStatus.objects.create(status_type=status_type, status_message=status_message , employee_id=employee_id,company_id=company_id)

    return JsonResponse({'status': 'success'})
 
def powercheck(request):
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'start':
            return start_activity_tracking(request)
        elif action == 'stop':
            return stop_and_save_activities(request)
        elif action == 'update_status':
            status_type = request.POST.get('status_type')
            status_message = request.POST.get('status_message')
            return update_status(request, status_type, status_message)

    return render(request, "employ-template/calendar.html")

def start_activity_tracking(request):
    global monitoring_status, init_app_process
    employee_id = request.user.employs.id
    company_id = request.user.employs.companyid.id

    json_data = {
         "employee_id": employee_id,
         "company_id":company_id
    }
    json_string = json.dumps(json_data)

    # Stop the existing initApp.py subprocess if it's running
    if init_app_process and init_app_process.poll() is None:
        init_app_process.terminate()


    # Start a new initApp.py subprocess
    init_app_process = subprocess.Popen(['python', 'engine/initApp.py', json_string])

    monitoring_status = True
    return JsonResponse({'status': 'success', 'message': 'Activities started.'})

def stop_and_save_activities(request):
    global monitoring_status, init_app_process
    employee_id = request.user.employs.id
    company_id = request.user.employs.companyid.id


    json_data = {
         "employee_id": employee_id,
         "company_id":company_id
    }
    json_string = json.dumps(json_data)

    # Stop the initApp.py subprocess if it's running
    if init_app_process and init_app_process.poll() is None:
        init_app_process.terminate()
        init_app_process.wait()

    # Start the destroyApp.py subprocess
    
    monitoring_status = False
    return JsonResponse({'status': 'success', 'message': 'Activities stopped.'})














# def employ_profile_save(request):
#     if request.method!="POST":
#         return HttpResponseRedirect(reverse("employ_profile"))
#     else:
#         first_name=request.POST.get("first_name")
#         last_name=request.POST.get("last_name")
#         password=request.POST.get("password")
#         location=request.POST.get("location")

        
#         try:
#             customuser=CustomUser.objects.get(id=request.user.id)
#             customuser.first_name=first_name
#             customuser.last_name=last_name
            
#             if password!=None and password!="":
#                 customuser.set_password(password)
#             customuser.save()

#             employ=Employs.objects.get(admin=customuser)
#             employ.first_name=first_name
#             employ.last_name=last_name
#             employ.location=location
#             employ
#             employ.save()
#             messages.success(request, "Successfully Updated Profile")
#             return HttpResponseRedirect(reverse("employ_profile"))
#         except:
#             messages.error(request, "Failed to Update Profile")
#             return HttpResponseRedirect(reverse("employ_profile"))

def employ_profile_save(request):
    if request.method == "POST":
        first_name=request.POST.get("first_name")
        last_name=request.POST.get("last_name")
        password=request.POST.get("password")
        location=request.POST.get("location")
        customuser=CustomUser.objects.get(id=request.user.id)
        customuser.first_name=first_name
        customuser.last_name=last_name
        if password!=None and password!="":
            customuser.set_password(password)
        customuser.save()
        employ=Employs.objects.get(admin=customuser)
        employ.first_name=first_name
        employ.last_name=last_name
        employ.location=location
        employ
        employ.save()
        messages.success(request, "Successfully Updated Profile")
    return redirect('/show_login')


def employ_add(request):
    
    employ_obj=Employs.objects.get(admin=request.user.id)
    try:
        company = Companys.objects.get(id=employ_obj.companyid.id)  
        organization_name = company.organizationname 
    except Companys.DoesNotExist:
        organization_name = "No Organization Name Available" 


    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)


    compid=data.companyid

   

    employ_obj=Employs.objects.get(admin=request.user.id)
    try:
        company = Companys.objects.get(id=employ_obj.companyid.id)  
        organization_name = company.organizationname 
    except Companys.DoesNotExist:
        organization_name = "No Organization Name Available" 

    



    
    if request.method=="POST":
        firstname1=request.POST["firstname1"]
        lastname1=request.POST["lastname1"]
        
        pan=request.POST["pan"]
        ifsecode=request.POST["ifsecode"]
        acno=request.POST["acno"]
        beneficiaryname=request.POST["beneficiaryname"]
        phno=request.POST["phno"]  
        gender1=request.POST.get('gender1')
        dob=request.POST["dob"]
        address1=request.POST["address1"]
        heq=request.POST["heq"]
        aadharno=request.POST["aadharno"] 
        bloodgroup=request.POST["bloodgroup"] 
        # fathername=request.POST.get("fathername")
        # fathersdob=request.POST.get("fathersdob")
        # mothername=request.POST.get("mothername")
        # mothersdob=request.POST.get("mothersdob")
        # Childdetails1=request.POST.get("Childdetails1")
        # Childdetails2=request.POST.get("Childdetails2")
        # maritalstatus=request.POST.get("maritalstatus")
        # workexperiance=request.POST.get("workexperiance")
        # previousemploye=request.POST.get("previousemploye")
        # previousdesignation=request.POST.get("previousdesignation")
        # Marriageannivarsary=request.POST.get("Marriageannivarsary")
        # emergencycontactname=request.POST.get("emergencycontactname")
        # emergencycontactnumber=request.POST.get("emergencycontactnumber")
        # emergencycontactrelation=request.POST.get("emergencycontactrelation")
        # nationality=request.POST.get('nationality')
        # qualification=request.POST.get("qualification")
        profile_pic=request.FILES.get('profile_pic')
        student_obj=Employs.objects.get(admin=request.user.id)
        try:
            emp_instance=employ_add_form.objects.get(student_id=student_obj)
            emp_instance.firstname1=firstname1
            emp_instance.lastname1=lastname1
            emp_instance.pan=pan
            emp_instance.ifsecode=ifsecode
            emp_instance.acno=acno
            emp_instance.beneficiaryname=beneficiaryname
            emp_instance.phno=phno
            emp_instance.gender1=gender1
            emp_instance.dob=dob
            emp_instance.address1=address1
            emp_instance.heq=heq
            emp_instance.aadharno=aadharno
            emp_instance.bloodgroup=bloodgroup
            # emp_instance.fathername=fathername
            # emp_instance.fathersdob=fathersdob
            # emp_instance.mothername=mothername
            # emp_instance.mothersdob=mothersdob
            # emp_instance.Childdetails1=Childdetails1
            # emp_instance.Childdetails2=Childdetails2
            # emp_instance.maritalstatus=maritalstatus
            # emp_instance.workexperiance=workexperiance
            # emp_instance.previousdesignation=previousdesignation
            # emp_instance.Marriageannivarsary=Marriageannivarsary
            # emp_instance.emergencycontactname=emergencycontactname
            # emp_instance.emergencycontactnumber=emergencycontactnumber
            # emp_instance.emergencycontactrelation=emergencycontactrelation
            # emp_instance.nationality=nationality
            # emp_instance.qualification=qualification
            emp_instance.profile_pic=profile_pic
            # emp_instance.previousemploye=previousemploye

            emp_instance.save()
        except employ_add_form.DoesNotExist:
            k=employ_add_form(profile_pic=profile_pic,student_id=student_obj,firstname1=firstname1,lastname1=lastname1,email2=student_obj.email,pan=pan,ifsecode=ifsecode,acno=acno,beneficiaryname=beneficiaryname,phno=phno,gender1=gender1,dob=dob,address1=address1,heq=heq,aadharno=aadharno,bloodgroup=bloodgroup)
            k.save();
   
        return HttpResponseRedirect(reverse("employ_profile"))
    return render(request,"employ-template/emp1form.html",{'data':data,'organization_name':organization_name,})


# def employ_add(request):
#     s=employnav.objects.all()
#     if request.method=="POST":
#         firstname1=request.POST["firstname1"]
#         lastname1=request.POST["lastname1"]
        
#         pan=request.POST["pan"]
#         ifsecode=request.POST["ifsecode"]
#         acno=request.POST["acno"]
#         beneficiaryname=request.POST["beneficiaryname"]
#         phno=request.POST["phno"]  
#         gender1=request.POST['gender1']
#         dob=request.POST["dob"]
#         address1=request.POST["address1"]
#         heq=request.POST["heq"]
#         aadharno=request.POST["aadharno"] 
#         bloodgroup=request.POST["bloodgroup"] 
#         profile_pic=request.FILES.get('profile_pic')
#         student_obj=Employs.objects.get(admin=request.user.id)
#         try:
#             emp_instance=employ_add_form.objects.get(student_id=student_obj)
#             emp_instance.firstname1=firstname1
#             emp_instance.lastname1=lastname1
#             emp_instance.pan=pan
#             emp_instance.ifsecode=ifsecode
#             emp_instance.acno=acno
#             emp_instance.beneficiaryname=beneficiaryname
#             emp_instance.phno=phno
#             emp_instance.gender1=gender1
#             emp_instance.dob=dob
#             emp_instance.address1=address1
#             emp_instance.heq=heq
#             emp_instance.aadharno=aadharno
#             emp_instance.bloodgroup=bloodgroup
#             emp_instance.profile_pic=profile_pic
            

#             emp_instance.save()
#         except employ_add_form.DoesNotExist:
#             k=employ_add_form(profile_pic=profile_pic,student_id=student_obj,firstname1=firstname1,lastname1=lastname1,email2=student_obj.email,pan=pan,ifsecode=ifsecode,acno=acno,beneficiaryname=beneficiaryname,phno=phno,gender1=gender1,dob=dob,address1=address1,heq=heq,aadharno=aadharno,bloodgroup=bloodgroup)
#             k.save();
   
#         return HttpResponseRedirect(reverse("employ_profile"))
#     return render(request,"employ-template/emp1form.html",{})


def edit_basic_info(request,id):
  k = Employs.objects.get(id=id);
  return render(request,"employ-template/update_basic_info.html",{'k':k})

def update_basic_info(request,id):
    if (request.method=="POST"):
        first_name=request.POST.get("first_name")
        last_name=request.POST.get("last_name")
        email=request.POST.get("email")
        k=Employs.objects.get(id=id);
        
        k.first_name=first_name
        k.last_name=last_name
        k.email=email
        k.save();
        return HttpResponseRedirect(reverse("employ_profile"))
    return render(request,"employ-template/update_basic_info.html")

def edit_pay_info(request,id):
  
    employ_obj=Employs.objects.get(admin=request.user.id)
    try:
        company = Companys.objects.get(id=employ_obj.companyid.id)  
        organization_name = company.organizationname 
    except Companys.DoesNotExist:
        organization_name = "No Organization Name Available" 



  
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)



    compid=data.companyid

    k = employ_add_form.objects.get(id=id);
    return render(request,"employ-template/update_pay_info.html",{'k':k,'organization_name':organization_name,
                                                                })

def update_pay_info(request,id):
    if (request.method=="POST"):
        pan=request.POST.get("pan")
        ifsecode=request.POST.get("ifsecode")
        acno=request.POST.get("acno")
        beneficiaryname=request.POST.get("beneficiaryname")
       
        k=employ_add_form.objects.get(id=id);
        
        k.pan=pan
        k.ifsecode=ifsecode
        k.acno=acno
        k.beneficiaryname=beneficiaryname
        k.save();
        return HttpResponseRedirect(reverse("employ_profile"))
    return render(request,"employ-template/update_pay_info.html")

def edit_other_info(request,id):
    employ_obj=Employs.objects.get(admin=request.user.id)
    try:
        company = Companys.objects.get(id=employ_obj.companyid.id)  
        organization_name = company.organizationname 
    except Companys.DoesNotExist:
        organization_name = "No Organization Name Available" 



  
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)



    compid=data.companyid

    k = employ_add_form.objects.get(id=id);
    return render(request,"employ-template/update_other_info.html",{'k':k,'organization_name':organization_name,
                                                                })

# def update_other_info(request,id):
#     if (request.method=="POST"):
#         phno=request.POST.get("phno")
#         gender1=request.POST.get("gender1")
#         dob=request.POST.get("dob")
#         address1=request.POST.get("address1")
#         heq=request.POST.get("heq")
#         aadharno=request.POST.get("aadharno")
#         bloodgroup=request.POST.get("bloodgroup")

#         k=employ_add_form.objects.get(id=id);
        
#         k.phno=phno
#         k.gender1=gender1
#         k.dob=dob
#         k.address1=address1
#         k.heq=heq
#         k.aadharno=aadharno
#         k.bloodgroup=bloodgroup
#         k.save();
#         return HttpResponseRedirect(reverse("employ_profile"))
#     return render(request,"employ-template/update_other_info.html")



def update_other_info(request,id):
    if (request.method=="POST"):
        phno=request.POST.get("phno")
        gender1=request.POST.get("gender1")
        dob=request.POST.get("dob")
        address1=request.POST.get("address1")
        heq=request.POST.get("heq")
        aadharno=request.POST.get("aadharno")
        fathername=request.POST.get("fathername")
        fathersdob=request.POST.get("fathersdob")
        mothername=request.POST.get("mothername")
        mothersdob=request.POST.get("mothersdob")
        Childdetails1=request.POST.get("Childdetails1")
        Childdetails2=request.POST.get("Childdetails2")
        maritalstatus=request.POST.get("maritalstatus")
        workexperiance=request.POST.get("workexperiance")
        previousemploye=request.POST.get("previousemploye")
        previousdesignation=request.POST.get("previousdesignation")
        Marriageannivarsary=request.POST.get("Marriageannivarsary")
        emergencycontactname=request.POST.get("emergencycontactname")
        emergencycontactnumber=request.POST.get("emergencycontactnumber")
        emergencycontactrelation=request.POST.get("emergencycontactrelation")
        nationality=request.POST.get('nationality')
        qualification=request.POST.get("qualification")
        bloodgroup=request.POST.get("bloodgroup")

        k=employ_add_form.objects.get(id=id);
        
        k.phno=phno
        k.gender1=gender1
        k.dob=dob
        k.address1=address1
        k.heq=heq
        k.aadharno=aadharno
        k.fathername=fathername
        k.fathersdob=fathersdob
        k.mothername=mothername
        k.mothersdob=mothersdob
        k.Childdetails1=Childdetails1
        k.Childdetails2=Childdetails2
        k.maritalstatus=maritalstatus
        k.workexperiance=workexperiance
        k.previousemploye=previousemploye
        k.previousdesignation=previousdesignation
        k.Marriageannivarsary=Marriageannivarsary
        k.emergencycontactname=emergencycontactname
        k.emergencycontactnumber=emergencycontactnumber
        k.emergencycontactrelation=emergencycontactrelation
        k.nationality=nationality
        k.qualification=qualification
        k.bloodgroup=bloodgroup
        k.save();
        return HttpResponseRedirect(reverse("employ_profile"))
    return render(request,"employ-template/update_other_info.html")

def payslip_table(request):
    user=CustomUser.objects.get(id=request.user.id)
    employ=Employs.objects.get(admin=user)
    email_id=request.session.get('e_mail')
    p=employ_payslip.objects.filter(email=email_id).values();
    
    return render(request,"employ-template/paysliptable.html",{'p':p,'employ':employ})


def send_payslip_mail(request):
    # Get the selected year and month from the request
    selected_year = int(request.GET.get('year', datetime.now().year))
    selected_month = int(request.GET.get('month', datetime.now().month))

    # Check if December of the current year is completed
    today = datetime.now()
    if today.month == 12 and today.day > 31:
        selected_month = 1  # Reset to January of the next year
        selected_year += 1  # Increment the year

    # Calculate the start date for filtering past months
    start_date = datetime(selected_year, selected_month, 1) - timedelta(days=1)

    # Calculate the end date based on the selected time frame (3 or 6 months)
    time_frame = request.GET.get('time_frame', '3')  # Default to 3 months if not specified
    if time_frame == '6':
        end_date = start_date - timedelta(days=181)
    else:
        end_date = start_date - timedelta(days=91)

    # Filter data from the `pay` model for the selected time frame
    # filtered_data = MonthlyTotal.objects.filter(date_field__range=[end_date, start_date])

    # Prepare data for the dropdowns
    year_range = range(2021, datetime.now().year + 1)
    month_range = range(1, 13)  # 1 to 12 for months

    # Retrieve data from the `MonthlyTotal` model for the selected time frame
    if time_frame == '6':
        # If the time frame is 6 months, fetch data for the past 6 months
        start_date_month = selected_month - 5 if selected_month >= 6 else 12 - (5 - selected_month)
        start_date_year = selected_year if selected_month >= 6 else selected_year - 1
    else:
        # Default to fetching data for the past 3 months
        start_date_month = selected_month - 2 if selected_month >= 3 else 12 - (2 - selected_month)
        start_date_year = selected_year if selected_month >= 3 else selected_year - 1

    monthly_total_data = MonthlyTotal.objects.filter(
        year__gte=start_date_year,
        year__lte=selected_year,
        month__gte=start_date_month,
        month__lte=selected_month
    )
   

    return render(request, 'employ-template/search_payslip.html', {
        'selected_year': selected_year,
        'selected_month': selected_month,
        # 'filtered_data': filtered_data,
        'year_range': year_range,
        'month_range': month_range,
        'time_frame': time_frame,
        'monthly_total_data': monthly_total_data,})


def send_payslip_to_employeemail(request):
    from_email = settings.EMAIL_HOST_USER
    email=request.session.get('email_2')
    e=Employs.objects.filter(email=email).values();
    recipient_list =[email]
    html_content = render_to_string("employ-template/employ_email_template.html",{'title':'test email','e':e})
    text_content=strip_tags(html_content)
    subject="YOUR PAYSLIP"
    email = EmailMultiAlternatives(
        subject,
        text_content,
        from_email,
        recipient_list,
        )
    email.attach_alternative(html_content,"text/html")
    email.fail_silently = True
    email.send() 
    return redirect('/send_payslip_mail',{'e':e})

def payslip_table(request):
    user=CustomUser.objects.get(id=request.user.id)
    employ=Employs.objects.get(admin=user)
    email=request.session.get('email_2')
    e=Employs.objects.filter(email=email).values()
    return render(request,"employ-template/paysliptable.html",{'e':e,'employ':employ})

def pdf_report_create(request):
    user=CustomUser.objects.get(id=request.user.id)
    employ=Employs.objects.get(admin=user)
    email=request.session.get('email_2')
    c=company_details.objects.all()
    e=Employs.objects.filter(email=email).values()
    # email_id=request.session.get('e_mail')
    # p=employ_payslip.objects.filter(email=email_id).values();
    template_path = 'employ-template/pay.html'

    context = {'employ':employ,'e':e,'c':c,}

    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = 'filename="products_report.pdf"'

    template = get_template(template_path)

    html = template.render(context)

    # create a pdf
    pisa_status = pisa.CreatePDF(
       html, dest=response)
    # if error then show some funy view
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response


def pay(request):
    user=CustomUser.objects.get(id=request.user.id)
    employ=Employs.objects.get(admin=user)
    email_id=request.session.get('e_mail')
    p=employ_payslip.objects.filter(email=email_id).values();
    
    return render(request,"employ-template/pay.html",{'p':p,'employ':employ})

def employ_payroll_table(request):
    user=CustomUser.objects.get(id=request.user.id)
    employ=Employs.objects.get(admin=user)
    email_id=request.session.get('e_mail')
    p=employ_payslip.objects.filter(email=email_id).values()
    return render(request,"employ-template/employ_table.html",{'p':p,'employ':employ})


from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.urls import reverse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib import messages
from django.db.models import Sum
import json
from .models import Employs, ad_salary, emidata, salary_struct, salary_deductions, company_details
from .models import Companys  # Assuming this is the correct import for your Companys model
from notifications.signals import notify
from datetime import datetime

def advancesalary_request(request):
    

    employs = Employs.objects.get(admin=request.user.id)
    data = Employs.objects.filter(admin=request.user.id).first()
    data1 = data.id
    data2 = data.hroptions
    project_manager = data.projectmanagerop
    compid = data.companyid
    regname = Companys.objects.filter(id=data.companyid.id).first()
    organization_name = regname.organizationname
    teamleadop = TeamMember.objects.filter(employee=data1, is_team_lead=1)
    company_id = data.companyid
    company_name = company_id.organizationname
    company_details_obj = company_details.objects.filter(companyid=company_id).first()
    rst = Employs.objects.filter(companyid=company_id, admin=request.user.id)
    salarper = salary_struct.objects.filter(companyid=company_id)
    salarded = salary_deductions.objects.filter(companyid=company_id)
    salbasic = salary_struct.objects.filter(companyid=company_id, salarycomponent="Basic Salary")
    salaryexbasic = salary_struct.objects.filter(companyid=company_id).exclude(salarycomponent="Basic Salary")
    alowsum = salary_struct.objects.filter(companyid=company_id).exclude(salarycomponent="Basic Salary").aggregate(Sum('percentageofCTC'))['percentageofCTC__sum']
    deduct = salary_deductions.objects.filter(companyid=company_id).exclude(percentageordfixed="Fixed").aggregate(Sum('percentageofdctc'))['percentageofdctc__sum']
    deductfix = salary_deductions.objects.filter(companyid=company_id, percentageordfixed="Fixed").aggregate(Sum('percentageofdctc'))['percentageofdctc__sum']
    empdata=Employs.objects.filter(admin=request.user.id).first()
    edata=empdata.companyid_id
    cmpdata=set_payroll_date.objects.filter(companyid_id=edata)
    net_pay_list = []
    for i in rst:
        net_pay = 0
        if deductfix or deduct:
            for st in salbasic:
                package_value = float(i.package) if i.package is not None else 0
                percentage_value = float(st.percentageofCTC) if st.percentageofCTC is not None else 0
                deduct_value = float(deduct) if deduct is not None else 0
                deductfix_value = float(deductfix) if deductfix is not None else 0
                
                all_alow = float(i.package) / 12 * 100 / 100 * float(alowsum) / 100
                all_ded = (package_value / 12 * percentage_value / 100 * deduct_value / 100) + deductfix_value
                aftersum = all_alow - all_ded
                  
                net_pay = max(net_pay, package_value / 12 * percentage_value / 100 + aftersum)
        else:
            net_pay = float(i.package) / 12 * 100 / 100
        net_pay_list.append(net_pay)

    max_net_pay = max(net_pay_list)
    net_pay_list_json = json.dumps(net_pay_list)

    if request.method == 'POST':
        amount = request.POST.get('amount')
        eminumber_id = request.POST.get('eminumber')
        reason = request.POST.get('reason')
        selected_months = None  # Default value

        if eminumber_id is not None:
            selected_months = int(eminumber_id)

        if selected_months is None:
            messages.error(request, "Please select the number of EMI months.")
            return HttpResponseRedirect(reverse("advancesalary_request"))
        elif selected_months <= 0:
            messages.error(request, "Number of EMI months must be greater than zero.")
            return HttpResponseRedirect(reverse("advancesalary_request"))
        elif not amount.isdigit() or int(amount) <= 0:
            messages.error(request, "Amount must be a positive number.")
            return HttpResponseRedirect(reverse("advancesalary_request"))
        else:
            max_value = max_net_pay * selected_months
            try:
                amount_int = int(amount)
                if amount_int > max_value:
                    messages.error(request, f"Amount cannot exceed the maximum value for the selected number of months: {max_value}")
                    return HttpResponseRedirect(reverse("advancesalary_request"))

                leave_data = ad_salary.objects.create(
                    amount=amount,
                    student_id=employs,
                    eminumber_id=eminumber_id,
                    reason=reason,
                    request_status=0,
                    company_id=request.user.employs.companyid
                )
   # Send notifications
              # Send notifications
                employsss = Employs.objects.get(id=employs.id)
                admin_obj = request.user.employs.companyid.id
                companys_admin = Companys.objects.get(id=admin_obj)
                employss = Employs.objects.filter(hroptions=1, companyid=companys_admin)
                for admin in employss:
                    notify.send(sender=employsss.admin, recipient=admin.admin, verb='Advance salary', description=f"{employsss.first_name}.{employsss.last_name} has applied for Advance Salary")
                notify.send(sender=employsss.admin, recipient=companys_admin.usernumber, verb='Advance salary', description=f"{employsss.first_name}.{employsss.last_name} has applied for Advance salary")

                messages.success(request, "Advance salary request submitted successfully")
            except Exception as e:
                messages.error(request, f"Failed To Apply for Advance Salary: {e}")

        return HttpResponseRedirect(reverse("advancesalary_request"))  # Redirect after POST

    staff_obj = Employs.objects.get(admin=request.user.id)
    leave_data = ad_salary.objects.filter(student_id=staff_obj)
    emidata_instance = emidata.objects.filter(companyid=company_id).first()

    eminumber_options = []
    if emidata_instance:
        eminumber_options = range(1, emidata_instance.eminumber + 1)

    paginator = Paginator(leave_data, 10)
    page_number = request.GET.get('page')
    try:
        leave_data = paginator.page(page_number)
    except PageNotAnInteger:
        leave_data = paginator.page(1)
    except EmptyPage:
        leave_data = paginator.page(paginator.num_pages)

    context = {
        'rst': rst,
        'salarper': salarper,
        'salarded': salarded,
        'salbasic': salbasic,
        'salaryexbasic': salaryexbasic,
        'alowsum': alowsum,
        'deduct': deduct,
        'eminumber_options': eminumber_options,
        'organization_name': company_name,
        'leave_data': leave_data,
        'data': data,
        'deductfix': deductfix,
        'company_details': company_details_obj,
        'max_val': max_net_pay,
        'net_pay_list_json': net_pay_list_json,
        'employs': employs,
        'compid': company_id,
        'organization_name': organization_name,
        'cmpdata':cmpdata
       

    }

    return render(request, "employ-template/employ_advsalayr_request.html", context)


def advancesalary_request_save(request):
    if request.method != "POST":
        return HttpResponseRedirect(reverse("advancesalary_request"))
    else:
        # Get the month and year as input, e.g., "2023-10"
        eminumber = request.POST.get("eminumber")
        amount = request.POST.get("amount")
        reason = request.POST.get("reason")

        # Convert the month and year string to a Python datetime object
        
        student_obj = request.user.employs.id
        admin_obj = request.user.employs.companyid.id
        employsss = Employs.objects.get(id=student_obj.id)
        companys_admin = Companys.objects.get(id=admin_obj)
        # employsshr = Employs.objects.filter(hroptions=1,companyid=companys_admin,admin=request.user.id)
        # employss1 = Employs.objects.filter(admin=request.user.id,hroptions=1,companyid=companys_admin)

        try:
            leave_report = ad_salary(eminumber=eminumber, student_id_id=student_obj.id, amount=amount, reason=reason, request_status=0,companyid_id=admin_obj)
            leave_report.save()
            for admin in employsss:
                   notify.send(sender=employsss.admin, recipient=admin.admin, verb='Advance salary', description=f"{employsss.first_name}.{employsss.last_name} has applied for Leave")

            messages.success(request, "Successfully Applied for Leave1")
            return HttpResponseRedirect(reverse("advancesalary_request"))
        except:
            messages.error(request, "Failed To Apply for Leave")
            return HttpResponseRedirect(reverse("advancesalary_request"))

     
def tax_slip(request):
    user = CustomUser.objects.get(id=request.user.id)
    emp=Employs.objects.get(admin=user)
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    email=Employs.objects.get(admin=request.user.id)
    tax=employ_tax_form.objects.filter(employ=email)
    return render(request,"employ-template/tax_table.html",{"tax":tax,'emp':emp})

def edit_home_rent(request,id):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    k = employ_tax_form.objects.get(id=id);
    return render(request,"employ-template/update_home_rent.html",{'k':k,})

def update_home_rent(request,id):
    if (request.method=="POST"):
        Current_Monthly_Rent=request.POST.get("Current_Monthly_Rent")
        Name_of_landlord=request.POST.get("Name_of_landlord")
        PAN_of_landlord=request.POST.get("PAN_of_landlord")
        Address_of_landlord=request.POST.get("Address_of_landlord")
        home_rent_proof=request.FILES.get("home_rent_proof")
        from_month=request.POST.get("from_month")
        to_month=request.POST.get("to_month")
        
        k=employ_tax_form.objects.get(id=id);
        
        k.Current_Monthly_Rent=Current_Monthly_Rent
        k.Name_of_landlord=Name_of_landlord
        k.PAN_of_landlord=PAN_of_landlord
        k.Address_of_landlord=Address_of_landlord
        k.home_rent_proof=home_rent_proof
        k.from_month=from_month
        k.to_month=to_month
        k.save();
        return HttpResponseRedirect(reverse("tax_slip"))
    return render(request,"employ-template/update_home_rent.html")

def edit_tax_sections(request,id):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    k = employ_tax_form.objects.get(id=id);
    return render(request,"employ-template/update_tax_sections.html",{'k':k,})

def update_tax_sections(request,id):
    if (request.method=="POST"):
        Section_80C=request.POST.get("Section_80C")
        Section_80CCD=request.POST.get("Section_80CCD")
        Section_80D=request.POST.get("Section_80D")
        Section_80DD=request.POST.get("Section_80DD")
        Section_80E=request.POST.get("Section_80E")
        Section_80EEB=request.POST.get("Section_80EEB")
        Section_80G=request.POST.get("Section_80G")
        Section_80U=request.POST.get("Section_80U")
        Section_80DDB=request.POST.get("Section_80DDB")
        Section_80TTA=request.POST.get("Section_80TTA")
        Section_80TTB=request.POST.get("Section_80TTB")
        Section_80C_proof=request.FILES.get("Section_80C_proof")
        Section_80D_proof=request.FILES.get("Section_80D_proof")
        Section_80CCD1BS_proof=request.FILES.get("Section_80CCD1BS_proof")
        Section_80DD_proof=request.FILES.get("Section_80DD_proof")
        Section_80E_proof=request.FILES.get("Section_80E_proof")
        Section_80G_proof=request.FILES.get("Section_80G_proof")
        Section_80U_proof=request.FILES.get("Section_80U_proof")
        Section_80EEB_proof=request.FILES.get("Section_80EEB_proof")
        Section_80DDB_proof=request.FILES.get("Section_80DDB_proof")
        Section_80TTA_proof=request.FILES.get("Section_80TTA_proof")
        Section_80TTB_proof=request.FILES.get("Section_80TTB_proof")

        k=employ_tax_form.objects.get(id=id);
        k.Section_80C=Section_80C
        k.Section_80CCD=Section_80CCD
        k.Section_80D=Section_80D
        k.Section_80DD=Section_80DD
        k.Section_80E=Section_80E
        k.Section_80EEB=Section_80EEB
        k.Section_80G=Section_80G
        k.Section_80U=Section_80U
        k.Section_80DDB=Section_80DDB
        k.Section_80TTA=Section_80TTA
        k.Section_80TTB=Section_80TTB
        k.Section_80C_proof=Section_80C_proof
        k.Section_80D_proof=Section_80D_proof
        k.Section_80CCD1BS_proof=Section_80CCD1BS_proof
        k.Section_80DD_proof=Section_80DD_proof
        k.Section_80E_proof=Section_80E_proof
        k.Section_80G_proof=Section_80G_proof
        k.Section_80U_proof=Section_80U_proof
        k.Section_80EEB_proof=Section_80EEB_proof
        k.Section_80DDB_proof=Section_80DDB_proof
        k.Section_80TTA_proof=Section_80TTA_proof
        k.Section_80TTB_proof=Section_80TTB_proof
        k.save();
        return HttpResponseRedirect(reverse("tax_slip"))
    return render(request,"employ-template/update_tax_sections.html")

def edit_home_loan(request,id):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    k = employ_tax_form.objects.get(id=id);
    return render(request,"employ-template/update_home_loan.html",{'k':k,})

def update_home_loan(request,id):
    if (request.method=="POST"):
        Annual_interest_payable=request.POST.get("Annual_interest_payable")
        Additional_benefit_under_Section=request.POST.get("Additional_benefit_under_Section")
        Name_of_lender=request.POST.get("Name_of_lender")
        PAN_of_lender=request.POST.get("PAN_of_lender")
        Address_of_lender=request.POST.get("Address_of_lender")
        Section_80EEA=request.POST.get("Section_80EEA")
        House_Property_proof=request.FILES.get("House_Property_proof")
        Section_80EE_proof=request.FILES.get("Section_80EE_proof")
        Section_80EEA_proof=request.FILES.get("Section_80EEA_proof")

        k=employ_tax_form.objects.get(id=id);
        k.Annual_interest_payable=Annual_interest_payable
        k.Additional_benefit_under_Section=Additional_benefit_under_Section
        k.Name_of_lender=Name_of_lender
        k.PAN_of_lender=PAN_of_lender
        k.Address_of_lender=Address_of_lender
        k.Section_80EEA=Section_80EEA
        k.House_Property_proof=House_Property_proof
        k.Section_80EE_proof=Section_80EE_proof
        k.Section_80EEA_proof=Section_80EEA_proof
        k.save();
        return HttpResponseRedirect(reverse("tax_slip"))
    return render(request,"employ-template/update_home_loan.html")


def edit_travel_allowance(request,id):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    k = employ_tax_form.objects.get(id=id);
    return render(request,"employ-template/update_travel_allowance.html",{'k':k,})

def update_travel_allowance(request,id):
    if (request.method=="POST"):
        Amount=request.POST.get("Amount")
        Origin=request.POST.get("Origin")
        Destination=request.POST.get("Destination")
        TravelStartDate=request.POST.get("TravelStartDate")
        k=employ_tax_form.objects.get(id=id);
        k.Amount=Amount
        k.Origin=Origin
        k.Destination=Destination
        k.TravelStartDate=TravelStartDate
        k.save();
        return HttpResponseRedirect(reverse("tax_slip"))
    return render(request,"employ-template/update_travel_allowance.html")

@csrf_exempt
def employ_fcmtoken_save(request):
    token=request.POST.get("token")
    try:
        employ=Employs.objects.get(admin=request.user.id)
        employ.fcm_token=token
        employ.save()
        return HttpResponse("True")
    except:
        return HttpResponse("False")


def pay(request):
    user=CustomUser.objects.get(id=request.user.id)
    employ=Employs.objects.get(admin=user)
    email_id=request.session.get('e_mail')
    p=employ_payslip.objects.filter(email=email_id).values();
    
    return render(request,"employ-template/pay.html",{'p':p,'employ':employ})

def employ_payroll_table(request):
    today = datetime.now()
    current_year = today.year
    user=CustomUser.objects.get(id=request.user.id)
    email_id=request.session.get('e_mail')
    user_obj=Employs.objects.get(admin=request.user.id)
    package = int(user_obj.package)
    # p=employ_payslip.objects.filter(email=email_id).values()
    return render(request,"employ-template/employ_table.html",{'user':user,'user_obj':user_obj})

def employ_all_notification(request):
    

        
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)

 
 

    employ=Employs.objects.get(admin=request.user.id)
    notifications=NotificationEmploy.objects.filter(employ_id=employ.id)
    notifications1=NotificationEmploy.objects.filter(employ_id=employ.id).count()
    notifications2=admin_project_create.objects.filter(is_status=0,read=0,admin_id=employ.id).count() 
    k=admin_project_create.objects.filter(admin_id=employ.id) 

    return render(request,"employ-template/all_notification.html",{'data':data,"notifications":notifications,"notifications1":notifications1,'notifications2':notifications2,'k':k, 


})



def mark_as_read_done(request, admin_project_create_id):
    admin_project_create_obj = get_object_or_404(admin_project_create, id=admin_project_create_id)
    admin_project_create_obj.read = True
    admin_project_create_obj.save()
    return HttpResponseRedirect('/Employ_home') 


def payslip_apply_view(request):
    

          
    employs = Employs.objects.all()
    employs = Employs.objects.get(admin=request.user.id)  # Assuming there's only one employ for each admin
    joining_date = employs.dateofjoining
    startingdate=joining_date.strftime('%Y-%m')
    curren=datetime.today()
    currentm=curren-timedelta(days=curren.day)
    currentmonth=currentm.strftime('%Y-%m')

    hr=HR.objects.all()
    emp = CustomUser.objects.filter(user_type=2,is_active=True).count()  # Count active employees
    dis = CustomUser.objects.filter(user_type=2,is_active=False).count()  # Count inactive employees

    staff_obj = Employs.objects.get(admin=request.user.id)
    leave_data=payslip_request.objects.filter(student_id=staff_obj)

    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    compid=data.companyid
    adm=admin_drop.objects.filter(name="Advance Salary").first()
    mon=Addonsuser.objects.filter(companyid=regname,plan="Advance Salary").first()

   
    return render(request,"employ-template/payslip_request_apply_view.html",{'currentmonth':currentmonth, 'joining_date':joining_date,'startingdate':startingdate,'adm':adm,'mon':mon, "data":data,'organization_name':organization_name,"data":data,"leave_data":leave_data,'employs':employs,'hr':hr,'emp':emp,'dis':dis, 


})


def payslip_apply_view_save(request):
    if request.method != "POST":
        return HttpResponseRedirect(reverse("payslip_apply_view"))
    else:
        # Get the month and year as input, e.g., "2023-10"
        startingdate = request.POST.get("startingdate")
        endingdate = request.POST.get("endingdate")
        
        reason = request.POST.get("reason")

        # Convert the month and year string to a Python datetime object
        try:
            startingdate1 = datetime.strptime(startingdate, "%Y-%m")
            endingdate1=datetime.strptime(endingdate, "%Y-%m")
        except ValueError:
            messages.error(request, "Invalid month and year format")
            return HttpResponseRedirect(reverse("payslip_apply_view"))

        student_obj=request.user.employs.id
        admin_obj = request.user.employs.companyid.id
        employsss = Employs.objects.get(id=student_obj)
        companys_admin = Companys.objects.get(id=admin_obj)
        try:
            leave_report = payslip_request(student_id_id=student_obj, startingdate=startingdate1,endingdate=endingdate1, reason=reason, status=0,companyid_id=admin_obj)
            leave_report.save()
            notify.send(sender=employsss.admin ,recipient=companys_admin.usernumber, verb='Payslip', description=f"{employsss.first_name}.{employsss.last_name} has requested for payslips" )
            messages.success(request, "Successfully applied for payslip")

            return HttpResponseRedirect(reverse("payslip_apply_view"))

        except:
            messages.error(request, "Failed To Apply for payslip")
            return HttpResponseRedirect(reverse("payslip_apply_view"))

def paysliprequest(request):
    

         

    staff_obj =Employs.objects.get(admin=request.user.id)

    k = types.objects.all()
    
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)





    st4 = request.POST.get("ss")
    st3 = request.POST.get("vk")
    st = request.POST.get("from_date")
    st1 = request.POST.get("to_date")

    leave_data = payslip_request.objects.filter(student_id=staff_obj).order_by('-startingdate')

    # Apply filters based on user input
    if st4 and st4 != '----Select----':
        leave_data = leave_data.filter(status=st4)
    if st3:
        leave_data = leave_data.filter(typea__icontains=st3)
    if st and st1:
        # Convert string dates to datetime objects
        start_date = datetime.strptime(st, "%Y-%m-%d")
        end_date = datetime.strptime(st1, "%Y-%m-%d")

        # Filter the data based on the date range
        leave_data = leave_data.filter(date__range=[start_date, end_date])

    total_approved = leave_data.filter(status=1)
    total_pending = leave_data.filter(status=0)


   
    items_per_page = 20  # Adjust the number of items per page as needed

    paginator = Paginator( leave_data, items_per_page)

    page = request.GET.get('page')

    try:

        leave_data = paginator.page(page)

    except PageNotAnInteger:

        leave_data = paginator.page(1)

    except EmptyPage:

        leave_data = paginator.page(paginator.num_pages)

    return render(request, "employ-template/payslipreq.html", {
        'leave_data': leave_data,
        'k': k,
        'data1':data1,'data':data,
        
        'organization_name':organization_name,
        'total': total_approved,
        'total1': total_pending,
      

        
    })



def help(request):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    return render(request,"employ-template/help.html",{})
    from django.shortcuts import render,redirect



from .models import employhelp
# Create your views here.
def data(request):
    


    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    compid=data.companyid
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
  
 

    redirect('/home')
    return render(request,"employhelp/help.html",{'organization_name':organization_name,'data':data,'data':data,
})
def home(request):
    
    return render(request,"employhelp/getstarted.html")
def home1(request):
    
    return render(request,"employhelp/leave.html")
def home2(request):
    
    return render(request,"employhelp/taxreduction.html")
def home3(request):
   
    return render(request,"employhelp/reimbesement.html")
def home4(request):
    
    return render(request,"employhelp/investments.html")
def home5(request):
    
    return render(request,"employhelp/mytax.html")
def home6(request):
    return render(request,"employhelp/search.html")
def home7(request):
    return render(request,"employhelp/gstarted.html")
def home8(request):
    return render(request,"employhelp/people.html")
def home9(request):
    return render(request,"employhelp/payroll.html")
def home10(request):
    return render(request,"employhelp/statutory.html")
def home11(request):
    return render(request,"employhelp/payment.html")
def home12(request):
    return render(request,"employhelp/selfservice.html")
def home13(request):
    return render(request,"employhelp/insurance.html")
def home14(request):
    return render(request,"employhelp/accountintegration.html")
def home15(request):
    return render(request,"employhelp/integration.html")
def home16(request):
    return render(request,"employhelp/modules.html")
def home17(request):
    return render(request,"employhelp/updates.html")
def info1(request):
    return render(request,"employhelp/contactsupport.html")
def info2(request):
    return render(request,"employhelp/guied1.html")
def info3(request):
    return render(request,"employhelp/guied2.html")
def info4(request):
    return render(request,"employhelp/guied3.html")
def demo1(request):
    return render(request,"employhelp/demo1.html")
def demo2(request):
    return render(request,"employhelp/demo2.html")
def demo3(request):
    return render(request,"employhelp/demo3.html")
def demo4(request):
    return render(request,"employhelp/demo4.html")
def demo5(request):
    return render(request,"employhelp/demo5.html")
def demo6(request):
    return render(request,"employhelp/demo6.html")
def demo7(request):
    return render(request,"employhelp/demo7.html")
def demo8(request):
    return render(request,"employhelp/demo8.html")
def people1(request):
    return render(request,"employhelp/people1.html")
def people2(request):
    return render(request,"employhelp/people2.html")
def people3(request):
    return render(request,"employhelp/people3.html")
def people4(request):
    return render(request,"employhelp/people4.html")
def people5(request):
    return render(request,"employhelp/people5.html")
def payroll1(request):
    return render(request,"employhelp/payroll1.html")
def payroll2(request):
    return render(request,"employhelp/payroll2.html")
def payroll3(request):
    return render(request,"employhelp/payroll3.html")
def payroll4(request):
    return render(request,"employhelp/payroll4.html")
def payroll5(request):
    return render(request,"employhelp/payroll5.html")
def payroll6(request):
    return render(request,"employhelp/payroll6.html")
def statutory1(request):
    return render(request,"employhelp/statutory1.html")
def statutory3(request):
    return render(request,"employhelp/statutory3.html")
def sfund1(request):
    return render(request,"employhelp/sfund1.html")
def sfund2(request):
    return render(request,"employhelp/sfund2.html")
def sfund3(request):
    return render(request,"employhelp/sfund3.html")
def esi1(request):
    return render(request,"employhelp/esi1.html")
def esi2(request):
    return render(request,"employhelp/esi2.html")
def esi3(request):
    return render(request,"employhelp/esi3.html")
def tds1(request):
    return render(request,"employhelp/tds1.html")
def tds2(request):
    return render(request,"employhelp/tds2.html")
def tds3(request):
    return render(request,"employhelp/tds3.html")
def tds4(request):
    return render(request,"employhelp/tds4.html")
def tax1(request):
    return render(request,"employhelp/tax1.html")
def tax2(request):
    return render(request,"employhelp/tax2.html")
def payment1(request):
    return render(request,"employhelp/payment1.html")
def payment2(request):
    return render(request,"employhelp/payment2.html")
def payment3(request):
    return render(request,"employhelp/payment3.html")
def payment4(request):
    return render(request,"employhelp/payment4.html")
def payment5(request):
    return render(request,"employhelp/payment5.html")
def billing1(request):
    return render(request,"employhelp/billing1.html")

def billing2(request):
    return render(request,"employhelp/billing2.html")

def billing3(request):
    return render(request,"employhelp/billing3.html")
def billing4(request):
    return render(request,"employhelp/billing4.html")

def billing5(request):
    return render(request,"employhelp/billing5.html")
def billing6(request):
    return render(request,"employhelp/billing6.html")
def billing7(request):
    return render(request,"employhelp/billing7.html")
def billing8(request):
    return render(request,"employhelp/billing8.html")
def billing9(request):
    return render(request,"employhelp/billing9.html")
def service1(request):
    return render(request,"employhelp/service1.html")
def service2(request):
    return render(request,"employhelp/service2.html")
def service3(request):
    return render(request,"employhelp/service3.html")
def service5(request):
    return render(request,"employhelp/service5.html")
def attendence1(request):
    return render(request,"employhelp/attendence1.html")
def attendence2(request):
    return render(request,"employhelp/attendence2.html")
def attendence3(request):
    return render(request,"employhelp/attendence3.html")
def attendence4(request):
    return render(request,"employhelp/attendence4.html")
def attendence5(request):
    return render(request,"employhelp/attendence5.html")
def attendence6(request):
    return render(request,"employhelp/attendence6.html")
def attendence7(request):
    return render(request,"employhelp/attendence7.html")
def attendence8(request):
    return render(request,"employhelp/attendence8.html")
def reimbus1(request):
    return render(request,"employhelp/reimbus1.html")
def reimbus2(request):
    return render(request,"employhelp/reimbus2.html")
def dletter1(request):
    return render(request,"employhelp/dletter1.html")
def dletter2(request):
    return render(request,"employhelp/dletter2.html")
def insurance1(request):
    return render(request,"employhelp/insurance1.html")
def insurance2(request):
    return render(request,"employhelp/insurance2.html")
def insurance3(request):
    return render(request,"employhelp/insurance3.html")
def insurance4(request):
    return render(request,"employhelp/insurance4.html")
def insurance5(request):
    return render(request,"employhelp/insurance5.html")
def insurance6(request):
    return render(request,"employhelp/insurance6.html")
def insurance7(request):
    return render(request,"employhelp/insurance7.html")
def insurance8(request):
    return render(request,"employhelp/insurance8.html")
def insurance9(request):
    return render(request,"employhelp/insurance9.html")
def insurance10(request):
    return render(request,"employhelp/insurance10.html")
def insurance11(request):
    return render(request,"employhelp/insurance11.html")
def insurance12(request):
    return render(request,"employhelp/insurance12.html")
def insurance13(request):
    return render(request,"employhelp/insurance13.html")
def insurance14(request):
    return render(request,"employhelp/insurance14.html")
def insurance15(request):
    return render(request,"employhelp/insurance15.html")
def insurance16(request):
    return render(request,"employhelp/insurance16.html")
def insurance17(request):
    return render(request,"employhelp/insurance17.html")
def insurance18(request):
    return render(request,"employhelp/insurance18.html")
def insurance19(request):
    return render(request,"employhelp/insurance19.html")
def slack(request):
    return render(request,"employhelp/slack.html")
def books1(request):
    return render(request,"employhelp/books1.html")
def books2(request):
    return render(request,"employhelp/books2.html")
def books3(request):
    return render(request,"employhelp/books3.html")
def books4(request):
    return render(request,"employhelp/books4.html")
def books5(request):
    return render(request,"employhelp/books5.html")
def books6(request):
    return render(request,"employhelp/books6.html")
def books7(request):
    return render(request,"employhelp/books7.html")
def books8(request):
    return render(request,"employhelp/books8.html")
def home18(request):
    return render(request,"employhelp/tax-regime.html")
def home19(request):
    return render(request,"employhelp/PMacces.html")

def home20(request):
    return render(request,"employhelp/PMassign.html")

def home21(request):
    return render(request,"employhelp/PMprogress.html")

def home22(request):
    return render(request,"employhelp/PMcommunicate.html")

def home23(request):
    return render(request,"employhelp/PMupdateproject.html")

def home24(request):
    return render(request,"employhelp/PMupload.html")

def home25(request):
    return render(request,"employhelp/PMatten.html")

def home26(request):
    return render(request,"employhelp/PMresponsibilities.html")

def home27(request):
    return render(request,"employhelp/PMimplement.html")
def home28(request):
    return render(request,"employhelp/PMcreateproject.html")



def setup_guide(request):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)

    return render(request,"employ-template/setup_guide.html",{})




from datetime import date, timedelta
from django.core.mail import send_mail
from django.shortcuts import render
from .models import employ_add_form  # Make sure to import the correct model

def send_birthday_wishes(request):
    today = date.today()
    next_week = today + timedelta(days=6)

    # Query employees whose birthdays fall within the next week
    employees2 = employ_add_form.objects.filter(dateofbirth__day__gte=today.day, dateofbirth__day__lte=next_week.day, dateofbirth__month=today.month).order_by('dob')

    for employee in employees2:
        # Check if email has been sent today
        if not employee.email_sent_today:
            # Check if today is the birthday
            if employee.dob.day == today.day:
                # Send email for today's birthday
                message = f"Dear {employee.firstname},\n\nToday is your birthday! Happy birthday!\n\nBest regards,\nThe HRMS Team"
                send_mail('Happy Birthday', message, 'saipathivada1234@gmail.com', [employee.email])
            else:
                # Send email for upcoming birthday (within 1 or 2 days)
                days_to_birthday = (employee.dateofbirth - today).days
                if 0 < days_to_birthday <= 2:
                    message = f"Dear {employee.firstname},\n\nYour birthday is coming up in {days_to_birthday} days! We wish you an advanced happy birthday!\n\nBest regards,\nThe HRMS Team"
                    send_mail('Upcoming Birthday', message, 'saipathivada1234@gmail.com', [employee.email])

            # Mark email as sent for the current day
            employee.email_sent_today = True
            employee.save()

    s = employnav.objects.all()  # Make sure to replace 'employnav' with the correct model
    return render(request, 'employ-template/birthday.html', {'employees2': employees2, 's': s})


# from django.shortcuts import render
# from datetime import date, timedelta, datetime
# from.models import Employee1
# from django.core.mail import send_mail
# from django.db.models import Q
# from django.utils import timezone
# from django_crontab.crontab import Crontab
# from django.http import HttpResponse, HttpResponseRedirect

# def upcoming_birthdays(request):
#     s=employnav.objects.all()
#     today = date.today()
#     next_week = today + timedelta(days=6)
#     employees = Employee1.objects.filter(Q(birthday_date__day__gte=today.day) & Q(birthday_date__day__lte=next_week.day))
#     upcoming_birthdays = Employee1.objects.filter(birthday_date__day=today.day)
#     for employee in upcoming_birthdays:
#         # Check if an email has already been sent to this employee
#         if Employee1.objects.filter(email=employee.email, sent_att__isnull=False).exists():
#             continue
#         # If an email has not been sent, send it now
#         html_content = render_to_string("email_template.html", {'employee': employee})
#         subject = "Happy birthday to our Best employee!"
#         from_email = settings.DEFAULT_FROM_EMAIL
#         recipient_list = [employee.email]
#         message = EmailMultiAlternatives(subject, html_content, from_email, recipient_list)
#         message.content_subtype = "html"
#         message.fail_silently = False
#         message.send()
#         # Set the sent_birthday_email field for the employee
#         employee.sent_att = timezone.now()
#         employee.save()
#         # Send a separate email to the second employee
#         second_employee = Employee1.objects.filter(birthday_date__day=today.day, id__in=[employee.id for employee in upcoming_birthdays]).exclude(id=employee.id).first()
#         if second_employee:
#             html_content = render_to_string("email_template.html", {'employee': second_employee})
#             subject = "Happy birthday to our Best employee!"
#             from_email = settings.DEFAULT_FROM_EMAIL
#             recipient_list = [second_employee.email]
#             message = EmailMultiAlternatives(subject, html_content, from_email, recipient_list)
#             message.content_subtype = "html"
#             message.fail_silently = False
#             message.send()
#             # Set the sent_birthday_email field for the second employee
#             second_employee.sent_att = timezone.now()
#             second_employee.save()
#     return render(request, 'birthday.html', {'employees': employees,'upcoming_birthdays':upcoming_birthdays,})



from django.http import HttpResponse


from datetime import datetime, timedelta, date
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect
from django.views import generic
from django.urls import reverse
from django.utils.safestring import mark_safe
import calendar
import requests
from django.views.generic import TemplateView
from django.template import Context, Template

from django.db.models import Count,Q
from .utils import Calendar,caltable
from django import forms
from .models import employlevsheet,editholiday12,customholidays,publicholidays
def index(request):
    return HttpResponse('hello')

class HomeForm(forms.Form):
    name=forms.CharField(max_length=100)
    email=forms.EmailField()

class CalendarView(TemplateView):
    template_name = "employ-template/calendar.html"
    
    def get_context_data(self, **kwargs,):
        data=Employs.objects.filter(admin=self.request.user.id).first()
        data1=data.id
        data2=data.hroptions
        project_manager=data.projectmanagerop
        recipient=data.receptionist_option


        compid=data.companyid
        regname=Companys.objects.filter(id=data.companyid.id).first()
        organization_name=regname.organizationname
        teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
        # if data2:
        #     s=employnav.objects.filter(is_name_exist=1,hr_options=1)
        # elif project_manager:
        #     s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
        # elif teamleadop:
        #     s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
        # elif recipient:
        #     s=employnav.objects.filter(is_name_exist=1,receptionist_option=1)

        # else:
        #     s=employnav.objects.filter(is_name_exist=1,employ_options=1)
    


    
    

        tlop=TeamMember.objects.filter(employee=data1,is_team_lead=1)
        tloptions=employnav.objects.filter(is_name_exist=1,is_tl_option=1)



        context = super(CalendarView,self).get_context_data(**kwargs)
        d = get_date(self.request.GET.get('month', None))
        cal = Calendar(d.year, d.month,d.day)
        html_cal = cal.formatmonth(self.request,withyear=True)
        days_list,weekoff_days,weekend_count = caltable(self,d.year,d.month,self.request)
        employlevsheetcausel = employlev.objects.filter(leave_id=1,companyid=compid).first()
        employlevsheetmedical=employlev.objects.filter(leave_id=2,companyid=compid).first()
        employlevsheetearned=employlev.objects.filter(leave_id=3,companyid=compid).first()
        employs = Employs.objects.get(admin=self.request.user)
        email = data.email
        todaysec = datetime.now().date()

        check_in = checkin.objects.filter(date=todaysec, empid=email).first()
        check_out = checkout.objects.filter(date=todaysec, empid=email).first()


        workings = working_shifts.objects.filter(shift_name=employs.working12,companyid=employs.companyid).first()

        if workings:
            starting_time = workings.starting_time
            before_time = workings.befor_time  # Use the correct field name
            starting_datetime = datetime.combine(todaysec, starting_time)
            check_in_time = starting_datetime - timedelta(minutes=before_time)
            check_in_cutoff = starting_datetime + timedelta(minutes=workings.cutoff_time)
            can_check_in = check_in_time <= datetime.now() <= check_in_cutoff
            extra_time1=workings.extra_time1 

            current_month = datetime.today().date().replace(day=1)
            next_month1 = (current_month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        else:
            starting_time=None
            before_time=None
            starting_datetime=None
            check_in_time=None
            check_in_cutoff=None
            can_check_in=None
            current_month=None
            next_month1=None
            extra_time1=None
        weekholiday=editholiday12.objects.first()
        leave_report_datamedical = LeaveReportEmploy.objects.filter(Q(leave_date__range=[current_month, next_month1]) | Q(to_date__range=[current_month, next_month1]),employ_id=employs.id, leave_status=1,leave_type=2)
        leave_report_datacausal = LeaveReportEmploy.objects.filter(Q(leave_date__range=[current_month, next_month1]) | Q(to_date__range=[current_month, next_month1]),employ_id=employs.id, leave_status=1,leave_type=1)
        leave_report_dataearned = LeaveReportEmploy.objects.filter(Q(leave_date__range=[current_month, next_month1]) | Q(to_date__range=[current_month, next_month1]),employ_id=employs.id, leave_status=1,leave_type=3)
        total_leave_duration = sum([leave.get_leave_duration() for leave in leave_report_datamedical])
        total_causel=sum([leave.get_leave_duration() for leave in leave_report_datacausal])
        total_earned=sum([leave.get_leave_duration() for leave in leave_report_dataearned])
        formatted_data =  f"Medical Leave Total Duration (Days): {total_leave_duration}"
        if employlevsheetmedical:
            usedmedical=employlevsheetmedical.defaultleave-total_leave_duration
        else:
            usedmedical=None
        if employlevsheetcausel:
            usedcausel=employlevsheetcausel.defaultleave-total_causel
        else:
           usedcausel=None 
        if employlevsheetearned:
              usedearned=employlevsheetearned.defaultleave-total_earned
        else:
            usedearned=None

        today = date.today()
        is_special_weekend = customholidays.objects.filter(date=today,companyid=compid).exists()
        is_public_holiday = publicholidays.objects.filter(publicholiday_date=today,companyid=compid).exists()
        is_weekoff_today =today.strftime('%A') in weekoff_days or is_special_weekend or is_public_holiday
        context['leave_report_data'] = formatted_data
        context['calendar'] = mark_safe(html_cal)
        context['prev_month'] = prev_month(d,self.request)
        context['next_month'] = next_month(d,self.request)
        context['employlevsheetcausel'] = employlevsheetcausel
        context['employlevsheetmedical']=employlevsheetmedical
        context['employlevsheetearned']=employlevsheetearned
        context['weekholiday']=weekholiday
        context['employs']=employs
        context['formatted_data']=formatted_data
        context['usedmedical']=usedmedical
        context['usedcausel']=usedcausel
        context['usedearned']=usedearned
        context['weekoff_days']=weekoff_days
        context['is_weekoff_today']=is_weekoff_today
        context['is_public_holiday']=is_public_holiday
        context['is_special_weekend']=is_special_weekend
        # context['s']=s
        context['tlop']=tlop
        context['tloptions']=tloptions
        context['data']=data
        context['recipient']=recipient

        context['check_in']=check_in
        context['check_out']=check_out
        context['check_in_time']=check_in_time
        context['check_in_cutoff']=check_in_cutoff
        context['can_check_in']=can_check_in
        context['organization_name']=organization_name
        context['extra_time1']=extra_time1
        if check_out and not check_in:
            messages.success(self.request, 'You have been automatically checked out as absent.')

            return redirect('/calendar/')




        inline_context={
            "days_list":"days_list",
            
            
        }
        inline_html_template=Template(f"{days_list}")
        inline_view=inline_html_template.render(Context(inline_context))
        context['inline_view']=inline_view
        return  context
    def post(self, request, *args, **kwargs):
        form = HomeForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            # Do something with the name and email
            return HttpResponseRedirect(reverse('calendar'))
        else:
            context = self.get_context_data(**kwargs)
            context['form'] = form
           
            context['check_in'] = check_in(self.request)
           
            return redirect('/calendar/')
    

class SecondCalendarView(TemplateView):
    template_name = "employ-template/secondcalendar.html"
    
    def get_context_data(self, **kwargs,):
        data=Employs.objects.filter(admin=self.request.user.id).first()
        data1=data.id
        data2=data.hroptions
        project_manager=data.projectmanagerop
        compid=data.companyid
        regname=Companys.objects.filter(id=data.companyid.id).first()
        organization_name=regname.organizationname
        teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
        # if data2:
        #     s=employnav.objects.filter(is_name_exist=1,hr_options=1)
        # elif project_manager:
        #     s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
        # elif teamleadop:
        #     s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
        # else:
        #     s=employnav.objects.filter(is_name_exist=1,employ_options=1)
    


    
    

        tlop=TeamMember.objects.filter(employee=data1,is_team_lead=1)
        tloptions=employnav.objects.filter(is_name_exist=1,is_tl_option=1)



        context = super(SecondCalendarView,self).get_context_data(**kwargs)
        d = get_date(self.request.GET.get('month', None))
        cal = Calendar(d.year, d.month,d.day)
        html_cal = cal.formatmonth(self.request,withyear=True)
        days_list,weekoff_days,weekend_count = caltable(self,d.year,d.month,self.request)
        employlevsheetcausel = employlev.objects.filter(leave_id=1,companyid=compid).first()
        employlevsheetmedical=employlev.objects.filter(leave_id=2,companyid=compid).first()
        employlevsheetearned=employlev.objects.filter(leave_id=3,companyid=compid).first()
        employs = Employs.objects.get(admin=self.request.user)
        email = data.email
        todaysec = datetime.now().date()

        check_in = checkin.objects.filter(date=todaysec, empid=email).first()
        check_out = checkout.objects.filter(date=todaysec, empid=email).first()


        workings = working_shifts.objects.filter(shift_name=employs.working12,companyid=employs.companyid).first()

        if workings:
            starting_time = workings.starting_time
            before_time = workings.befor_time  # Use the correct field name
            starting_datetime = datetime.combine(todaysec, starting_time)
            check_in_time = starting_datetime - timedelta(minutes=before_time)
            check_in_cutoff = starting_datetime + timedelta(minutes=workings.cutoff_time)
            can_check_in = check_in_time <= datetime.now() <= check_in_cutoff
            current_month = datetime.today().date().replace(day=1)
            next_month1 = (current_month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        else:
            starting_time=None
            before_time=None
            starting_datetime=None
            check_in_time=None
            check_in_cutoff=None
            can_check_in=None
            current_month=None
            next_month1=None
        weekholiday=editholiday12.objects.first()
        leave_report_datamedical = LeaveReportEmploy.objects.filter(Q(leave_date__range=[current_month, next_month1]) | Q(to_date__range=[current_month, next_month1]),employ_id=employs.id, leave_status=1,leave_type=2)
        leave_report_datacausal = LeaveReportEmploy.objects.filter(Q(leave_date__range=[current_month, next_month1]) | Q(to_date__range=[current_month, next_month1]),employ_id=employs.id, leave_status=1,leave_type=1)
        leave_report_dataearned = LeaveReportEmploy.objects.filter(Q(leave_date__range=[current_month, next_month1]) | Q(to_date__range=[current_month, next_month1]),employ_id=employs.id, leave_status=1,leave_type=3)
        total_leave_duration = sum([leave.get_leave_duration() for leave in leave_report_datamedical])
        total_causel=sum([leave.get_leave_duration() for leave in leave_report_datacausal])
        total_earned=sum([leave.get_leave_duration() for leave in leave_report_dataearned])
        formatted_data =  f"Medical Leave Total Duration (Days): {total_leave_duration}"
        if employlevsheetmedical:
            usedmedical=employlevsheetmedical.defaultleave-total_leave_duration
        else:
            usedmedical=None
        if employlevsheetcausel:
            usedcausel=employlevsheetcausel.defaultleave-total_causel
        else:
           usedcausel=None 
        if employlevsheetearned:
              usedearned=employlevsheetearned.defaultleave-total_earned
        else:
            usedearned=None

        today = date.today()
        is_special_weekend = customholidays.objects.filter(date=today,companyid=compid).exists()
        is_public_holiday = publicholidays.objects.filter(publicholiday_date=today,companyid=compid).exists()
        is_weekoff_today =today.strftime('%A') in weekoff_days or is_special_weekend or is_public_holiday
        context['leave_report_data'] = formatted_data
        context['calendar'] = mark_safe(html_cal)
        context['prev_month'] = prev_month(d,self.request)
        context['next_month'] = next_month(d,self.request)
        context['employlevsheetcausel'] = employlevsheetcausel
        context['employlevsheetmedical']=employlevsheetmedical
        context['employlevsheetearned']=employlevsheetearned
        context['weekholiday']=weekholiday
        context['employs']=employs
        context['formatted_data']=formatted_data
        context['usedmedical']=usedmedical
        context['usedcausel']=usedcausel
        context['usedearned']=usedearned
        context['weekoff_days']=weekoff_days
        context['is_weekoff_today']=is_weekoff_today
        context['is_public_holiday']=is_public_holiday
        context['is_special_weekend']=is_special_weekend
        # context['s']=s
        context['tlop']=tlop
        context['tloptions']=tloptions
        context['data']=data

        context['check_in']=check_in
        context['check_out']=check_out
        context['check_in_time']=check_in_time
        context['check_in_cutoff']=check_in_cutoff
        context['can_check_in']=can_check_in
        context['organization_name']=organization_name



        inline_context={
            "days_list":"days_list",
            
            
        }
        inline_html_template=Template(f"{days_list}")
        inline_view=inline_html_template.render(Context(inline_context))
        context['inline_view']=inline_view
        return  context
    def post(self, request, *args, **kwargs):
        form = HomeForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            # Do something with the name and email
            return HttpResponseRedirect(reverse('secondcalendar'))
        else:
            context = self.get_context_data(**kwargs)
            context['form'] = form
           
            context['check_in'] = check_in(self.request)
           
            return redirect('/secondcalendar/')
 
from.models import employlev,halfldayvreason
def employ_apply_leave(request):

    
    employs = Employs.objects.all()
    hr=HR.objects.all()
    emp = CustomUser.objects.filter(user_type=2,is_active=True).count()  # Count active employees
    dis = CustomUser.objects.filter(user_type=2,is_active=False).count()  # Count inactive employees



    staff_obj = Employs.objects.get(admin=request.user.id)
    leave_data=payslip_request.objects.filter(student_id=staff_obj)
  

    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    half=halfldayvreason.objects.filter(companyid=compid).first()
    half1=employlev.objects.filter(companyid=compid)
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)



    staff_obj = Employs.objects.get(admin=request.user.id)
    leave_data=LeaveReportEmploy.objects.filter(employ_id=staff_obj).order_by('-created_at')
    if request.method == 'POST':
        status_filter = request.POST.get('status')
        leave_type_filter = request.POST.get('leave_type')
        
        # Apply filters
        if status_filter or leave_type_filter:
            leave_data = leave_data.filter(Q(leave_status=status_filter) | Q(leave_type__iexact=leave_type_filter))  
    items_per_page = 10

    page = request.GET.get('page', 1)

    paginator = Paginator(leave_data, items_per_page)

    try:
        leave_data_page = paginator.page(page)
    except PageNotAnInteger:
        leave_data_page = paginator.page(1)
    except EmptyPage:
        leave_data_page = paginator.page(paginator.num_pages)


    return render(request,"employ-template/employ_apply_leave.html",{"leave_data":leave_data,'half':half,'half1':half1,"leave_data": leave_data_page,'data':data,'employs':employs,'hr':hr,'emp':emp,'dis':dis
})



    
def get_date(req_month):
    if req_month:
        year, month = (int(x) for x in req_month.split('-'))
        curr_day = datetime.now().day
        curr_year = datetime.now().year
        curr_month = datetime.now().month
        return date(year, month, day=1)

    return datetime.today()

def prev_month(d,request):
    first = d.replace(day=1)
    prev_month = first - timedelta(days=1)
    month = 'month=' + str(prev_month.year) + '-' + str(prev_month.month)
    request.session['year']=str(prev_month.year);
    return month

def next_month(d,request):
    days_in_month = calendar.monthrange(d.year, d.month)[1]
    last = d.replace(day=days_in_month)
    next_month = last + timedelta(days=1)
    month = 'month=' + str(next_month.year) + '-' + str(next_month.month)
    request.session['month']=str(next_month.month);
    request.session['year']=str(next_month.year);
    return month



def sidenav(request):
    s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
    return render(request,'employ-template/calendar.html',{})

        

# def check_in(request):
#     today = datetime.now().date()
#     now = datetime.now().time()
#     stat = "present"
#     email = request.session.get('email_2')

#     if request.method == 'POST' and 'vk' in request.POST:
#         check = checkin.objects.filter(date=today, empid=email).first()
#         if not check:
#             checkin.objects.create(date=today, time=now, status=stat, empid=email)
#             formatted_time = now.strftime("%I:%M %p")  # Format time as HH:MM AM/PM
#             messages.success(request,f'You have successfully checked in at {formatted_time}.')
#             return redirect('/calendar/')
#         else:
#             messages.warning(request,'You have already checked in today.')
#             return redirect('/calendar/')


#     if request.method == 'POST' and 'vk1' in request.POST:
#         checkt = checkout.objects.filter(date=today, empid=email).first()
#         if not checkt:
#             checkout.objects.create(date=today, time=now, empid=email,date_value=1)
#             formatted_time = now.strftime("%I:%M %p")  # Format time as HH:MM AM/PM
#             messages.success(request, f'You have successfully checked out at  {formatted_time}.')
#             return redirect('/calendar/')

#         else:
#             messages.warning(request, 'You have already checked out today.')
#             return redirect('/calendar/')

#     return render(request, 'employ-template/calendar.html')


from datetime import datetime, timedelta
from django.contrib import messages
from django.shortcuts import redirect, render
from .models import working_shifts, checkin, checkout, Employs

def check_in(request):
    user = request.user
    employs = Employs.objects.get(admin=user)
    today = datetime.now().date()
    now = datetime.now().time()
    stat="present"
    email =employs.email

    # Fetch the working shift data based on 'employs.working12' value
    workings = working_shifts.objects.filter(shift_name=employs.working12,companyid=employs.companyid).first()
    extra_time_slots = ExtraTimeSlot.objects.filter(employees=employs)


    if workings:
        if request.method == 'POST':
            if 'vk' in request.POST:
                # Check if the user has already checked in
                check_in = checkin.objects.filter(date=today, empid=email).first()
                if not check_in:
                    # Calculate the check-in and cutoff times
                    starting_time = workings.starting_time
                    before_time = workings.befor_time  # Use the correct field name
                    starting_datetime = datetime.combine(today, starting_time)
                    check_in_time = starting_datetime - timedelta(minutes=before_time)
                    check_in_cutoff = starting_datetime + timedelta(minutes=workings.cutoff_time)

                    if check_in_time <= datetime.now() <= check_in_cutoff:
                        checkin.objects.create(date=today, time=now,status=stat, empid=email, shift_name=workings.shift_name,companyid=employs.companyid,is_employee="1")
                        formatted_time = now.strftime("%I:%M %p")  # Format time as HH:MM AM/PM
                        messages.success(request, f'You have successfully checked in at {formatted_time}.')
                    else:
                        messages.warning(request, 'It is too early/late to check in.')
                else:
                    messages.warning(request, 'You have already checked in today.')

            if 'vk1' in request.POST:
                # Check if the user has already checked out
                check_in = checkin.objects.filter(date=today, empid=email).first()
                check_out = checkout.objects.filter(date=today, empid=email).first()
                if check_in and not check_out:
                    extra_time1= workings.extra_time1
                    extra_time_minutes = sum(slot.duration for slot in extra_time_slots) + extra_time1
                    ending_time = workings.ending_time
                    ending_datetime = datetime.combine(today, ending_time)
                    extra_time_datetime = ending_datetime + timedelta(minutes=extra_time_minutes)

                    if datetime.now() <= extra_time_datetime:
                        checkin_status = 'present'
                    else:
                        checkin_status = 'absent'

                    checkout.objects.create(date=today, time=now, empid=email, shift_name=workings.shift_name, date_value=1, companyid=employs.companyid, is_employee="1")
                    check_in.status = checkin_status
                    check_in.save()
                    formatted_time = now.strftime("%I:%M %p")
                    messages.success(request, f'You have successfully checked out at {formatted_time}.')
                else:
                    messages.warning(request, 'You have already checked out today.')

            return redirect('/calendar/')
    else :
        messages.warning(request, 'You are not allowed to check in or out because your shift is not valid.')
        return redirect('/calendar/')
    return render (request,'employ-template/calendar.html')



from django.shortcuts import render
from datetime import datetime
from django.db.models import Sum

from .models import  MonthlyTotal,company_details_first
from .utils import get_current_month_holidays_count

def count_business_days(year, month):
    total_days = calendar.monthrange(year, month)[1]
    business_days = 0
    
    for day in range(1, total_days + 1):
        if datetime(year, month, day).weekday() < 5:  # Monday to Friday (0 to 4)
            business_days += 1
    
    return business_days

def get_month_name(month_number):
    return calendar.month_abbr[month_number]

def emp_payslip(request):
    today = datetime.now()
    
    current_year = today.year
    current_month = today.month
    company_detail=company_details_first.objects.all()
    objs=Employs.objects.get(admin=request.user.id)
    details=employ_add_form.objects.filter(student_id=objs)
    # Filter data for the current month
    # data_for_month = CheckData.objects.filter(
    #     created_at__year=current_year,
    #     created_at__month=current_month
    # )
    data_for_months = checkout.objects.filter(
        date__year=current_year,
        date__month=current_month,
        empid=objs.email
    )
    india_holidays_count = get_current_month_holidays_count('india')
    user_obj=Employs.objects.get(admin=request.user.id)
    package = int(user_obj.package)
    tax= int(package * 0.66667)
    
    b_tax=int(tax * 0.5)
    h_tax=int(b_tax * 0.5)
    s_tax=int(b_tax * 0.3)
    l_tax=int(b_tax * 0.2)
    standard_deductions= 50000
    net_taxable_income=tax - standard_deductions
    quotient = int(package) / 12
    main_salary=quotient
    # Multiply the remainder by 0.5%
    month_salary = main_salary * 0.5
    value=(main_salary * 0.25) +  (main_salary * 0.15) + (main_salary * 0.10)
    rent=int(main_salary * 0.25)
    travel=int(main_salary * 0.15)
    leave_travel=int(main_salary * 0.10)
    # Get the number of days in the current month
    business_days_in_current_month = count_business_days(current_year, current_month)

    # Adjust the amount based on the number of days in the current month
    quotients = month_salary / business_days_in_current_month 
    try:
        monthly_total_result = data_for_months.aggregate(total=Sum('date_value'))
        monthly_total = monthly_total_result['total'] + india_holidays_count
    except TypeError:
        monthly_total = 0
    totalss = int(monthly_total * quotients)
    
    total= int(totalss + value )
    month_and_year = f"{get_month_name(current_month)} {current_year}"
    if monthly_total is None:
        monthly_total = 0
    monthly_total_obj, created = MonthlyTotal.objects.update_or_create(
        year=current_year,
        month=current_month,
        student_id=user_obj,
        month_and_year = month_and_year,
        defaults={'total': totalss}
    )
    data=MonthlyTotal.objects.get(student_id=user_obj,month=current_month)
    return render(request, 'employ-template/emp_payslip.html', {'standard_deductions':standard_deductions,'net_taxable_income':net_taxable_income,'tax':tax,'b_tax':b_tax,'h_tax':h_tax,'s_tax':s_tax,'l_tax':l_tax,'travel':travel,'leave_travel':leave_travel,'rent':rent,'details':details,'user_obj':user_obj,'data':data,'company_detail':company_detail,'total':total,'monthly_total': monthly_total_obj.total,'totalss':totalss})


from .models import TeamMember , Task , Project
def employee_record_emp(request, employee_id):
    employee = Employs.objects.get(id=employee_id)
    is_team_lead = TeamMember.objects.filter(employee=employee, is_team_lead=True).exists()
    # Get employee's overall performance data
    performance_data = employee.calculate_overall_performance()

    # Get the list of projects associated with the employee
    projects = Project.objects.filter(teammember__employee=employee)
    team_members = TeamMember.objects.filter(employee=employee)
    total_score_across_all_projects = sum(member.calculate_total_score_across_all_projects() for member in team_members)
    # Get the list of tasks associated with the employee
    tasks = Task.objects.filter(e_id=employee)

    context = {
        'employee': employee,
        'performance_data': performance_data,
        'projects': projects,
        'tasks': tasks,
        'total_score_across_all_projects':total_score_across_all_projects,
        'is_team_lead':is_team_lead
    }


    return render(request,"employ-template/employee_record.html" , context)



def team_lead_read_proj(request, teamlead_id):
    projects = Project.objects.filter(teammember__employee=teamlead_id)
    completed_count = projects.filter(status='completed').count()
    ongoing_count = projects.filter(status='ongoing').count()
    featured_count = projects.filter(status='featured').count()
    if request.method == 'GET':
        status = request.GET.get('status')

        if status == 'completed':
            projects = projects.filter(status='completed')
        elif status == 'ongoing':
            projects = projects.filter(status='ongoing')
        elif status == 'featured':
            projects = projects.filter(status='featured')
        else:
            # If the status is not specified or is invalid, show all projects
            projects = Project.objects.filter(teammember__employee=teamlead_id)

        # Count the number of projects for each status
        
    return render(request, 'employ-template/TeamLeadViewProjects.html', {'projects': projects,"status":status,"completed_count":completed_count,"ongoing_count":ongoing_count,"featured_count":featured_count})

from django.utils import timezone

def team_lead_projectwise_task(request, pid):
    
    
    
    team_members = TeamMember.objects.filter(project=pid)
    team_member_scores = [(team_member, team_member.calculate_total_score_in_project()) for team_member in team_members]
    team_member_scores = sorted(team_member_scores, key=lambda x: x[1], reverse=True)
    project_details=Project.objects.filter(id=pid)
    task_details = Task.objects.filter( id=pid).all()

    if request.method == 'GET':
        t_status = request.GET.get('t_status', 'total')  # Default to 'total' if not provided

        if t_status == 'total':
            task_details = Task.objects.filter( p_id=pid).all()
        elif t_status == 'completed':
            task_details = Task.objects.filter( p_id=pid, t_status='completed')
        elif t_status == 'in-progress':
            # task_details = Task.objects.filter( p_id=pid, t_status='in-progress').all()
            task_details = Task.objects.filter( p_id=pid).exclude(t_status='completed')

        

        tasks = task_details
        count_no_of_total_tasks = Task.objects.filter( p_id_id=pid).count()
        count_no_of_completed_tasks = Task.objects.filter( p_id_id=pid, t_status="completed").count()
        count_no_of_pending_tasks = count_no_of_total_tasks - count_no_of_completed_tasks
        task_data = []

        for task in tasks:
            now = timezone.now()
            task_score = task.calculate_task_score()
            remaining_time = task.calculate_remaining_time()
            progress_percentage = min(task_score / 10 * 100, 100)
            excess_points = max(task_score - 10, 0)  # Calculate the excess points (if any)
            neg_points = max(-task_score, 0)

            task_data.append({
                'task': task,
                'progress_percentage': progress_percentage,
                'excess_points': excess_points,
                'neg_points': neg_points,
                'task_score': task_score,
                'remaining_time': remaining_time,
            })
            
        
        context = {
        
        "project_details": project_details,
        'task_total': count_no_of_total_tasks,
        'task_completed': count_no_of_completed_tasks,
        'task_pending': count_no_of_pending_tasks,
        'team_members': team_members,
        
        't_status':t_status,
        "task_details": task_details,
        'task_data': task_data,
        
        'team_member_scores': team_member_scores,

        }
    
    return render(request, 'employ-template/TeamLeadViewProjectwiseTasks.html', context)

def team_lead_delete_user(request,employee_id,project_id):
        # Assuming your model has a field 'id' representing the user's ID
        employee = TeamMember.objects.get(employee_id=employee_id,project_id=project_id)
        employee.delete()
        referer = request.META.get('HTTP_REFERER')
        if referer:
          return redirect(referer)
        else:
          return redirect(reverse('home'))

from django.http import HttpResponse, HttpResponseRedirect, JsonResponse

def tl_get_team_members(request, project_id):
    team_members = TeamMember.objects.filter(project_id=project_id)
    data = [{'id': member.employee.id, 'name': f"{member.employee.first_name} {member.employee.last_name}"}
            for member in team_members]
    return JsonResponse({'team_members': data})

from django.core.mail import send_mail
from django.conf import settings

def team_lead_create_task(request,pid):
    
    boards = HR.objects.all()
    projects = Project.objects.filter(id=pid ).values()
    employees = Employs.objects.all()
    context = {"boards": boards, "projects": projects, "employees": employees}
    if request.method == 'POST':
        t_name = request.POST['t_name']
        t_desc = request.POST['t_desc']
        t_deadline_date = request.POST['t_deadline_date']
        t_status = "todo"
        t_priority = request.POST['t_priority']
        b_id = request.POST['b_id']
        p_id = pid
        e_id = request.POST['e_id']
        taskObj = Task.objects.create(t_name=t_name, t_desc=t_desc, t_deadline_date=t_deadline_date,
                                      t_status=t_status, t_priority=t_priority,  b_id_id=b_id, p_id_id=p_id, e_id_id=e_id)
        if taskObj:
            empDetails = Employs.objects.filter(id=e_id).values()
            subject = 'DevelopTrees - New Task Created for you'
            message = f'Hi {empDetails[0]["first_name"]} , Your organization as created a new task : {t_name} , description : {t_desc}, priority : {t_priority} and deadline for task is : {t_deadline_date}, Login in your account to get more information. From: DevelopTrees. '
            email_from = settings.EMAIL_HOST_USER
            recipient_list = [empDetails[0]["email"], ]
            send_mail(subject, message, email_from, recipient_list)
            messages.success(request, "Task was created successfully!")
            dynamic_url = reverse ('team_lead_projectwise_tasks' , args=[pid])
            return HttpResponseRedirect(dynamic_url)
        else:
            messages.error(request, "Some Error was occurred!")
            return HttpResponseRedirect('/team_lead_create_task')
    return render(request, 'employ-template/TeamLeadCreateTask.html', context)

def team_lead_delete_task(request, pk):
    try:
        # Fetch the task and its project ID
        task = Task.objects.get(id=pk)
        pid = task.p_id.id  # Assuming Task has a foreign key to Project model

        # Check if the user has the necessary permissions (you might want to customize this logic)
        
        task.delete()
            

    except Task.DoesNotExist:
        messages.error(request, "Task does not exist.")
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")

    dynamic_url = reverse('team_lead_projectwise_tasks', args=[pid])
    return HttpResponseRedirect(dynamic_url)

def team_lead_update_task(request, pk):
    # try:
    tasks = Task.objects.get(id=pk)
    p_id = tasks.p_id
    pid=tasks.p_id.id
    projects_emp_link = TeamMember.objects.filter(project_id=p_id)
     
    if request.method == 'POST':
        t_name = request.POST['t_name']
        t_desc = request.POST['t_desc']
        t_deadline_date = request.POST['t_deadline_date']
        t_priority = request.POST['t_priority']
        e_id = request.POST['e_id']
        employ_instance = Employs.objects.get(id=e_id)
        t_status = request.POST['t_status']
        tasks.t_deadline_date = datetime.strptime(t_deadline_date, '%Y-%m-%dT%H:%M')
        tasks.t_name = t_name
        tasks.t_desc = t_desc
        
        tasks.t_priority = t_priority
        tasks.t_status = t_status
        tasks.e_id = employ_instance
        tasks.save()
        if tasks:
            empDetails = Employs.objects.filter(id=e_id).values()
            subject = 'DevelopTrees - Task Updated for you'
            message = f'Hi {empDetails[0]["first_name"]} , Your organization as updated a task : {t_name} , description : {t_desc}, priority : {t_priority} and deadline for task is : {t_deadline_date}, Login in your account to get more information. From: DevelopTrees. '
            email_from = settings.EMAIL_HOST_USER
            recipient_list = [empDetails[0]["email"], ]
            send_mail(subject, message, email_from, recipient_list)
            messages.success(request, "Task was updated successfully!")
            dynamic_url = reverse('team_lead_projectwise_tasks', args=[pid])
            return HttpResponseRedirect(dynamic_url)

        else:
            messages.error(request, "Some Error was occurred!2")
            
    else:
        if tasks:
            return render(request, 'employ-template/TeamLeadUpdateTask.html', {"tasks": tasks , "projects_emp_link": projects_emp_link } )
        else:
            messages.error(request, "Some Error was occurred!1")
            
    dynamic_url = reverse('team_lead_projectwise_tasks', args=[pid])
    return HttpResponseRedirect(dynamic_url)

def employ_upload_photo(request):
    

         
    employ_obj=Employs.objects.get(admin=request.user.id)
    try:
        company = Companys.objects.get(id=employ_obj.companyid.id)  
        organization_name = company.organizationname 
    except Companys.DoesNotExist:
        organization_name = "No Organization Name Available" 

    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)



    compid=data.companyid

    # data = Employs.objects.get(admin=request.user.id)
  


    
    if request.method == "POST":
        profile_pic = request.FILES.get("profile_pic")
        if profile_pic:
           
            data.profile_pic = profile_pic
            
                # Create a new instance if it doesn't exist
            data,_ = Employs.objects.get_or_create(admin=request.user.id)
            data.profile_pic=profile_pic

            data.save()
            messages.success(request, "Profile Picture Uploaded Successfully.")
        else:
            messages.error(request, "Profile Picture Upload Failed.")
        return redirect('/Employ_home')
    return render(request, "employ-template/employprofile_pic.html", {'data':data,'data1':data1,'organization_name':organization_name, 


})
def employdelete_profile_pic(request):
    user_obj = Employs.objects.get(admin=request.user.id)
    try:
       
        user_obj.profile_pic.delete()  # Delete the profile picture file
        user_obj.save()
        messages.success(request, "Profile picture removed successfully.")
    except Employs.DoesNotExist:
        messages.error(request, "User data not found. Please add user data first.")
    return redirect('/Employ_home')
    # return redirect("employ_upload_photo")



def team_lead_update_project(request, pid):
    # Get the project to be updated
    project = get_object_or_404(Project, id=pid)
    team_leader_info=project.get_team_leader()
    
    # Get available team members excluding team leads
    teams=Employs.objects.filter(Q(teammember__isnull=True) | Q(teammember__project=project)).distinct()
    team_members = TeamMember.objects.filter(project=pid)

    if request.method == 'POST':

       
        # Retrieve updated project details from the form
        p_name = request.POST['p_name']
        p_desc = request.POST['p_desc']
        pr_deadline = request.POST['project_deadline_date']
        project_manager = request.POST['manager_name']
        status = request.POST.get('status')

        # Update the project
        project.p_name = p_name
        project.p_desc = p_desc
        project.project_deadline = pr_deadline
        project.project_manager = project_manager
        project.status = status
        project.save()  
            # Update team lead if changed
        selected_team_lead_id = request.POST.get('team_lead')
        if selected_team_lead_id:
            try:
                team_lead = Employs.objects.get(id=selected_team_lead_id)
                project.teammember_set.update(is_team_lead=False)  # Remove team leader status from existing team leader

                # Create a new TeamMember if it doesn't exist
                team_member, created = TeamMember.objects.get_or_create(project=project, employee=team_lead)
                team_member.is_team_lead = True
                team_member.save()
            except Employs.DoesNotExist:
                # Handle the case where the selected team leader doesn't exist
                messages.error(request, 'Selected team leader does not exist.')
                return redirect('/read-proj')

        # Update selected team members
        selected_employee_ids = request.POST.getlist('selected_employees[]')
        existing_team_member_ids = TeamMember.objects.filter(project=project).values_list('employee_id', flat=True)
        for emp_id in selected_employee_ids:
            emp_id = int(emp_id)
            if emp_id not in existing_team_member_ids:
                employee = get_object_or_404(Employs, id=emp_id)
                TeamMember.objects.create(project=project, employee=employee, is_team_lead=False)

        messages.success(request, 'Project updated successfully and employees assigned.')
        return redirect('/read-proj')

    # Fetch available team members (exclude team leads)
    # selected_team_lead_id = project.team_lead.id if project.team_lead else None
    return render(request, 'employ-template/team_lead_update_project.html', {"project": project,"teams":teams, "team_members": team_members,'team_leader_info':team_leader_info})

from .models import Project,TeamMember
def read_proj_hr(request):
    projects = Project.objects.all()
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')


    return render(request, 'employ-template/ViewProjects1.html',{'projects':projects,'admin_drops':admin_drops})

def reports1(request):
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname

    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)




    return render(request,"employ-template/reports.html",{'organization_name':organization_name,'regname':regname, 'data':data})

def projectwise_task_1(request, pid):
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')

    project_details=Project.objects.filter(id=pid)
    team_members = TeamMember.objects.filter(project=pid)
    team_member_scores = [(i, i.calculate_total_score_in_project()) for i in team_members]
    team_member_scores = sorted(team_member_scores, key=lambda x: x[1], reverse=True)
    return render(request, 'employ-template/ViewProjectwiseTasks1.html', {'project_details':project_details,'team_members':team_members,'team_member_scores':team_member_scores,'admin_drops':admin_drops})


from .models import projecttask,tlassigntask
from datetime import datetime
from django.db import transaction
def tlassigntask1(request):
    rty = None
    data = Employs.objects.filter(admin=request.user.id).first()
    data1 = data.id
    data2 = data.hroptions
    project_manager = data.projectmanagerop
    compid = data.companyid
    employee_name = data.first_name  
    today = datetime.today().date()
    regname = Companys.objects.filter(id=data.companyid.id).first()
    organization_name = regname.organizationname
    teamleadop = TeamMember.objects.filter(employee=data1, is_team_lead=1)

    # Fetch tasks based on user role
    if data2:
        s = employnav.objects.filter(is_name_exist=1, hr_options=1)
    elif project_manager:
        s = employnav.objects.filter(is_name_exist=1, projectmanager_options=1)
    elif teamleadop:
        s = employnav.objects.filter(is_name_exist=1, is_tl_option=1)

    team_members = TeamMember.objects.filter(is_team_lead=1, employee=data1).first()

    if team_members:
        team1 = team_members.project
        project_ids = TeamMember.objects.filter(project=team1).values_list('project_id', flat=True)
        employee_ids = TeamMember.objects.filter(project=team1).values_list('employee_id', flat=True)
        tlemp_id = TeamMember.objects.filter(project=team1, is_team_lead=1).values_list('employee_id', flat=True)
        employees = Employs.objects.filter(id__in=employee_ids).exclude(id=data1) 
        projects = Project.objects.filter(id__in=project_ids).first()
        tlid = Employs.objects.filter(id__in=tlemp_id).first()
        selected_employees = [task.employid for task in tlassigntask.objects.all()]
        tlidc = tlid.id
        project_id1 = projects.id
        # Filter projecttask by date
        today = datetime.now()
        task4 = projecttask.objects.filter(l_name=data1,task_date=today).values()

        k1 = teamtask.objects.filter(l_name=employee_name,tdate=today).values()
    else:
        # No team members found
        team1 = None
        employees = None
        projects = None
        project_id1 = None
        task4 = None
        k1 = None

    # items_per_page = 1  # Adjust the number of items per page as needed
    # paginator = Paginator(k1, items_per_page)
    # page = request.GET.get('page')

    # try:
    #     task = paginator.page(page)
    # except PageNotAnInteger:
    #     task = paginator.page(1)
    # except EmptyPage:
    #     task = paginator.page(paginator.num_pages)

    tlallready = tlassigntask.objects.filter(project_id=project_id1, task_date=today).values_list('employid', flat=True)
    taskallready = tlassigntask.objects.filter(project_id=project_id1, task_date=today)
    employalreadyin = Employs.objects.filter(id__in=tlallready)
    taskalreadyin = projecttask.objects.filter(id__in=taskallready).values_list('id', flat=True)

    if request.method == "POST":
        # Handle creating a new task
        if 'task' in request.POST and 'description' in request.POST:
            task_name = request.POST.get('task')
            description = request.POST.get('description')
            if task_name and description:
                new_task = teamtask(
                    task=task_name,
                    description=description,
                    companyid=compid,
                    l_name=employee_name
                )
                new_task.save()
                messages.success(request, "Task created successfully")
                return redirect('/tlassigntask')
            else:
                messages.error(request, "Task name and description are required.")
        


        # Handle assigning tasks to employees
        if 'project_id' in request.POST:
            project_id = request.POST.get('project_id')
            p_id = Project.objects.get(id=project_id)
            with transaction.atomic():
                employee_selected = False
                existing_employee_task = None
                for key in request.POST.keys():
                    if key.startswith('employees_'):
                        task_id = key.split('_')[1].rstrip('[]')
                        selected_employees = request.POST.getlist(key)
                        description = request.POST.get(f"description_{task_id}")
                        current_date = datetime.now().date()
                        if selected_employees:
                            employee_selected = True
                            for emp_id in selected_employees:
                                existing_employee_task = tlassigntask.objects.filter(task_id=task_id, employid=emp_id, deadline__gt=current_date).exists()
                                if not existing_employee_task:
                                    try:
                                        selected_employee = Employs.objects.get(id=emp_id)
                                        tlassigntask.objects.create(
                                            task_id=task_id, 
                                            employid=selected_employee, 
                                            project_id=p_id, 
                                            companyid=compid, 
                                            description=description,
                                            teamid=employee_name
                                        )
                                    except Employs.DoesNotExist:
                                        messages.error(request, f"Employee with id {emp_id} does not exist.")
                        if not existing_employee_task:
                            if employee_selected:
                                messages.success(request, "Task was Sent Successfully")
                            return redirect('/Employ_home')
                        else:
                            rty = "The task has already been assigned to the employee."

    success_message_displayed = request.session.pop('success_message_displayed', False)

    return render(request, "employ-template/tltaskassign.html", {
        'rty': rty,
        'taskalreadyin': taskalreadyin,
        'employalreadyin': employalreadyin,
        'tlidc': tlidc,
        'organization_name': organization_name,
        'selected_employees': selected_employees,
        'task4': task4,
        'team1': team1,
        'project_id1': project_id1,
        'employees': employees,
        'data': data,
        'k1': k1
    })

from.models import employperformance
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from.models import employperformance
def table1(request):
    employ_obj = Employs.objects.get(admin=request.user.id)
    try:
        company = Companys.objects.get(id=employ_obj.companyid.id)
        organization_name = company.organizationname
    except Companys.DoesNotExist:
        organization_name = "No Organization Name Available"

    data = Employs.objects.filter(admin=request.user.id).first()
    data1 = data.id
    data2 = data.hroptions
    project_manager = data.projectmanagerop
    teamleadop = TeamMember.objects.filter(employee=data1, is_team_lead=True)
    compid = data.companyid

    is_team_lead = teamleadop.exists()
    is_project_manager = project_manager == 1

    emm = Employs.objects.filter(admin=request.user.id).first()
    em1 = emm.companyid

    tsk1_data = employperformance.objects.filter(employ_name_id__companyid=em1)

    if is_team_lead:
        tsk1_data = tsk1_data.filter(tl_name_id=data1)
    elif is_project_manager:
        pass

    items_per_page = 10  # Adjust the number of items per page as needed
    paginator = Paginator(tsk1_data, items_per_page)
    page = request.GET.get('page')

    try:
        tsk1_data = paginator.page(page)
    except PageNotAnInteger:
        tsk1_data = paginator.page(1)
    except EmptyPage:
        tsk1_data = paginator.page(paginator.num_pages)

    if request.method == "POST":
        srem = request.POST.get("se")
        if srem:
            if is_team_lead:
                tsk1_data = employperformance.objects.filter(employ_name_id__companyid=em1, tl_name_id=data1, employ_name__first_name__icontains=srem)
            elif is_project_manager:
                tsk1_data = employperformance.objects.filter(employ_name_id__companyid=em1, employ_name__first_name__icontains=srem)

    context = {
        'tsk1_data': tsk1_data,
        "data": data,
        'data1': data1,
        'organization_name': organization_name,
    }
    return render(request, 'employ-template/emptable.html', context)


from .models import admin_project_create,Meeting

def displayed(request):
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    compid=data.companyid
    pmanager = admin_project_create.objects.filter(admin_id=request.user.id)
    datas=Employs.objects.get(admin=request.user.id)
    if request.method == 'POST':
           meeting_url = request.POST.get('meeting_url')
           team_member = request.POST.get('team_member')
           

        # Create a new project
           p1 = Meeting(
               meeting_url=meeting_url,
               team_member=team_member,
              
            )
           p1.save()
           messages.success(request, "Successfully Sent Meeting Link")

    description = request.session.get('meeting_description', "")
    url = request.session.get('meeting_url', "")

    team_members = TeamMember.objects.filter(is_team_lead=1,employee=datas).first()
    if team_members:
        team1=team_members.project
        employee_ids=TeamMember.objects.filter(project=team1,is_team_lead=0).values_list('employee_id',flat=True)
        employees=Employs.objects.filter(id__in=employee_ids)
    else:
        employees=None
        
        return redirect('/Employ_home')

    return render(request, 'employ-template/data3.html', {'datas':datas,'project_manager':project_manager,'data2':data2, 'teamleadop':teamleadop,'organization_name':organization_name, 'employees':employees, 'pmanager': pmanager,'compid':compid,  'team_members': TeamMember.objects.all(),'description': description,'url': url})


from django.db.models import Avg

def performancetask(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    compid=data.companyid
    project_manager=data.projectmanagerop
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    elif project_manager:
        s=employnav.objects.filter(is_name_exist=1,projectmanager_options=1)
    elif teamleadop:
        s=employnav.objects.filter(is_name_exist=1,is_tl_option=1)
    employ_obj=Employs.objects.get(admin=request.user.id)


    team=TeamMember.objects.filter(employee=data1,is_team_lead=1)  
    projects = Project.objects.all()  
    items_per_page = 10
    performance_data = employperformance.objects.filter(employ_name_id__companyid=em1)
    paginator = Paginator(performance_data, items_per_page)
    page = request.GET.get('page')

    try:
        performance_data = paginator.page(page)
    except PageNotAnInteger:
        performance_data = paginator.page(1)
    except EmptyPage:
        performance_data = paginator.page(paginator.num_pages)
    return render(request, 'employ-template/emptaskper.html', {'organization_name':organization_name, 'team':team,'employ_obj':employ_obj,'projects':projects,'data':data,'data1':data1, 'performance_data': performance_data})



def performanceproject(request):

    employ_obj = Employs.objects.get(admin=request.user.id)

    try:

        company = Companys.objects.get(id=employ_obj.companyid.id)

        organization_name = company.organizationname

    except Companys.DoesNotExist:

        organization_name = "No Organization Name Available"
    
    emm = Employs.objects.filter(admin=request.user.id).first()
    em1 = emm.companyid
    projects_drops = project_drop.objects.filter(parent_category=None).order_by('id')
    data = Employs.objects.filter(admin=request.user.id).first()
    data1 = data.id
    data2 = data.hroptions
    project_manager = data.projectmanagerop
    compid = data.companyid
    teamleadop = TeamMember.objects.filter(employee=data1, is_team_lead=1)
    
    # if data2:
    #     s = employnav.objects.filter(is_name_exist=1, hr_options=1)
    # elif project_manager:
    #     s = employnav.objects.filter(is_name_exist=1, projectmanager_options=1)
    # elif teamleadop:
    #     s = employnav.objects.filter(is_name_exist=1, is_tl_option=1)
    # else:
    #     s = employnav.objects.filter(is_name_exist=1, employ_options=1)

    a = company_details.objects.filter(companyid=compid).first()
    employ_obj = Employs.objects.get(admin=request.user.id)
    admin_drops = employ_drop.objects.filter(parent_category=None).order_by('id')
    team = TeamMember.objects.filter(employee=data1, is_team_lead=1)
    # projects = Project.objects.filter(o_id__companyid=em1)
    all_projects = Project.objects.filter(o_id__companyid=em1)
    unique_projects = {}
    for project in all_projects:
        if project.p_name not in unique_projects:
            unique_projects[project.p_name] = project

    projects = unique_projects.values()

    items_per_page = 10
    selected_project = None
    if request.method == 'POST':
        selected_project_id = request.POST.get('project')  # Get the selected project ID

        if selected_project_id:
            selected_project = Project.objects.get(id=selected_project_id)
            project_name = selected_project.p_name
            # Fetch the employee average performances for the selected project
            performances_for_project = employperformance.objects.filter(project_name=selected_project)
            employee_average_performances = performances_for_project.values('employ_name').annotate(average_performance=Avg('performance'))

            # Retrieve employee names and IDs
            for emp in employee_average_performances:
                employ_id = emp['employ_name']
                try:
                    employee = Employs.objects.get(id=employ_id)
                    emp['employee_name'] = employee.first_name
                    emp['employee_id'] = employee.empid
                except Employs.DoesNotExist:
                    emp['employee_name'] = "N/A"
                    emp['employee_id'] = "N/A"
        else:
            # Handle the case where there is no selected project
            performances_for_project = None
            project_name = None
            employee_average_performances = None
    else:
        performances_for_project = None
        project_name = None
        employee_average_performances = None

    if employee_average_performances is not None:
        paginator = Paginator(employee_average_performances, items_per_page)
        page = request.GET.get('page')

        try:
            employee_average_performances = paginator.page(page)
        except PageNotAnInteger:
            employee_average_performances = paginator.page(1)
        except EmptyPage:
            employee_average_performances = paginator.page(paginator.num_pages)
    else:
        # Set employee_average_performances to an empty page if it's None
        employee_average_performances = Paginator([], items_per_page).page(1)

    return render(request, 'employ-template/empprojectper.html', {
        'organization_name': organization_name,
        # 's': s,
        'team': team,
        'employ_obj': employ_obj,
        'projects': projects,
        'project_name': project_name,
        'employee_average_performances': employee_average_performances,
        'data': data,
        'data1': data1,
        'selected_project': selected_project,
    })
def performanceallproject(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid

    employ_obj=Employs.objects.get(admin=request.user.id)
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    compid=data.companyid
    project_manager=data.projectmanagerop
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)

    team=TeamMember.objects.filter(employee=data1,is_team_lead=1)
    a = company_details.objects.filter(companyid=compid).first()
    projects = Project.objects.all()  
    items_per_page = 10
    tsk1_data = employperformance.objects.filter(employ_name_id__companyid=em1).values('employ_name').annotate(avg_performance=Avg('performance'))
    employees_data = []
    for task in tsk1_data:
        employ_id = task['employ_name']
        try:
            employee = Employs.objects.get(id=employ_id)
            task['employee_name'] = employee.first_name
            task['employee_id'] = employee.empid
            project_names = employperformance.objects.filter(employ_name=employee).values('project_name__p_name')
            task['project_names'] = project_names

        except Employs.DoesNotExist:
            task['employee_name'] = "N/A"
            task['employee_id'] = "N/A"
            task['project_names'] = "N/A"
        employees_data.append(task)
      
    items_per_page = 10
    paginator = Paginator(employees_data, items_per_page)
    page = request.GET.get('page')

    try:
        employees_data = paginator.page(page)
    except PageNotAnInteger:
        employees_data = paginator.page(1)
    except EmptyPage:
        employees_data = paginator.page(paginator.num_pages)
    return render(request, 'employ-template/empavgper.html', {'organization_name':organization_name, 'employ_obj':employ_obj,'projects':projects,'data':data,'data1':data1, 'tsk1_data': employees_data,'team':team})

def reports1(request):
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    compid=data.companyid
    data2=data.hroptions
    project_manager=data.projectmanagerop
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    a = company_details.objects.filter(companyid=compid).first()
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)




    return render(request,"employ-template/reports.html",{'data':data,'organization_name':organization_name})

# from django.shortcuts import render
from django.db import models
# from.models import Employee,Department
from datetime import date, timedelta,timezone
from datetime import datetime, timedelta
from django.shortcuts import render
from django.db.models import Sum
from .models import adminnav, admin_drop, Department, Employs

def search1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    compid=data.companyid
    a = company_details.objects.filter(companyid=compid).first()

    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)



    departments = Department.objects.all()
    employees = Employs.objects.filter(companyid=em1)
    
    search_term = request.GET.get('search_term')
    designation = request.GET.get('department')
    distinct_department_designations = employees.values('designation').distinct()
    selected_date_range = request.GET.get('dateofjoining')
    
    employees_by_designation = {}

    if designation == "":  # If "All Departments" is selected
        designation_employees = employees  # Use all employees without filtering
    else:
        designation_employees = employees.filter(designation=designation)

    if selected_date_range:
        today = datetime.now().date()
        start_date = today - timedelta(days=int(selected_date_range) * 30)
        designation_employees = designation_employees.filter(dateofjoining__gte=start_date)

    if search_term:
        designation_employees = designation_employees.filter(first_name__icontains=search_term)

    total_count = designation_employees.count()
    total_package = designation_employees.aggregate(Sum('package'))['package__sum']

    employees_by_designation[designation] = designation_employees

    labels = []
    salaries = []    
    items_per_page = 10  # Adjust the number of items per page as needed
    paginator = Paginator(employees, items_per_page)
    page = request.GET.get('page')

    try:
        employees = paginator.page(page)
    except PageNotAnInteger:
        employees = paginator.page(1)
    except EmptyPage:
        employees = paginator.page(paginator.num_pages)


    context = {
        'employees': employees_by_designation,
        'total_count': total_count,
        'total_package': total_package,
        'labels': labels,
        'salaries': salaries,
        'departments': distinct_department_designations,
        'selected_date_range': selected_date_range, 
        'data':data, 
        'organization_name':organization_name,
        
    }

    return render(request, 'employ-template/variance_report.html', context)

def documentreport1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    compid=data.companyid
    a = company_details.objects.filter(companyid=compid).first()

    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)



    employ=empdocs.objects.filter(employ_id__companyid=em1).order_by("empid")
    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator(employ, items_per_page)
    page = request.GET.get('page')

    try:
        employ = paginator.page(page)
    except PageNotAnInteger:
        employ = paginator.page(1)
    except EmptyPage:
        employ = paginator.page(paginator.num_pages)
    if request.method == "POST":
        srem = request.POST.get("se")
        if srem:
            employ = empdocs.objects.filter(employ_id__companyid=em1, employ_id__id__icontains=srem)

    return render (request,"employ-template/docreport.html",{'employ':employ,'organization_name':organization_name, 'data':data})
from django.shortcuts import get_object_or_404
def delete_docreport(request,id):
    document = get_object_or_404(empdocs, id=id)
    document.delete()
    messages.success(request,"Your data was successfully deleted")
    return redirect('/documentreport1/')

import os
import zipfile
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.conf import settings
from io import BytesIO
from .models import Employs  # Import your Employs model

def download_images_zip1(request, employ_id):
    employ = get_object_or_404(Employs, id=employ_id)
      
    # Create a BytesIO stream to write the zip file.
    output = BytesIO()
    zip_file = zipfile.ZipFile(output, 'w')

    # Iterate over 'empdocs' related to the employee and add their image files to the zip.
    for empdoc in employ.empdocs_set.all():
        document_type = empdoc.documenttype1
        image_path = empdoc.imagefile.path
        image_name = os.path.basename(image_path)
        
        # Construct the relative path inside the zip file.
        relative_path = os.path.join(employ.first_name, image_name)
        zip_file.write(image_path, relative_path)

    zip_file.close()

    # Set response headers to serve the zip file.
    response = HttpResponse(output.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename={employ.first_name}_images.zip'
    return response

import os
import zipfile
from io import BytesIO
from django.http import HttpResponse
from django.conf import settings
# from .models import Employs
# import mimetypes
# def view_document1(request, employ_id):
#     employ = get_object_or_404(Employs, id=employ_id)
#     # print("employ:",employ)
    
#     empdoc = employ.empdocs_set.first()
    
#     if empdoc:
#         document_path = empdoc.imagefile.path
#         print("document_path",document_path)

#         content_type, _ = mimetypes.guess_type(document_path)

#         print("Content Type:", content_type)

#         if not content_type:
#             content_type = 'application/octet-stream'

#         with open(document_path, 'rb') as document_file:
#             document_content = document_file.read()

#         response = HttpResponse(document_content, content_type=content_type)
#         return response
#     else:
#         return HttpResponse("No document found")

import mimetypes
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from .models import Employs  # Replace 'yourapp' with the name of your Django app

def view_document1(request, employ_id):    
    employ = get_object_or_404(Employs, id=employ_id)
    
    empdoc = employ.empdocs_set.first()
    
    if empdoc:
        document_path = empdoc.imagefile.path

        try:
            with open(document_path, 'rb') as document_file:
                document_content = document_file.read()
        except FileNotFoundError:
            return HttpResponse("Error: Document file not found", status=404)

        content_type, _ = mimetypes.guess_type(document_path)

        if not content_type:
            content_type = 'application/octet-stream'

        response = HttpResponse(document_content, content_type=content_type)
        return response
    else:
        return HttpResponse("No document found")

def download_all_employee_data1(request):
    # Create a BytesIO stream to write the zip file.
    output = BytesIO()
    zip_file = zipfile.ZipFile(output, 'w')

    # Directory structure for the zip file.
    base_dir = os.path.join(settings.MEDIA_ROOT, 'employees')

    # Iterate over all employees.
    for employ in Employs.objects.all():
        employee_dir = os.path.join(base_dir, employ.first_name)

        # Add employee documents to the zip file.
        for empdoc in employ.empdocs_set.all():
            image_path = empdoc.imagefile.path
            image_name = os.path.basename(image_path)
            relative_path = os.path.join(employ.first_name, image_name)
            zip_file.write(image_path, relative_path)

    zip_file.close()

    # Set response headers to serve the zip file.
    response = HttpResponse(output.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=all_employee_data.zip'
    return response
from django.shortcuts import render
from datetime import date, timedelta
from .models import Employs, checkin, LeaveReportEmploy, editholiday12, customholidays, publicholidays



def paid1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    compid=data.companyid
    a = company_details.objects.filter(companyid=compid).first()
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    results = []
    # Query your data with date filtering
    employees = Employs.objects.filter(companyid=em1).order_by("empid")
    # current_day=datetime.now().day
    current_month = datetime.now().month
    current_year = datetime.now().year

    _, num_days = calendar.monthrange(current_year, current_month)
    start_date = datetime(current_year, current_month, 1)
    
    days_of_month = [start_date + timedelta(days=i) for i in range(num_days)]

    checkin_statuses = {}  # A dictionary to store check-in status for each day

    # Iterate through each employee
    for employee in employees:
        empid = employee.empid
        first_name = employee.first_name

        # Calculate the current month
        today = date.today()
        current_month = today.month
        current_year = today.year

        last_day_of_month = date(current_year, current_month, 1)
        last_day_of_month = last_day_of_month.replace(day=28)
        while last_day_of_month.month == current_month:
            last_day_of_month += timedelta(days=1)
        last_day_of_month -= timedelta(days=1)

        # Calculate the number of days in the current month
        days_in_month = (last_day_of_month - date(current_year, current_month, 1)).days + 1

        # Query your database for admin-configured holidays
        editholidays = editholiday12.objects.first()
        admin_holidays = sum(getattr(editholidays, day, False) or 0 for day in ['sun', 'sat1', 'sat2', 'sat3', 'sat4', 'sat5', 'alsat', 'mon', 'tue', 'web', 'thu', 'fri'])

        # Query your database for custom holidays
        custom_holidays = customholidays.objects.filter(date__month=current_month).count()

        # Query your database for public holidays
        public_holidays = publicholidays.objects.filter(publicholiday_date__month=current_month).count()

        # Calculate the remaining days
        remaining_days = days_in_month - admin_holidays - custom_holidays - public_holidays

        # Calculate the total check-ins for this employee (based on email in the checkin table)
        total_checkins = checkin.objects.filter(empid=employee.email, date__month=current_month).count()

        # Query for leaves for this employee in the current month
        half_leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Half-Day").count()
        unpaid_leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Unpaid_leave").count()
        leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Leave").count()
        open_request = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=0, leave_date__month=current_month).count()

        checkin_status_for_employee = []  # Store check-in status for this employee
       
        # Iterate through each day in the month
        for day in days_of_month:
            # Check if there's a check-in record for this employee on this day
            checkin_exists = checkin.objects.filter(
                empid=employee.email,  # Match email
                date=day
            ).exists()

            checkin_status_for_employee.append('Present' if checkin_exists else '-NA-')

        checkin_statuses[employee] = checkin_status_for_employee


        results.append({
            'empid': empid,
            
            'first_name': first_name,
            'total_checkins': total_checkins,
            'remaining_days': remaining_days,
            'original_working_days': remaining_days - total_checkins,
            'half_leave': half_leave,
            'unpaid_leave': unpaid_leave,
            'leave': leave,
            'open_request': open_request,
           
           
        })
    items_per_page = 10
    page = request.GET.get('page')
    paginator = Paginator(results, items_per_page)

    try:
        results = paginator.page(page)
    except PageNotAnInteger:
        results = paginator.page(1)
    except EmptyPage:
        results = paginator.page(paginator.num_pages)
          

    return render(request, "employ-template/paid.html", {
        'days_of_month': days_of_month,
        'employ_id': employees,
        'checkin_statuses': checkin_statuses,
        'results': results,
        
        
        'data':data,
        'organization_name':organization_name
    })
from datetime import datetime, timedelta
from calendar import monthrange
from django.shortcuts import render
from .models import Employs, checkin

def paid3(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    a = company_details.objects.filter(companyid=compid).first()
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)


    employees = Employs.objects.filter(companyid=em1)
   
    # Get start_date and end_date from query parameters (if provided)
    start_date_param = request.GET.get('start_date')
    end_date_param = request.GET.get('end_date')
    checkin_statuses = {}
    
    # Determine today's date
    today = datetime.now()

    # Calculate start_date and end_date based on provided parameters or default to the current month
    if start_date_param and end_date_param:
        start_date = datetime.strptime(start_date_param, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_param, '%Y-%m-%d')
    else:
        current_month = today.month
        current_year = today.year
        _, num_days = monthrange(current_year, current_month)
        start_date = datetime(current_year, current_month, 1)
        end_date = datetime(current_year, current_month, num_days)
    
    # Create a list of dates in the date range
    date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

    # Categorize dates into "previous," "present," or "future"
    previous_dates = [date for date in date_range if date < today]
    present_dates = [date for date in date_range if date == today]
    future_dates = [date for date in date_range if date > today]
    items_per_page = 3
    page = request.GET.get('page')
    paginator = Paginator(employees, items_per_page)

    try:
        employees_page = paginator.page(page)
    except PageNotAnInteger:
        employees_page = paginator.page(1)
    except EmptyPage:
        employees_page = paginator.page(paginator.num_pages)
    for employee in employees_page:
        checkin_status_for_employee = []

        for day in date_range:
            checkin_exists = checkin.objects.filter(
                empid=employee.email,
                date=day
            ).exists()
            checkin_status_for_employee.append('Present' if checkin_exists else '-NA-')

        checkin_statuses[employee] = checkin_status_for_employee

    # Determine the category of the date range
    if start_date < today and end_date < today:
        date_category = "Previous"
    elif start_date > today and end_date > today:
        date_category = "Future"
    else:
        date_category = "Present"

    return render(request, "employ-template/paid2.html", {
        'date_category': date_category,
        'date_range': date_range,
        'employ_id': employees,
        'checkin_statuses': checkin_statuses,
        'previous_dates':previous_dates,
        'present_dates':present_dates,
        'future_dates':future_dates,
        'checkin_statuses_page': employees_page,

        
        'data':data,
        
        'organization_name':organization_name,

    })



def balance1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid

    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    a = company_details.objects.filter(companyid=compid).first()
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)

    results = []
   

    # Query your data with date filtering
    employees = Employs.objects.filter(companyid=em1)
    
    # current_day=datetime.now().day
    current_month = datetime.now().month
    current_year = datetime.now().year

    _, num_days = calendar.monthrange(current_year, current_month)
    start_date = datetime(current_year, current_month, 1)
    
    days_of_month = [start_date + timedelta(days=i) for i in range(num_days)]
    for employee in employees:
        empid = employee.empid
        first_name = employee.first_name

        # Calculate the current month
        today = date.today()
        current_month = today.month
        current_year = today.year

        last_day_of_month = date(current_year, current_month, 1)
        last_day_of_month = last_day_of_month.replace(day=28)
        while last_day_of_month.month == current_month:
            last_day_of_month += timedelta(days=1)
        last_day_of_month -= timedelta(days=1)

        # Calculate the number of days in the current month
        days_in_month = (last_day_of_month - date(current_year, current_month, 1)).days + 1
   

        # Calculate the number of days in the current month
        half_leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Half-Day").count()
        unpaid_leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Unpaid_leave").count()
        leave = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=1, leave_date__month=current_month, leave_type="Leave").count()
        open_request = LeaveReportEmploy.objects.filter(employ_id=employee, leave_status=0, leave_date__month=current_month).count()
       


        # Add information to the results
        results.append({
            'empid': empid,
            'first_name': first_name,
            'half_leave': half_leave,
            'unpaid_leave': unpaid_leave,
            'leave': leave,
            'open_request': open_request,
            
            'data':data,
            
            'organization_name':organization_name,

           
            
        })
    items_per_page = 4 # Adjust the number of items per page as needed

    paginator = Paginator(results, items_per_page)
    page = request.GET.get('page')

    try:
        results = paginator.page(page)
    except PageNotAnInteger:
        results = paginator.page(1)
    except EmptyPage:
        results = paginator.page(paginator.num_pages)

    return render(request, 'employ-template/balance.html', {'results': results})


from datetime import datetime, timedelta
from calendar import monthrange
from django.shortcuts import render
from .models import Employs, checkin

def opeanrequest1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    a = company_details.objects.filter(companyid=compid).first()
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)


    employees = Employs.objects.filter(companyid=em1)
   
    # Get start_date and end_date from query parameters (if provided)
    start_date_param = request.GET.get('start_date')
    end_date_param = request.GET.get('end_date')
    checkin_statuses = {}
    
    # Determine today's date
    today = datetime.now()

    # Calculate start_date and end_date based on provided parameters or default to the current month
    if start_date_param and end_date_param:
        start_date = datetime.strptime(start_date_param, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_param, '%Y-%m-%d')
    else:
        current_month = today.month
        current_year = today.year
        _, num_days = monthrange(current_year, current_month)
        start_date = datetime(current_year, current_month, 1)
        end_date = datetime(current_year, current_month, num_days)
    
    # Create a list of dates in the date range
    date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

    # Categorize dates into "previous," "present," or "future"
    previous_dates = [date for date in date_range if date < today]
    present_dates = [date for date in date_range if date == today]
    future_dates = [date for date in date_range if date > today]
    items_per_page = 2
    page = request.GET.get('page')
    paginator = Paginator(employees, items_per_page)

    try:
        employees_page = paginator.page(page)
    except PageNotAnInteger:
        employees_page = paginator.page(1)
    except EmptyPage:
        employees_page = paginator.page(paginator.num_pages)
    for employee in employees_page:
        checkin_status_for_employee = []

        for day in date_range:
            checkin_exists = checkin.objects.filter(
                empid=employee.email,
                date=day
            ).exists()
            checkin_status_for_employee.append('Present' if checkin_exists else '-NA-')

        checkin_statuses[employee] = checkin_status_for_employee

    # Determine the category of the date range
    if start_date < today and end_date < today:
        date_category = "Previous"
    elif start_date > today and end_date > today:
        date_category = "Future"
    else:
        date_category = "Present"

    return render(request, "employ-template/openanrequest.html", {
        'date_category': date_category,
        'date_range': date_range,
        'employ_id': employees,
        'checkin_statuses': checkin_statuses,
        'previous_dates':previous_dates,
        'present_dates':present_dates,
        'future_dates':future_dates,
        
        'checkin_statuses_page': employees_page,

        'data':data,
        'organization_name':organization_name,
        

    })    

from django.core.mail import send_mail
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Employs, empdocs, employ_add_form
from django.http import Http404

from django.core.exceptions import ObjectDoesNotExist

def all_employees_missing_info1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid
    a = company_details.objects.filter(companyid=compid).first()
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)


    if request.method == 'POST':
        selected_employee_ids = request.POST.getlist('selected_employees')

        # Loop through the selected employee IDs and send emails
        for employee_id in selected_employee_ids:
            try:
                employee = Employs.objects.get(id=employee_id)
                employee_email = employee.email
                try:
                    employ_add_form_info = employ_add_form.objects.get(student_id=employee)
                except employ_add_form.DoesNotExist:
                    employ_add_form_info = None

                empdocs_info = empdocs.objects.filter(employ_id=employee)

                # Compose the email message with missing information
                missing_info = []
                if employee.missing_info():
                    missing_info.append(", ".join(employee.missing_info().keys()))
                if employ_add_form_info and employ_add_form_info.missing_info():
                    missing_info.append(", ".join(employ_add_form_info.missing_info().keys()))
                for empdoc in empdocs_info:
                    if empdoc.missing_info():
                        missing_info.append(", ".join(empdoc.missing_info().keys()))
                        
                if missing_info:
                    subject = f"Mr/Miss {employee.first_name}.{employee.last_name} Please Update Your Missing Information."
                    message = f"Dear {employee.first_name} .{employee.last_name},\n\n"
                    message += f"The following information is missing or incomplete:\n"
                    message += "\n".join(missing_info)
                    # message += "\n\nPlease Login to DevelopTrees HRMS to update your information as soon as possible."
                    message += "\n\nPlease click the following link to update your information as soon as possible.\n"
                    # Generate the profile URL using the employee's ID
                    # message += "\n"
                    # profile_url = reverse('employ_profile', kwargs={'employee_id': employee_id})
                    # message += request.build_absolute_uri(profile_url)

                    from_email = settings.EMAIL_HOST_USER  # Replace with your email address
                    recipient_list = [employee.email]  # Send email to the employee's email address

                    send_mail(subject, message, from_email, recipient_list)

            except Employs.DoesNotExist:
                pass  # Handle the case where an employee does not exist

        return redirect('all_employees_missing_info1')

    employees = Employs.objects.filter(companyid=em1)
    employee_data = []

    for employee in employees:
        try:
            employ_add_form_info = employ_add_form.objects.get(student_id=employee)
        except ObjectDoesNotExist:
            employ_add_form_info = None

        employee_info = {
            'employs': employee,
            'empdocs_info': empdocs.objects.filter(employ_id=employee),
            'employ_add_form_info': employ_add_form_info,
        }
        employee_data.append(employee_info)
    items_per_page = 10  # Adjust the number of items per page to 10
    paginator = Paginator(employee_data, items_per_page)  #

    page = request.GET.get('page')
    try:
        employee_data = paginator.page(page)
    except PageNotAnInteger:
        employee_data = paginator.page(1)
    except EmptyPage:
        employee_data = paginator.page(paginator.num_pages)



    context = {
        'employee_data': employee_data,
        
        
        
        'data':data,
        'organization_name':organization_name,
        
    }

    return render(request, 'employ-template/all_employees_missing_info1.html', context)

def all_employees_reimbursement_report1(request):
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    compid=data.companyid
    a = company_details.objects.filter(companyid=compid).first()
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)

    
    staff_obj = Employs.objects.all()

   
    k = reimbursementsetup1.objects.filter(companyid=compid)

    st4 = request.POST.get("ss")
    st3 = request.POST.get("vk")
    st = request.POST.get("d1")
    st1 = request.POST.get("d2")
    leave_data = Reimbursement.objects.filter(employ_id__companyid=em1).order_by("employ_id")
    # Apply filters based on user input
    if st4 and st4 != '----Select----':  # Check if a valid status is selected
        leave_data = leave_data.filter(reimbursement_status=st4)
    if st3:  # If "Select Type" is selected
        leave_data = leave_data.filter(typea__icontains=st3)
    if st and st1:
        leave_data = leave_data.filter(date__range=[st, st1])

    total_approved = leave_data.filter(reimbursement_status=1).aggregate(Sum('amount'))['amount__sum'] or 0
    total_pending = leave_data.filter(reimbursement_status=0).aggregate(Sum('amount'))['amount__sum'] or 0
    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator(leave_data, items_per_page)
    page = request.GET.get('page')

    try:
        leave_data = paginator.page(page)
    except PageNotAnInteger:
        leave_data = paginator.page(1)
    except EmptyPage:
        leave_data = paginator.page(paginator.num_pages)

    return render(request, "employ-template/all_employees_reimbursement_report.html", {
        'leave_data': leave_data,
 
        'k': k,
        'total': total_approved,
        'total1': total_pending,
        'staff_obj':staff_obj,
        
        'data':data,
        'organization_name':organization_name,
        
    })
from django.db import connection
from.models import companylogo
# def yearmonth_uploaded1(request):
#     data=Employs.objects.filter(admin=request.user.id).first()
# 
#     s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)


#     a=companylogo.objects.all()
#     ad=ad_salary.objects.first()
#     b=Reimbursement.objects.first()
#     x=connection.cursor()
#     x.execute("SELECT ehrms_employs.empid,ehrms_employs.first_name,ehrms_employs.package,ehrms_employs.dateofjoining,ehrms_checkin.status,ehrms_reimbursement.reimbursement_status,ehrms_employs.id FROM ehrms_employs INNER JOIN ehrms_checkin on ehrms_employs.email=ehrms_checkin.empid INNER JOIN ehrms_reimbursement on ehrms_employs.id=ehrms_reimbursement.employ_id_id;")
#     rs=x.fetchall()  
#     return render(request,"employ-template/salaryreport.html",{'data':data,'rs':rs,'b':b,'ad':ad})
def yearmonth_uploaded1(request):
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    comp_id=data.companyid
    a=company_details.objects.filter(companyid=comp_id).first()
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)


   
    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid

    ad=ad_salary.objects.first()
    b=Reimbursement.objects.first()

   
   
   
    rst = Employs.objects.filter(companyid=em1).order_by("empid")
    salarper=salary_struct.objects.filter(companyid=em1)
    salarded=salary_deductions.objects.filter(companyid=em1)
    salbasic=salary_struct.objects.filter(companyid=em1,salarycomponent="Basic Salary")
    salaryexbasic=salary_struct.objects.filter(companyid=em1).exclude(salarycomponent="Basic Salary")
    alowsum=salary_struct.objects.filter(companyid=em1).exclude(salarycomponent="Basic Salary").aggregate(Sum('percentageofCTC'))['percentageofCTC__sum']
    deduct=salary_deductions.objects.filter(companyid=em1).exclude(percentageordfixed="Fixed").aggregate(Sum('percentageofdctc'))['percentageofdctc__sum']
    deductfix=salary_deductions.objects.filter(companyid=em1,percentageordfixed="Fixed").aggregate(Sum('percentageofdctc'))['percentageofdctc__sum']
   
    page = request.GET.get('page', 1)
    items_per_page = 10  # Adjust the number of items per page as needed
    paginator = Paginator(rst, items_per_page)
    rst = Employs.objects.filter(companyid=em1)
    salarper=salary_struct.objects.filter(companyid=em1)
    salarded=salary_deductions.objects.filter(companyid=em1)
    salbasic=salary_struct.objects.filter(companyid=em1,salarycomponent="Basic Salary")
    salaryexbasic=salary_struct.objects.filter(companyid=em1).exclude(salarycomponent="Basic Salary")
    alowsum=salary_struct.objects.filter(companyid=em1).exclude(salarycomponent="Basic Salary").aggregate(Sum('percentageofCTC'))['percentageofCTC__sum']
    deduct=salary_deductions.objects.filter(companyid=em1).exclude(percentageordfixed="Fixed").aggregate(Sum('percentageofdctc'))['percentageofdctc__sum']
    deductfix=salary_deductions.objects.filter(companyid=em1,percentageordfixed="Fixed").aggregate(Sum('percentageofdctc'))['percentageofdctc__sum']

    try:
        rst = paginator.page(page)
    except PageNotAnInteger:
        rst = paginator.page(1)
    except EmptyPage:
        rst = paginator.page(paginator.num_pages)
    st=request.POST.get('sai')
    if st :
        rst=Employs.objects.filter(empid__contains=(st.capitalize())).values()
        return render(request,"employ-template/salaryreport.html",{'organization_name':organization_name,'salarper':salarper,'salarded':salarded,'salbasic':salbasic,'salaryexbasic':salaryexbasic,'alowsum':alowsum,'deduct':deduct,'deductfix':deductfix,'rst':rst,'b':b,'ad':ad,'data':data,'paginator': paginator})

    return render(request,"employ-template/salaryreport.html",{'organization_name':organization_name,'salarper':salarper,'salarded':salarded,'salbasic':salbasic,'salaryexbasic':salaryexbasic,'alowsum':alowsum,'deduct':deduct,'deductfix':deductfix,'rst':rst,'b':b,'ad':ad,'data':data,'paginator': paginator})


# def masterctcs1(request):
# 
#     s=employnav.objects.filter(is_name_exist=1,is_tl_option=0)
#     data=Employs.objects.filter(admin=request.user.id).first()

#     a=companylogo.objects.all()
#     k=connection.cursor()
#     k.execute("SELECT ehrms_employs.empid,ehrms_customuser.first_name,ehrms_customuser.email,ehrms_employs.role,ehrms_employs.location,ehrms_employs.status,ehrms_employs.dateofjoining,ehrms_employs.dateofbirth,ehrms_employs.package FROM ehrms_employs INNER JOIN ehrms_customuser ON ehrms_employs.id=ehrms_customuser.id;")

#     rs=k.fetchall()
#     return render(request,'employ-template/masterctc.html',{'rs':rs,'data':data})

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

def masterctcs1(request):
    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    comp_id=data.companyid
    a=company_details.objects.filter(companyid=comp_id).first()
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)



    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    rst = Employs.objects.filter(companyid=em1).order_by("empid")

    # Pagination
    page = request.GET.get('page', 1)
    items_per_page = 10
    paginator = Paginator(rst, items_per_page)
    rst = Employs.objects.filter(companyid=em1)
    salarper=salary_struct.objects.filter(companyid=em1)
    salarded=salary_deductions.objects.filter(companyid=em1)
    salbasic=salary_struct.objects.filter(companyid=em1,salarycomponent="Basic Salary")
    salaryexbasic=salary_struct.objects.filter(companyid=em1).exclude(salarycomponent="Basic Salary")
    alowsum=salary_struct.objects.filter(companyid=em1).exclude(salarycomponent="Basic Salary").aggregate(Sum('percentageofCTC'))['percentageofCTC__sum']
    deduct=salary_deductions.objects.filter(companyid=em1).exclude(percentageordfixed="Fixed").aggregate(Sum('percentageofdctc'))['percentageofdctc__sum']
    deductfix=salary_deductions.objects.filter(companyid=em1,percentageordfixed="Fixed").aggregate(Sum('percentageofdctc'))['percentageofdctc__sum']
   

    try:
        rst = paginator.page(page)
    except PageNotAnInteger:
        rst = paginator.page(1)
    except EmptyPage:
        rst = paginator.page(paginator.num_pages)
    if request.method == "POST":
        srem = request.POST.get("se")
        if srem:
            rst =Employs.objects.filter(companyid=em1, empid__icontains=srem)

    return render(request, 'employ-template/masterctc.html', {'organization_name':organization_name, 'data': data, 'rst': rst,'a': a,
                                                              'salarper':salarper,'salarded':salarded,'salbasic':salbasic,'salaryexbasic':salaryexbasic,
                                                              'alowsum':alowsum,'deduct':deduct,'deductfix':deductfix})





def register2(request):

    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    data2=data.hroptions
    project_manager=data.projectmanagerop
    comp_id=data.companyid
    a=company_details.objects.filter(companyid=comp_id).first()
    regname=Companys.objects.filter(id=data.companyid.id).first()
    organization_name=regname.organizationname
    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)

    rts=Employs.objects.filter(companyid=regname).order_by("empid")
    query = request.GET.get('query')
    # x=connection.cursor()
    # x.execute("SELECT ehrms_employs.id,ehrms_employs.empid,ehrms_employs.first_name,ehrms_employs.last_name,ehrms_employs.email,ehrms_employ_add_form.phno,ehrms_employs.gender,ehrms_employ_add_form.pan,ehrms_employs.dateofjoining,ehrms_employs.role FROM ehrms_employs INNER JOIN ehrms_employ_add_form ON ehrms_employs.email=ehrms_employ_add_form.email2;")
    # rs=x.fetchall()
    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator(rts, items_per_page)
    page = request.GET.get('page')

    try:
        rts = paginator.page(page)
    except PageNotAnInteger:
        rts = paginator.page(1)
    except EmptyPage:
        rts = paginator.page(paginator.num_pages)
    if request.method == "POST":
        srem = request.POST.get("se")
        if srem:
            rts =Employs.objects.filter(companyid=regname, empid__icontains=srem)
    return render(request, 'employ-template/register_data.html', {'data':data,'organization_name':organization_name, 'rts': rts,'query':query})




from .models import  CustomUser, Employs,Progress,admin_home_drop
def add_employ(request):
    
    employ_obj=Employs.objects.get(admin=request.user.id)
    try:
        company = Companys.objects.get(id=employ_obj.companyid.id)  
        organization_name = company.organizationname 
    except Companys.DoesNotExist:
        organization_name = "No Organization Name Available"  

    emm=Employs.objects.filter(admin=request.user.id).first()
    em1=emm.companyid
    li=list.objects.all()

    data=Employs.objects.filter(admin=request.user.id).first()
    data1=data.id
    plantype=data.companyid.plantype
    data2=data.hroptions
    project_manager=data.projectmanagerop
    compid=data.companyid

    regname=Companys.objects.filter(id=data.companyid.id).first()

    teamleadop= TeamMember.objects.filter(employee=data1,is_team_lead=1)
    aa=employ_designation3.objects.filter(companyid=compid)



    existing_user_email = None
    existing_user_username = None
    existing_user_empid = None

    if request.method == "POST":
        email = request.POST.get("email")
        username = request.POST.get("username")
        empid = request.POST.get("empid")

        if email:
            try:
                existing_user_email = CustomUser.objects.get(email=email)
                messages.info(request, 'Email already exists')
                return redirect("/add_employ")  # Redirect here to avoid processing the rest of the view
            except ObjectDoesNotExist:
                existing_user_email = None
            except MultipleObjectsReturned:
                messages.error(request, ' Email already exists.')
                return redirect("/add_employ")  # Redirect here to avoid processing the rest of the view

        if username:
            try:
                existing_user_username = CustomUser.objects.get(username=username)
                messages.error(request, 'Username already exists')
                return redirect("/add_employ")  # Redirect here to avoid processing the rest of the view
            except ObjectDoesNotExist:
                existing_user_username = None
            except MultipleObjectsReturned:
                messages.error(request, 'Both username and email already exist. Contact the administrator.')
                return redirect("/add_employ")  # Redirect here to avoid processing the rest of the view
        if empid:
            try:

                existing_user_empid = Employs.objects.get(empid=empid)
                messages.error(request, 'empid already exists')
                return redirect("/add_employ")  # Redirect here to avoid processing the rest of the view
            except ObjectDoesNotExist:
                existing_user_empid = None
            except MultipleObjectsReturned:
                messages.error(request, 'empid already exists.')
                return redirect("/add_employ")  # Redirect here to avoid processing the rest of the view


    if request.method == "POST":
        first_name = request.POST["first_name"]
        last_name = request.POST["last_name"]
        username = request.POST["username"]
        email = request.POST["email"]
        web_mail = request.POST["web_mail"]
        password = request.POST["password"]
        address = request.POST["address"]
        empid = request.POST["empid"]
        designation_id = request.POST.get("designation")  
        location = request.POST["location"]
        package = request.POST["package"]
        pincode = request.POST["pincode"]
        contactno = request.POST["contactno"]
        bloodgroup = request.POST.get("bloodgroup")
        dateofjoining = request.POST['dateofjoining']
        dateofbirth = request.POST['dateofbirth']
        gender = request.POST["gender"]
        role = request.POST.get('role') # Added field for role
        
        if role == "Employee":
            user = CustomUser.objects.create_user(username=username, password=password, email=email, user_type=2)
            user.employs.first_name = first_name
            user.employs.last_name = last_name
            user.employs.email = email
            user.employs.password = password
            user.employs.address = address
            user.employs.empid = empid
            user.employs.web_mail = web_mail
            user.employs.designation_employ_id = int(designation_id)
            user.employs.location = location
            user.employs.package = package
            user.employs.pincode = pincode
            user.employs.contactno = contactno
            user.employs.bloodgroup = bloodgroup
            user.employs.dateofjoining = dateofjoining
            user.employs.dateofbirth = dateofbirth
            user.employs.gender = gender
            user.employs.role = role 
            user.employs.hroptions=0
            user.employs.companyid=em1
            user.employs.username=username

            user.save()
            messages.success(request, "Employee Added Successfully")
                
        elif role == "HR":
            user = CustomUser.objects.create_user(username=username, password=password, email=email, user_type=2)
            user.employs.first_name = first_name
            user.employs.last_name = last_name
            user.employs.email = email
            user.employs.password = password
            user.employs.address = address
            user.employs.empid = empid
            user.employs.web_mail = web_mail
            user.employs.location = location
            user.employs.package = package
            user.employs.pincode = pincode
            user.employs.contactno = contactno
            user.employs.bloodgroup = bloodgroup
            user.employs.dateofjoining = dateofjoining
            user.employs.gender = gender
            user.employs.role = role  
            user.employs.hroptions=1
            user.employs.companyid=em1
            user.employs.dateofbirth = dateofbirth
            user.employs.username=username

            user.save()
            messages.success(request, "Successfully Added HR")

            
        elif role == "Project Manager":
            user = CustomUser.objects.create_user(username=username, password=password, email=email, user_type=1)
            user.adminhod.firstname = first_name
            user.adminhod.lastname = last_name
            user.adminhod.email = email
            user.adminhod.password = password
            user.adminhod.address = address
            user.adminhod.empid = empid
            user.adminhod.web_mail = web_mail
            user.adminhod.location = location
            user.adminhod.package = package
            user.adminhod.pincode = pincode
            user.adminhod.contactno = contactno
            user.adminhod.bloodgroup = bloodgroup
            user.adminhod.dateofjoining = dateofjoining
            user.adminhod.gender = gender
            user.adminhod.role = role  
            user.adminhod.options=1
            user.employs.companyid=em1
            user.employs.dateofbirth = dateofbirth
            user.employs.username=username

            user.save()
            messages.success(request, "Successfully Added Project Manager")
            
        else:
            user = CustomUser.objects.create_user(username=username, password=password, email=email, user_type=1)
            user.adminhod.firstname = first_name
            user.adminhod.lastname = last_name
            user.adminhod.email = email
            user.adminhod.password = password
            user.adminhod.address = address
            user.adminhod.empid = empid
            user.adminhod.web_mail = web_mail
            user.adminhod.location = location
            user.adminhod.package = package
            user.adminhod.pincode = pincode
            user.adminhod.contactno = contactno
            user.adminhod.bloodgroup = bloodgroup
            user.adminhod.dateofjoining = dateofjoining
            user.adminhod.gender = gender
            user.adminhod.role = role  
            user.adminhod.options=0
            user.employs.companyid=em1
            user.employs.dateofbirth = dateofbirth
            user.employs.username=username

            user.save()
            messages.success(request, "Successfully Added Project Manager")            

       
        html_content = render_to_string("email_template.html", {'title': 'test email', 'first_name': first_name,
                                                                'last_name': last_name, 'empid': empid,'orgname':organization_name,'password':password,})
        text_content = strip_tags(html_content)
        subject = organization_name
        email = EmailMultiAlternatives(
            subject,
            text_content,
            settings.EMAIL_HOST_USER,
            [email],
        )
        email.attach_alternative(html_content, "text/html")
        email.fail_silently = True
        email.send()
        progress = 20
        val = 4
        value_message = 1
        progress_obj, _ = Progress.objects.get_or_create(pk=1, defaults={'value': 0})
        progress_obj.value3 = value_message
        progress_obj.progress_form4 = progress
        progress_obj.save()

        # Try to get the admin_home_drop object with id=5
        instance = admin_home_drop.objects.filter(id=5).first()

        if instance is not None:
            # If the object exists, update its progress_value
            instance.progress_value = val
            instance.save()
        else:
            # Handle the case where the object is not found
            # Decide what action to take or redirect to an appropriate page
            return redirect("/add_employ")  # Redirect to another page

        return redirect("/add_employ")

    

    # If the request method is not POST, render the form
    # s = adminnav.objects.all()
    form = Employs.objects.all()
    a1=working_shifts.objects.all()
    hrform = HR.objects.all()
    role = list.objects.all()
    emp_length=employedata.objects.filter(companyid=regname).first()

    # admin_drops = admin_drop.objects.filter(parent_category=None).order_by('id')
    return render(request, "employ-template/add_employ.html",
                  { "emp_length":emp_length,'plantype':plantype, 'organization_name':organization_name,'role': role, 'hrform': hrform, "form": form,  'li':li,'a1':a1,'data':data,'aa':aa})






def home21(request):
    return render(request,"employhelp/PMatten.html")
def home22(request):
    return render(request,"employhelp/home22.html")

# def home23(request):
#     return render(request,"employhelp/home23.html")

def home24(request):
    return render(request,"employhelp/PMupload.html")

def home25(request):
    return render(request,"employhelp/home25.html")
#hrside

def hrs1(request):
    return render(request,"employhelp/hrs1.html")

def hrs2(request):
    return render(request,"employhelp/hrs2.html")

def hrs3(request):
    return render(request,"employhelp/hrs3.html")

def hrs4(request):
    return render(request,"employhelp/hrs4.html")

def hrs5(request):
    return render(request,"employhelp/hrs5.html")

def hrr7(request):
     return render(request,"employhelp/hrr7.html")


def hrr8(request):
     return render(request,"employhelp/hrr8.html")


def hrs9(request):
     return render(request,"employhelp/hrs9.html")

def hrs10(request):
     return render(request,"employhelp/hrs10.html")
 
def hrs11(request):
     return render(request,"employhelp/hrs11.html")


def testingcards1(request):
    # person=Employs.objects.get(companyid_id=request.user.id)
    employ_obj=Employs.objects.get(admin=request.user.id)



    data=Employs.objects.filter(admin=request.user.id).first()
    
    data2=data.hroptions
    compid=data.companyid
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
 
 



    # hr_user = Employs.objects.filter(id=request.user.id, role__icontains='HR').first()
    persons = Employs.objects.filter(role__icontains='Employ',companyid=request.user.employs.companyid,admin__is_active=True)
    for person in persons:
        # Check if there is a check-in record for today
        checkin_today = checkin.objects.filter(empid=person.email, date=datetime.now().date()).first()

        # Check if there is a check-out record for today
        checkout_today = checkout.objects.filter(empid=person.email, date=datetime.now().date()).first()

        # Determine if the person is currently online
        person.online = checkin_today is not None and checkout_today is None
        person.absent = checkin_today is  None and checkout_today is None
        person.offline = checkin_today is not None and checkout_today is not None


    return render(request,"employ-template/testingcards1.html",{"persons":persons,'employ_obj':employ_obj,'data2':data2,'data':data })

def moni_dashboard1(request, employee_id ):
    employ_obj=Employs.objects.get(admin=request.user.id)

    data=Employs.objects.filter(admin=request.user.id).first()
    data2=data.hroptions
    compid=data.companyid
    companypaln=Companys.objects.filter(id=data.companyid.id).first()
    mon=Addonsuser.objects.filter(companyid=compid,plan="screen monitoring").first()
    mon1=Addonsuser.objects.filter(companyid=compid,plan="detailed monitoring").first()
    mon2=Addonsuser.objects.filter(companyid=compid,plan="powerlogs").first()
    mon3=Addonsuser.objects.filter(companyid=compid,plan="screenshots").first()
    screen=screenmon.objects.filter(id=1).first()
    screen1=screenmon.objects.filter(id=2).first()
    screen2=screenmon.objects.filter(id=3).first()
    screen3=screenmon.objects.filter(id=4).first()
   
       
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
 

    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    try:
        emp_dash = Employs.objects.get(id=employee_id)
    
 
        if emp_dash:
            # Check if there is a check-in record for today
            checkin_today = checkin.objects.filter(empid=emp_dash.email, date=datetime.now().date()).first()

            # Check if there is a check-out record for today
            checkout_today = checkout.objects.filter(empid=emp_dash.email, date=datetime.now().date()).first()

            # Determine if the emp_dash is currently online
            emp_dash.online = checkin_today is not None and checkout_today is None
            emp_dash.absent = checkin_today is  None and checkout_today is None
            emp_dash.offline = checkin_today is not None and checkout_today is not None


        return render(request ,"employ-template/moni_dashboard1.html",{"s":s,"employee_id": employee_id ,'mon':mon,'mon1':mon1,'mon2':mon2,'mon3':mon3,'screen':screen,'screen1':screen1,'screen2':screen2,'screen3':screen3,'companypaln':companypaln, "emp_dash":emp_dash,"compid":compid,"employs_all":employs_all,"h":h,'employ_obj':employ_obj,'data':data,'data2':data2})
    except Employs.DoesNotExist:
        return HttpResponse("Employee not found.")
    
from .models import Monitoring,WorkProductivityDataset,SystemStatus    
def normalize_url(url):
    if url.startswith('http://'):
        url = url[7:]
    elif url.startswith('https://'):
        url = url[8:]
    if url.startswith('www.'):
        url = url[4:]
    return url

def view_app_web1(request, employee_id):
    try:
        employ_obj = Employs.objects.get(admin=request.user.id)
        data = Employs.objects.filter(admin=request.user.id).first()
        data2 = data.hroptions
        compid = data.companyid
        if data2:
            s = employnav.objects.filter(is_name_exist=1, hr_options=1)

        user = CustomUser.objects.filter(id=request.user.id).first()
        userid1 = user.id

        projectm = admin_project_create.objects.filter(admin_id=userid1)
        h = HR.objects.all()
        employs_all = Employs.objects.all()
        compid = Companys.objects.filter(id=request.user.employs.companyid.id).first()

        emp_eyed = Employs.objects.get(id=employee_id)

        date_log = request.GET.get('date_log', '')
        if request.method == 'POST':
            date_log = request.POST.get('date_log', '')

        if date_log:
            m_date_f1 = datetime.strptime(date_log, '%Y-%m-%d')
            m_date_f2 = datetime.strftime(m_date_f1, '%Y-%m-%d')

            moni_details = Monitoring.objects.filter(
                employee_id=employee_id,
                m_log_ts__startswith=m_date_f2
            ).exclude(m_title="").order_by('-m_log_ts').values()
        else:
            moni_details = Monitoring.objects.filter(
                employee_id=employee_id,
            ).order_by('-m_log_ts').values()

        productive_urls = WorkProductivityDataset.objects.filter(w_type='productive', companyid=compid).values_list('w_pds', flat=True)
        unproductive_urls = WorkProductivityDataset.objects.filter(w_type='unproductive', companyid=compid).values_list('w_pds', flat=True)

        productive_urls_normalized = [normalize_url(url) for url in productive_urls]
        unproductive_urls_normalized = [normalize_url(url) for url in unproductive_urls]

        productive_count = 0
        unproductive_count = 0

        for result in moni_details:
            result_title_normalized = normalize_url(result['m_title'])
            result['is_productive'] = any(result_title_normalized.startswith(url) for url in productive_urls_normalized)
            result['is_unproductive'] = any(result_title_normalized.startswith(url) for url in unproductive_urls_normalized)

            if result['is_productive']:
                productive_count += 1
            elif result['is_unproductive']:
                unproductive_count += 1

        undefined_count = len(moni_details) - productive_count - unproductive_count

        # Get a set of all w_pds values
        w_pds_values = set(WorkProductivityDataset.objects.values_list('w_pds', flat=True))

        # Pass moni_details, w_pds_values, and counts to the template
        paginator = Paginator(moni_details, 8)  # Show 8 records per page
        page = request.GET.get('page')

        try:
            moni_details = paginator.page(page)
        except PageNotAnInteger:
            moni_details = paginator.page(1)
        except EmptyPage:
            moni_details = paginator.page(paginator.num_pages)

        return render(request, 'employ-template/ViewMoniLogs1.html', {
            "msg": moni_details,
            "productive_count": productive_count,
            "unproductive_count": unproductive_count,
            "undefined_count": undefined_count,
            "w_pds_values": w_pds_values,
            "emp_eyed": emp_eyed,
            'data': data,
            'data2': data2,
            "compid": compid,
            "s": s,
            "employs_all": employs_all,
            "h": h,
            "date_log": date_log
        })

    except Employs.DoesNotExist:
        return HttpResponse("Employee not found.")
    
def depth_view_app_web1(request, employee_id):
    data=Employs.objects.filter(admin=request.user.id).first()
    
    data2=data.hroptions
    compid=data.companyid
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
 

    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    compid=Companys.objects.filter(usernumber=request.user.id).first()

    try:
        emp_eyed = Employs.objects.get(id=employee_id)



        m_date_f2 = None
        moni_details = []

        date_log = request.GET.get('date_log', '')
        if request.method == 'POST':
            date_log = request.POST.get('date_log', '')

        if date_log:
            m_date_f1 = datetime.strptime(date_log, '%Y-%m-%d')
            m_date_f2 = datetime.strftime(m_date_f1, '%Y-%m-%d')

            moni_details = Monitoring.objects.filter(
                employee_id=employee_id, m_log_ts__startswith=m_date_f2
            ).exclude(m_title="").order_by('-m_log_ts').values()
        else:
            moni_details = Monitoring.objects.filter(
                employee_id=employee_id, 
            ).order_by('-m_log_ts').values()

        for result in moni_details:
            result['start_time'] = result['m_start_time'].strftime("%Y-%m-%d %H:%M:%S")
            result['end_time'] = result['m_end_time'].strftime("%Y-%m-%d %H:%M:%S")

       

        # Get a set of all w_pds values
        w_pds_values = set(WorkProductivityDataset.objects.values_list('w_pds', flat=True))
        paginator = Paginator(moni_details, 8)  # Show 10 records per page
        page = request.GET.get('page')
        
        try:
            moni_details = paginator.page(page)
        except PageNotAnInteger:
            moni_details = paginator.page(1)
        except EmptyPage:
            moni_details = paginator.page(paginator.num_pages)


        return render(request, 'employ-template/ViewDepthMoniLogs1.html', {
            "msg": moni_details,
            "emp_eyed": emp_eyed,
            
            'data':data,
            'data2':data2,
            "w_pds_values": w_pds_values,
            "m_date_f2": m_date_f2, 
            "compid":compid,"s":s,"employs_all":employs_all,"h":h,'date_log':date_log # Include m_date_f2 for the template if needed
        })
    except Employs.DoesNotExist:
        return HttpResponse("Employee not found.")
from .models import Screenshots

from .models import Screenshots
from .models import Screenshots
from django.shortcuts import render, HttpResponse
from datetime import datetime, timedelta
from .models import Employs, Screenshots
from django.shortcuts import render, HttpResponse
from datetime import datetime, timedelta
from .models import Employs, Screenshots

def ss_monitoring1(request, employee_id):


    data=Employs.objects.filter(admin=request.user.id).first()
    
    data2=data.hroptions
    compid=data.companyid
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
 

    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()
    try:
        emp_eyed = Employs.objects.get(id=employee_id)

        

        ss_moni_details = Screenshots.objects.filter(employee_id=employee_id).order_by('-timestamp')

        date_log = request.GET.get('date_log', '')
        if request.method == 'POST':
            date_log = request.POST.get('date_log', '')

        if date_log:
            try:
                ss_moni_details = Screenshots.objects.filter(employee_id=employee_id, timestamp__icontains=date_log).order_by('-timestamp').values()
            except ValueError:
                return HttpResponse("Invalid date format. Please use YYYY-MM-DD.")
        paginator = Paginator(ss_moni_details, 8)  # Show 10 records per page
        page = request.GET.get('page')
            
        try:
            ss_moni_details = paginator.page(page)
        except PageNotAnInteger:
            ss_moni_details = paginator.page(1)
        except EmptyPage:
            ss_moni_details = paginator.page(paginator.num_pages)

        return render(request, 'employ-template/ViewSSMoniLogs1.html', {"msg": ss_moni_details, "emp_eyed": emp_eyed,'data':data,'data2':data2,'h':h,'date_log':date_log})

    except Employs.DoesNotExist:
        return HttpResponse("Employee not found.")



def ss_delete(request,employee_id):
    one_week_ago = datetime.now() - timedelta(days=7)
    Screenshots.objects.filter(employee_id=employee_id,timestamp__lt=one_week_ago).delete()
    return redirect(reverse("ss_monitoring", args=[employee_id]))
def power_monitoring1(request , employee_id):
 
    
    data=Employs.objects.filter(admin=request.user.id).first()
    
    data2=data.hroptions
    compid=data.companyid
    if data2:
       s=employnav.objects.filter(is_name_exist=1,hr_options=1)
    user = CustomUser.objects.filter(id=request.user.id).first()
    userid1 = user.id
 

    projectm = admin_project_create.objects.filter(admin_id=userid1)
    h = HR.objects.all()
    employs_all = Employs.objects.all()

    try:    
        emp_eyed = Employs.objects.get(id=employee_id) 


        
        date_log = request.GET.get('date_log', '')
        if request.method == 'POST':
            date_log = request.POST.get('date_log', '')

        if date_log:
            try:
                # Convert the selected date string to a datetime object
                # selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
                
                # Filter data based on the selected date
                ss_power_details = SystemStatus.objects.filter(employee_id=employee_id,  timestamp__icontains=date_log).order_by('-timestamp').values()
            except ValueError:
                # Handle invalid date format
                return HttpResponse("Invalid date format. Please use YYYY-MM-DD.")
            
        else:
            # If the request is not a POST request, show all records
            ss_power_details = SystemStatus.objects.filter(employee_id=employee_id ).order_by('-timestamp').values() 
        paginator = Paginator(ss_power_details, 8)  # Show 10 records per page
        page = request.GET.get('page')
            
        try:
            ss_power_details = paginator.page(page)
        except PageNotAnInteger:
            ss_power_details = paginator.page(1)
        except EmptyPage:
            ss_power_details = paginator.page(paginator.num_pages)
        return render(request, 'employ-template/ViewPowerMoniLogs1.html', {"msg": ss_power_details , "emp_eyed":emp_eyed,"compid":compid,"s":s,"employs_all":employs_all,"h":h,'data':data,'data2':data2,'date_log':date_log})
    except Employs.DoesNotExist:
        return HttpResponse("Employee not found.")


from notifications.models import Notification

@csrf_exempt
# def employ_mark_as_read(request,notifi_id):
#     if request.method=="POST":  
#         mark = Notification.objects.get(id=notifi_id)
#         if mark.verb == "Payslip Status Approve" or mark.verb == "Payslip Status Reject" :
#             mark.mark_as_read()
#             return redirect("/paysliprequest")
#         elif mark.verb == "Reimbursement Status Approve" or mark.verb == "Reimbursement Status Reject":
#             mark.mark_as_read()
#             return redirect("/reg")
#         elif mark.verb == "Leave Status Accept" or mark.verb == "Leave Status Reject":
#             mark.mark_as_read()
#             return redirect("/employ_apply_leave/")
def employ_mark_as_read(request,notifi_id):
    if request.method=="POST":  
        mark = Notification.objects.get(id=notifi_id)
        sss=mark.actor_object_id
        if  mark.verb == "Admin Notification":
            adminnotificationid=mark.action_object.id
        else:
            adminnotificationid=None
        employ=Employs.objects.filter(admin=sss).first()
        md=admin_message.objects.filter(id=adminnotificationid).first() 
        if mark.verb == "Payslip Status Approve" or mark.verb == "Payslip Status Reject" :
            mark.mark_as_read()
            return redirect("/paysliprequest/")
        elif mark.verb == "Reimbursement Status Approve" or mark.verb == "Reimbursement Status Reject":
            mark.mark_as_read()
            return redirect("/reg")
        elif mark.verb == "Leave Status Accept" or mark.verb == "Leave Status Reject":
            mark.mark_as_read()
            return redirect("/employ_apply_leave/")
        elif mark.verb == "Advancesalary Status Approve" or mark.verb == "Advancesalary Status disapprove":
            mark.mark_as_read()
            return redirect("/advancesalary_request/")
        elif mark.verb == "Leave":
            mark.mark_as_read()
            return HttpResponseRedirect("/individual_employ_leaveReport/"+ str(employ.id))
        elif mark.verb == "Reimbursement":
            mark.mark_as_read()
            return redirect("/individual_reimbursement_view1/"+ str(employ.id))
        elif mark.verb == "Advance salary":
            mark.mark_as_read()
            return redirect("/hr_advancesalary_apply_view/")
        elif mark.verb == "Admin Notification":
            mark.mark_as_read()
            return redirect("/display_message/"+ str(md.id))
        elif mark.verb == "Admin Project Create":
            mark.mark_as_read()
            return redirect("/create-proj1/")
        elif mark.verb =="Extra Time Assigned":
            mark.mark_as_read()
            return redirect("/extratime_notification/")
        elif mark.verb =="Checkout Reminder":
            mark.mark_as_read()
            return redirect("/calendar/")
        elif mark.verb =="CheckInRequest":
            mark.mark_as_read()
            return redirect("checkin_request")
        elif mark.verb == "Admin Project Delete":
            mark.mark_as_read()
            return redirect("/delete_project/")
        elif mark.verb == "Admin Project On Hold":
            mark.mark_as_read()
            return redirect("/onhold_project/")
        elif mark.verb == "gatepass":
            mark.mark_as_read()
            return redirect("/gatepass_data/")
        elif mark.verb == "gatepass_approved":
            mark.mark_as_read()
            return redirect("/gatepass_approve_data/")

def hr_advancesalary_apply_view(request):

  
    employs = Employs.objects.filter(admin=request.user.id).first()
    leaves = ad_salary.objects.all()
    data = CustomUser.objects.filter(id=request.user.id, user_type=1)
    pending = ad_salary.objects.filter(request_status=0).count()
    total_amount = ad_salary.objects.filter(request_status=1).aggregate(Sum('amount'))
    emp_req = ad_salary.objects.filter(student_id__companyid=employs.companyid)
    

    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator(emp_req, items_per_page)
    page = request.GET.get('page')

    try:
        leaves = paginator.page(page)
    except PageNotAnInteger:
        leaves = paginator.page(1)
    except EmptyPage:
        leaves = paginator.page(paginator.num_pages)

    
    return render(request, "employ-template/hradvancesalary.html", {

        'emp_req': emp_req,
        'employs': employs,
        'leaves': leaves,
        'pending_count': pending,
        'total_amount': total_amount['amount__sum'] if total_amount['amount__sum'] else 0
    })



def hradvance_salary_status_approve(request,leave_id):
    leave=ad_salary.objects.get(id=leave_id)
    leave.request_status=1
    leave.save()
    employsss=leave.student_id
    emplo=Employs.objects.filter(id=employsss.id).first()
    compid=emplo.companyid.id
    orgname=Companys.objects.filter(id=compid).first()
    orgnam=orgname.organizationname
    # companys_admin=leave.company_id
    admin_obj = request.user.employs.companyid.id
    companys_admin = Companys.objects.get(id=admin_obj)
    notify.send(sender=companys_admin.usernumber, recipient=employsss.admin, verb='Advancesalary Status Approve', description=f"{orgnam}: Your  request for Advancesalary has been approved" )
    return HttpResponseRedirect(reverse("hr_advancesalary_apply_view"))

def hradvance_salary_status_disapprove(request,leave_id):
    leave=ad_salary.objects.get(id=leave_id)
    leave.request_status=2
    leave.save()
    employsss=leave.student_id
    emplo=Employs.objects.filter(id=employsss.id).first()
    compid=emplo.companyid.id
    orgname=Companys.objects.filter(id=compid).first()
    orgnam=orgname.organizationname
    # companys_admin=leave.company_id
    admin_obj = request.user.employs.companyid.id
    companys_admin = Companys.objects.get(id=admin_obj)
    notify.send(sender=companys_admin.usernumber, recipient=employsss.admin, verb='Advancesalary Status disapprove', description=f"{leave.company_id.organizationname}: Your  request for Advancesalary has been Disapproved" )
    return HttpResponseRedirect(reverse("hr_advancesalary_apply_view"))


            




def home30(request):
    return render(request,"employhelp/home30.html")
def home31(request):
    return render(request,"employhelp/home31.html")
def home32(request):
    return render(request,"employhelp/home32.html")
def home33(request):
    return render(request,"employhelp/home33.html")
def home34(request):
    return render(request,"employhelp/home34.html")
def home35(request):
    return render(request,"employhelp/home35.html")
def home36(request):
    return render(request,"employhelp/home36.html")
def home37(request):
    return render(request,"employhelp/home37.html")


def empleave(request):
    return render(request,"employhelp/empleave.html")
def empreimbursement(request):
    return render(request,"employhelp/empreimbursement.html")
def emppayslip(request):
    return render(request,"employhelp/emppayslip.html")
def empphoto(request):
    return render(request,"employhelp/empphoto.html")
def empinf(request):
    return render(request,"employhelp/empinfo.html")
def empdocument(request):
    return render(request,"employhelp/empdocument.html")
def empattendance(request):
    return render(request,"employhelp/empattendance.html")
def emptask(request):
    return render(request,"employhelp/emptask.html")
def hrempper(request):
    return render(request,"employhelp/hrempper.html")
def hrmonitoring(request):
    return render(request,"employhelp/hrmonitoring.html")
def hrmissinfo(request):
    return render(request,"employhelp/hrmissinfo.html")
def hrphoto(request):
    return render(request,"employhelp/hrphoto.html")


def tlleave(request):
    return render(request,"employhelp/tlleave.html")
def tldoc(request):
    return render(request,"employhelp/tldoc.html")
def tlreimbursment(request):
    return render(request,"employhelp/tlreimbursment.html")
def tlpayslip(request):
    return render(request,"employhelp/tlpayslip.html")
def tlempperform(request):
    return render(request,"employhelp/tlempperform.html")
def tlpersonal(request):
    return render(request,"employhelp/tlpersonal.html")
def tlphoto(request):
    return render(request,"employhelp/tlphoto.html")
def tltaskbypm(request):
    return render(request,"employhelp/tltaskbypm.html")
def tlassign(request):
    return render(request,"employhelp/tlassign.html")
def tlattendance(request):
    return render(request,"employhelp/tlattendance.html")
def tlmeeting(request):
    return render(request,"employhelp/tlmeeting.html")
def tldailytask(request):
    return render(request,"employhelp/tldailytask.html")



import uuid
import os
import zipfile

from django.views.decorators.csrf import csrf_exempt



@csrf_exempt
def generate_user_identifier(request):
    user_identifier = str(uuid.uuid4())
    print(f"Generated User Identifier: {user_identifier}")
    
    secret_file_path = 'secret_user_identifier.txt'
    with open(secret_file_path, 'w') as secret_file:
        secret_file.write(user_identifier)

    return user_identifier





import io
from pathlib import Path

def generate_device_identifier():
    device_identifier = str(uuid.uuid4())
    return device_identifier




from django.http import HttpResponse

import subprocess


@csrf_exempt
def download_installer(request):
  
    exe_file_path = 'dist/buglegal_monitor_app.exe'  # Path to the executable file
    

    if not request.user.employs.desktop_user_identifier:
        # Call the generate_user_identifier view to generate and set the identifier
        user_identifier = generate_user_identifier(request)
        request.user.employs.desktop_user_identifier = user_identifier
        request.user.employs.save()
    else:
        user_identifier = request.user.employs.desktop_user_identifier

    

    user_identifier_str = str(user_identifier)
 

    # Create a zip file containing the executable and secret file in memory
    zip_file = io.BytesIO()
    with zipfile.ZipFile(zip_file, 'w') as zipf:
        zipf.writestr('buglegal_monitor_app.exe', open(exe_file_path, 'rb').read())
        
        zipf.writestr('secret_user_identifier.txt', user_identifier_str.encode())

    # Serve the zip file directly in the response
    response = HttpResponse(zip_file.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=buglegal_monitor.zip'

    return response




def get_employee_and_company(request):
    user_identifier = request.GET.get('user_identifier', None)
    if user_identifier:
        try:
            employee_info = Employs.objects.get(desktop_user_identifier=user_identifier)

            
            employee_id = employee_info.id
            company_id = employee_info.companyid.id
            today = datetime.now().date()
            check_in = checkin.objects.filter(date=today, empid=employee_info.email).first()
            check_out = checkout.objects.filter(date=today, empid=employee_info.email).first()

            response_data = {
                'employee_id': employee_id,
                'company_id': company_id,
                'checkin': check_in is not None and check_out is None,  # True if check_in exists, otherwise False
                'checkout': check_out is not None,  # True if check_out exists, otherwise False
            }

            return JsonResponse(response_data)
            
           

        except Employs.DoesNotExist:
            return JsonResponse({'error': 'User not found'}, status=404)

    return JsonResponse({'error': 'Invalid user_identifier parameter'}, status=400)


@csrf_exempt
def check_launch(request):
    if request.method == 'POST':
        is_launched = request.POST.get('is_launched')
        employee_id = request.POST.get('employee_id')
        company_id = request.POST.get('company_id')
        employee = Employs.objects.filter(id=employee_id, companyid_id=company_id).first()
        
        try:
            # Update the is_launched status for the employee
            employee = Employs.objects.get(id=employee_id, companyid_id=company_id)
            employee.is_launched = is_launched
            employee.save()

            return JsonResponse({'status': 'success'})
        except Employs.DoesNotExist:
            return JsonResponse({'error': 'Employee not found'}, status=404)
    else:
        return JsonResponse({'error': 'Method Not Allowed'}, status=405)




from django.shortcuts import render
from .models import admin_message

def display_admin_messages(request):
    

    data=Employs.objects.filter(admin=request.user.id).first()

    data5=admin_message.objects.filter(role=data.role,companyid_id=data.companyid).order_by("-date")


    return render(request, 'employ-template/admin_messages.html', {'admin_messages': data5,})


def display_message(request,id):
    
    md=admin_message.objects.get(id=id)
    return render(request,"employ-template/admin_messages1.html",{'md':md})


    

from django.shortcuts import render, redirect
from .models import admin_message

def delete_admin_message(request, message_id):
    if request.method == 'POST':
        message = admin_message.objects.get(pk=message_id)
        message.delete()
        messages.error(request,"Removed Successfully")
    return redirect('/display_admin_messages') 



def adminprojects(request):
    

         
    data=Employs.objects.filter(admin=request.user.id).first()
    projectid=admin_project_create.objects.filter(admin_id=data)
    return render(request,"employ-template/pmproject.html",{'projectid':projectid,
            
})

def adminprojects1(request):
    admin_obj = request.user.employs.companyid.id
    companys_admin = Companys.objects.get(id=admin_obj)
         
    data=Employs.objects.filter(admin=request.user.id).first()
    employsss = Employs.objects.get(id=data.id)
    projectid=admin_project_create.objects.filter(admin_id=data)
    return render(request,"employ-template/pmproject1.html",{'projectid':projectid})

def hold_emp(request, pname):
    data=Employs.objects.filter(admin=request.user.id).first()
    mn = TeamMember.objects.filter(companyid=data.companyid,project_id__p_name=pname)
    return render(request, "employ-template/hold_emp.html", {'mn': mn})

def emp_delete(request,id):
    mns=TeamMember.objects.get(id=id)
    mns.delete()
    return redirect('/adminprojects1')
    
    
from datetime import date, datetime, timedelta
from django.shortcuts import render
from django.views import View
from .models import Companys, Employs, LeaveReportEmploy
from .utils import caltablesalary

class SalaryFinalViewemp(View):
    def get(self, request,id):
        data = Employs.objects.filter(admin=request.user.id).first()
        data1 = data.companyid
        salbasic = salary_struct.objects.filter(companyid=data1, salarycomponent="Basic Salary")
        salaryexbasic = salary_struct.objects.filter(companyid=data1).exclude(salarycomponent="Basic Salary")
        alowsum = salary_struct.objects.filter(companyid=data1).exclude(salarycomponent="Basic Salary").aggregate(Sum('percentageofCTC'))['percentageofCTC__sum']
        deduct = salary_deductions.objects.filter(companyid=data1).exclude(percentageordfixed="Fixed").aggregate(Sum('percentageofdctc'))['percentageofdctc__sum']
        deductfix = salary_deductions.objects.filter(companyid=data1, percentageordfixed="Fixed").aggregate(Sum('percentageofdctc'))['percentageofdctc__sum']
        slary = []
        today = date.today()
        payslip=payslip_request.objects.filter(student_id=data,companyid=data1,status=1,id=id).first()
        if payslip:
                        mindate=payslip.startingdate
                        maxdate=payslip.endingdate
                        mindateform=mindate.strftime('%Y-%m')
                        maxdateform=maxdate.strftime('%Y-%m')
        else:
                        mindate=None
                        maxdate=None
                        mindateform=None
                        maxdateform=None
        selected_month = request.GET.get('selected_month', mindateform)
        if selected_month:
           request.session['selected_month']=selected_month
           current_year,current_month = map(int, selected_month.split('-'))
           current_date_formatted = datetime(current_year, current_month, 1).strftime('%B %Y')
           selectcur=current_year+current_month
           curmonth=today.year+today.month
           if selectcur == curmonth :
               num_days_in_month = today.day - 1
               days_list, weekoff_days, weekend_count = caltablesalary(self,current_year, current_month, self.request)

           else:
                num_days_in_month = calendar.monthrange(current_year, current_month)[1]
                days_list, weekoff_days, weekend_count = caltable(self,current_year, current_month, self.request)

           end_date = datetime(current_year, current_month, num_days_in_month).date()
           employees = Employs.objects.filter(companyid=data1,dateofjoining__lte=end_date,admin=request.user.id)


        else:
            selected_month=None
            request.session['selected_month']=selected_month
            current_year=None
            current_month = None
            current_date_formatted =None
            selectcur=None
            curmonth=None
            end_date=None
            employees=None
        if employees:
                for employee in employees:
                    employid=employee.id
                    empid = employee.empid
                    first_name = employee.first_name
                    package = employee.package
                    email = employee.email
                    shifttime=employee.working12
                    dateofjoining=employee.dateofjoining
                    compid=employee.companyid

                    # Calculate total leave days for the selected month
                    start_date = datetime(current_year, current_month, 1).date()
                    end_date = (datetime(current_year, current_month, 1) + timedelta(days=32)).date() - timedelta(days=1)
                    empleaves = LeaveReportEmploy.objects.filter(
                        employ_id=employee,
                        leave_status=1,  # Assuming 1 represents approved leave
                        leave_date__range=[start_date, end_date]
                    )
                    employcheckin = checkin.objects.filter(empid=email, status="present", date__range=[start_date, end_date]).count()
                    employhalfche = checkin.objects.filter(empid=email, status="Halfday", date__range=[start_date, end_date]).count()
                    employcheckout = checkout.objects.filter(empid=email, date__range=[start_date, end_date])
                


                    total_leave_days = sum([leave.get_leave_duration() for leave in empleaves])

                    slary.append({
                        'empid': empid,
                        'first_name': first_name,
                        'total_leave_days': total_leave_days,
                        'selected_month': selected_month,
                        'package': package,
                        'employcheckin': employcheckin,
                        'employhalfche':employhalfche,
                        'current_date_formatted':current_date_formatted,
                        'dateofjoining':dateofjoining,
                        'employid':employid,
                        
                        
                    })
        else:
            employcheckin=None
            employhalfche=None
            num_days_in_month=None
            weekend_count=None
            alowsum=None
            deduct=None
            deductfix=None
            slary=None
            employees=None
            selected_month=None
            current_year=None
            current_month=None
            salbasic=None
            salaryexbasic=None
            payslip=None
       
            
        return render(request, "employ-template/salaryfinalemp.html",{'payslip':payslip,'mindateform':mindateform,'maxdateform':maxdateform,'employcheckin':employcheckin,'employees':employees,'employhalfche':employhalfche,'num_days_in_month':num_days_in_month,'weekend_count': weekend_count, 'alowsum': alowsum, 'deduct': deduct, 'deductfix': deductfix, 'slary': slary, 'employees': employees, 'selected_month': selected_month,'current_year': current_year, 'current_month': current_month, 'salbasic': salbasic, 'salaryexbasic': salaryexbasic})


class emp_payslipemp(View):
    def get(self, request,id):
        data = Employs.objects.filter(admin=request.user.id).first()
        data1 = data.companyid
        employees = Employs.objects.filter(companyid=data1,id=id)
        salbasic = salary_struct.objects.filter(companyid=data1, salarycomponent="Basic Salary")
        salaryexbasic = salary_struct.objects.filter(companyid=data1).exclude(salarycomponent="Basic Salary")
        alowsum = salary_struct.objects.filter(companyid=data1).exclude(salarycomponent="Basic Salary").aggregate(Sum('percentageofCTC'))['percentageofCTC__sum']
        deduct = salary_deductions.objects.filter(companyid=data1).exclude(percentageordfixed="Fixed").aggregate(Sum('percentageofdctc'))['percentageofdctc__sum']
        deductfix = salary_deductions.objects.filter(companyid=data1, percentageordfixed="Fixed").aggregate(Sum('percentageofdctc'))['percentageofdctc__sum']
        slary = []
        today = date.today()
        selected_month=request.session.get('selected_month')
        current_year, current_month = map(int, selected_month.split('-'))
        selectcur=current_year+current_month
        curmonth=today.year+today.month
        if selectcur == curmonth :
               num_days_in_month = today.day-1
               days_list, weekoff_days, weekend_count = caltablesalary(self,current_year, current_month, self.request)

        else:
                num_days_in_month = calendar.monthrange(current_year, current_month)[1]
                days_list, weekoff_days, weekend_count = caltable(self,current_year, current_month, self.request)
        for employee in employees:
            id=employee.id
            empid = employee.empid
            first_name = employee.first_name
            last_name=employee.last_name
            designation=employee.designation
            email=employee.email
            package = employee.package
            email = employee.email

            # Calculate total leave days for the selected month
            start_date = datetime(current_year, current_month, 1).date() 
            end_date = (datetime(current_year, current_month, 1) + timedelta(days=32)).date() - timedelta(days=1)
            empleaves = LeaveReportEmploy.objects.filter(
                employ_id=employee,
                leave_status=1,  # Assuming 1 represents approved leave
                leave_date__range=[start_date, end_date]
            )
            employcheckin = checkin.objects.filter(empid=email, status="present", date__range=[start_date, end_date]).count()
            employhalfche = checkin.objects.filter(empid=email, status="Halfday", date__range=[start_date, end_date]).count()
            employcheckout = checkout.objects.filter(empid=email, date__range=[start_date, end_date])

            total_leave_days = sum([leave.get_leave_duration() for leave in empleaves])

            slary.append({
                'id':id,
                'empid': empid,
                'first_name': first_name,
                'last_name' : last_name,
                'designation': designation,
                'email': email,
                'total_leave_days': total_leave_days,
                'selected_month': selected_month,
                'package': package,
                'employcheckin': employcheckin,
                'employhalfche':employhalfche
            })

        return render(request, "employ-template/employee_detailsemp.html", {'num_days_in_month':num_days_in_month,'weekend_count': weekend_count, 'alowsum': alowsum, 'deduct': deduct, 'deductfix': deductfix, 'slary': slary, 'employees': employees, 'selected_month': selected_month, 'current_year': current_year, 'current_month': current_month, 'salbasic': salbasic, 'salaryexbasic': salaryexbasic})


from .models import create_link  
from datetime import datetime
from django.contrib import messages
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from urllib.parse import quote

def meeting_link(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        date = request.POST['date']
        time = request.POST['time']
        time1 = request.POST['time1']

        p1 = create_link(
            title=title,
            date=date,
            time=time,
            time1=time1,
        )
        p1.save()

        description = "Meeting description goes here."
        event_start = datetime.strptime(f"{date} {time}", "%Y-%m-%d %H:%M")
        event_end = datetime.strptime(f"{date} {time1}", "%Y-%m-%d %H:%M")

        SCOPES = ['https://www.googleapis.com/auth/calendar.events']
        SERVICE_ACCOUNT_INFO = {
            "type": "service_account",
            "project_id": "erudite-acre-421411",
            "private_key_id": "1156ed43b854cd871cfbfdfd4c670860453043dd",
            "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvgIBADANBgkqhkiG9w0BAQEFAASCBKgwggSkAgEAAoIBAQC+OmIK3hzekMD9\nudHk2NlYudg1paeUWvx7c2pAjgdYXXSk/0N+LiFyPNqe0rNftQ/Ycmk4uHgggHCs\nTEehi7vbd0lp/H3ZSsz2iwReXzzsqXN6c736GkJLiyHb1BeZlKAkxiS9qexbGcpC\nmeIAXCs7srxgxIurV7ovpku/Mvd5I3OAEbt2UL0WG8d25Vo0vlJYLV7JbnaYihql\nZWDg4uczwFHZC/UgOtzRL/aalHvPeN5Zwt7lysYW2/FWewTF4/G8xBHatVCkkoSx\nh32S0yHNknFlZO74F5W0QQ1cxqiN9MBvy886ydIpG6J/DJX+tg05WkTdjEB0wXPy\nV7DIol7NAgMBAAECggEAA3BZQOKKiZB62KthCr7L0RGNAc1Elgc+DLpFp7SZl+Sa\nnPf7X57UXnUQq7HZ6cdDadGJquuUylugDdFUMyysQPerVVXp7YFvap6RPeaBWrLm\ngkZw3wJWA/kYK9nt8koQShHl7KLTgcMZ17Hh9skzw+DDjIlJEBgmJsHeGS8B78fm\n+VPZdy+vAH1gdBeFlyDNxXRCAMb/uSCl7CLWgQqSbaApWk0rK2jGXC0t/G4FaV0W\nXPhTR2YpssLou3VKDhOzcB35FNpve5FSVtldBXbyG+Hh/JuCFYHW08SsxdtMFVIN\noXq1rx8oLhl51GyKHKQxwIGyCZstxpkc5yXnkrl44QKBgQD+qeqJRdqZN0bRCoIH\nlV3LY7NMv4AJB4qQQy8eIRhiccFV0G7iXg3TImIdmOh53lDb5f3hC2fMBbIzjKK5\nTMP93zPxSuRjeFrhz5W44ZjNx5Q1qFsIrDqcUDbNoNRVnNaCDe++95Eln72wig8g\nvZwXbASbdIarpp62Dd9dZEz64QKBgQC/OelxKzB3x6ErekbCyhNaw9pJ/ukwHleH\nMGdHzJPNln2clq5dqswfP0xvX8XdBHEMpkqVOOFBEML/PDNC6bNsbE16U4a9GCoL\nGp8XirgEOCAT+2DwLbyZGq4+ifaM7wkvvGCXx9/aqwqpbhOnBYFgjk4JEXaCG8lI\n4hSqVgItbQKBgQDc71sLlu/MRxmIF/6xR5Ok9a1DY0xYCRPcjnzdPcci+R/LLFZW\ny1RhHQ5UMSlXOVmbPVBhonx9k1+Du3odGo1avf8ZUmkMJwlpqKdzSz0/IeviFVpN\nDuaWYDz7bpFoaPXsxNx/J9bj9k8iTtRLClTWOR9SIukHGDnPJf/nzTG/4QKBgBTH\n5FiKCwBZT2h8J0hv0V6hvg0giJLUGUr1taOV1dIXsprbvSMstnYG2PWZgNUd8HMh\ndyV0QvZhLe+yyRdSck8sDddnUFMoVxlR0UwrpQfAWawSO1yee8pfnePp9kMmjucD\nH7Qe4nLO95+l+LjD3mVx0YUmwqws/NjQyNIh+qaZAoGBAPKvcSR+47xM1veVX9L/\np8A/I1GNAKShdX3tIAAiuHS6p0fOk5TEqMa+5TusKW5a+f+uo5Z6azmOCPYgroPX\nrPUwmHGufQaHaOj1icmaJUcQV5NGdVqDm5qZr1XHC+adsgEGeb8YK9Uh5XHTDV1+\nV0YP0nOwRC4dk5vC8jeUnLnh\n-----END PRIVATE KEY-----\n",
            "client_email": "fathima123@erudite-acre-421411.iam.gserviceaccount.com",
            "client_id": "115766482548577355347",
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
            "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
            "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/fathima123%40erudite-acre-421411.iam.gserviceaccount.com",
            "universe_domain": "googleapis.com"
        }

        try:
            credentials = service_account.Credentials.from_service_account_info(
                SERVICE_ACCOUNT_INFO, scopes=SCOPES)
            service = build('calendar', 'v3', credentials=credentials)
            meeting_event = {
                'summary': title,
                'description': description,
                'start': {
                    'dateTime': event_start.isoformat(),
                    'timeZone': 'UTC',
                },
                'end': {
                    'dateTime': event_end.isoformat(),
                    'timeZone': 'UTC',
                },
            }
            event = service.events().insert(calendarId='primary', body=meeting_event).execute()
            event_id = event['id']
            start_iso = event_start.isoformat().replace(":", "").replace("-", "")
            end_iso = event_end.isoformat().replace(":", "").replace("-", "")
            meet_link = f"https://calendar.google.com/calendar/u/0/r/eventedit?text={quote(title)}&dates={start_iso}/{end_iso}"

   
            return redirect(meet_link)

        except HttpError as error:
            messages.error(request, f"Google Calendar API error: {error}")
            return redirect('/meeting_link')
        except Exception as e:
            messages.error(request, f"An error occurred: {e}")
            return redirect('/meeting_link')

    return render(request, 'employ-template/meeting_link.html')






from django.shortcuts import render, redirect
from .models import visitor
from django.utils.timezone import now

def visitor_register(request):
    if request.method == 'POST':
        visitorname = request.POST['visitorname']
        phonenumber = request.POST['phonenumber']
        email = request.POST['email']
        hostname = request.POST['hostname']
        reason = request.POST['reason']
        place = request.POST['place']
        captured_image = request.POST['captured_image']  # Retrieve base64 encoded image data

        v = visitor(
            visitorname=visitorname,
            phonenumber=phonenumber,
            email=email,
            hostname=hostname,
            reason=reason,
            place=place,
            datetime=now(),          )

        if captured_image:
            import base64
            from django.core.files.base import ContentFile

            format, imgstr = captured_image.split(';base64,')
            ext = format.split('/')[-1]
            data = ContentFile(base64.b64decode(imgstr), name=f'{visitorname}_{date}_{time}.{ext}')
            v.camera = data

        v.save()

        return redirect('visitor_register')  # Replace with your desired redirect URL

    return render(request, 'employ-template/visitor_register.html')

def visitor_list(request):
    visitors = visitor.objects.order_by('-datetime')  # Assuming your model is named Visitor
    
    start_date_param = request.GET.get('start_date')
    end_date_param = request.GET.get('end_date')
    message = None
    
    start_date = None
    end_date = None
    
    if start_date_param:
        try:
            start_date = datetime.strptime(start_date_param, '%Y-%m-%d').date()
            visitors = visitors.filter(datetime__date__gte=start_date)
        except ValueError:
            message = "Invalid start date format."
    
    if end_date_param:
        try:
            end_date = datetime.strptime(end_date_param, '%Y-%m-%d').date()
            if start_date_param and start_date:
                if start_date <= end_date:
                    visitors = visitors.filter(datetime__date__lte=end_date)
                else:
                    message = "End Date should be after Start Date."
            else:
                message = "Please select a Start Date before selecting an End Date."
        except ValueError:
            message = "Invalid end date format."
    
    # Debugging: Print SQL query generated
    print(visitors.query)
    
    # Pagination
    items_per_page = 5
    paginator = Paginator(visitors, items_per_page)
    page = request.GET.get('page')
    
    try:
        v = paginator.page(page)
    except PageNotAnInteger:
        v = paginator.page(1)
    except EmptyPage:
        v = paginator.page(paginator.num_pages)
    
    return render(request, 'employ-template/visitor_list.html', {
        'v': v,
        'start_date': start_date_param,
        'end_date': end_date_param,
        'message': message,
    })



    return render(request, 'employ-template/visitor_list.html',{'v':v})
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from .models import visitor
from django.shortcuts import render, get_object_or_404, redirect

def visitor_del(request, id):
    
    visitor1 = get_object_or_404(visitor, id=id)
    
    if request.method == "POST":
        visitor1.delete()
        return redirect("visitor_list")
    
    return render(request, 'employ-template/confirm_delete.html', {'visitor1': visitor1})

def extratime_notification(request):
    user = request.user
    employs = Employs.objects.get(admin=user)
    tasks_with_status=ExtraTimeSlot.objects.all()
    items_per_page = 10  # Adjust the number of items per page as needed

    paginator = Paginator(tasks_with_status, items_per_page)
    page = request.GET.get('page')

    try:
        tasks_with_status = paginator.page(page)
    except PageNotAnInteger:
        tasks_with_status = paginator.page(1)
    except EmptyPage:
        tasks_with_status = paginator.page(paginator.num_pages)

    return render(request,"employ-template/other_notification.html" ,{'employs':employs,'tasks_with_status':tasks_with_status})

def checkin_request(request):
    empid=Employs.objects.filter(admin=request.user.id).first()
    compid=empid.companyid
    shfit=empid.working12
    shiftid=working_shifts.objects.get(shift_name=shfit,companyid=compid)
    if request.method == 'POST':
        cutoff_time = request.POST.get('cutoff_time')
        shift_id = request.POST.get('shift_id')
        request = CheckInRequest.objects.create(
            companyid=compid,
            employ_id=empid,
            cutoff_time=cutoff_time,
            shift_id=shiftid.id,
        )
        notify.send(sender=empid.admin, recipient=compid.usernumber, verb='CheckInRequest', description=f"{empid.first_name}.{empid.last_name} has applied for check_in request")

        return redirect('Employ_home')  
    return render(request,'employ-template/checkin_request.html')



from django.shortcuts import render, redirect
from .models import make_complaints

def make_as_complaints(request):
    empid=Employs.objects.filter(admin=request.user.id).first()
    compid=empid.companyid
    emp_id=empid.empid
    if request.method == 'POST':
        
        description = request.POST.get('description')
        correction = request.POST.get('correction')

        a = make_complaints.objects.create(
            companyid=compid,
            employ_id=emp_id,
            name=empid.first_name+' '+empid.last_name,
            email=empid.email,
            description=description,
            correction=correction
        )
        notify.send(sender=empid.admin, recipient=compid.usernumber, verb='make_complaints', description=f"{empid.first_name}.{empid.last_name} has raised an complaint")
        a.save()
        messages.success(request, 'Complaint has been raised successfully')
        return redirect('/Employ_home')
    return render(request, 'employ-template/make_complaints.html')

def delete_project(request):
    return render(request,"employ-template/admin_delete_project.html")
            
def onhold_project(request):
    return render(request,"employ-template/onhold_project.html")

def delete_data1(request,pid):
    project = get_object_or_404(Project, id=pid)
    project.delete()
    # same_project = admin_project_create.objects.get(project_name=project)
    # same_project.delete()
    messages.success(request, "Project Deleted Successfully")

    return redirect("/Employ_home")


from .models import SOSEvent



def timer(request):
    return render(request, 'employ-template/timer.html')



@csrf_exempt
def sos_trigger(request):
    empid = Employs.objects.filter(admin=request.user.id).first()
    compid = empid.companyid


    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            user_id = data.get('user_id')

            if not user_id:
                return JsonResponse({'success': False, 'error': 'User ID not provided'})

            employ_obj = Employs.objects.get(admin_id=user_id)
            company_obj = employ_obj.companyid
            admin_obj = AdminHod.objects.filter(admin__companys__id=company_obj.id).first()

            if not admin_obj:
                return JsonResponse({'success': False, 'error': 'No AdminHod matches the given query'})

            sos_event = SOSEvent.objects.create(
                employ=employ_obj,
                company=company_obj,
                admin=admin_obj
            )
            notify.send(sender=empid.admin,recipient=compid.usernumber, verb = 'SOS trigger',description=f'{empid.first_name}.{empid.last_name} has triggred the SOS')


            return JsonResponse({'success': True})
        except Employs.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'No Employs matches the given query'})
        except AdminHod.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'No AdminHod matches the given query'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})



def check_sos_events(request):
    sos_events = SOSEvent.objects.filter(seen=False)
    sos_data = [
        {
            "employee": event.employ.first_name + " " + event.employ.last_name,
            "email": event.employ.email,
            "triggered_at": event.triggered_at,
            "company": event.company.organizationname,
            "admin": event.admin.admin.username  
        }
        for event in sos_events
    ]
    return JsonResponse({"new_sos": bool(sos_events), "sos_events": sos_data})

import io
from django.http import HttpResponse
from PIL import Image, ImageDraw, ImageFont
from datetime import datetime
from django.core.mail import EmailMessage

def emp_coupons(request):
    emp = Employs.objects.filter(admin=request.user.id).first()
    compid = emp.companyid
    current_month = datetime.now().month
    current_year = datetime.now().year

    all_coupns = Coupon_emp_all.objects.filter(emp_id=emp, companyid=compid, month=current_month, year=current_year)
    sumofamounts = sum(i.amount for i in all_coupns)
    return render(request, 'employ-template/emp_coupons.html', {'all_coupns': all_coupns, 'sumofamounts': sumofamounts})

def delete_and_download_coupon(request, coupon_id):
    try:
        coupon = get_object_or_404(Coupon_emp_all, id=coupon_id)

        img = Image.new('RGB', (500, 250), color=(255, 255, 255))
        background = Image.open(os.path.join('images', 'coupon.jpg'))
        img.paste(background, (0, 0))  

        d = ImageDraw.Draw(img)
        font_title = ImageFont.truetype('static/plugins/Edu_AU_VIC_WA_NT_Hand/static/EduAUVICWANTHand-Bold.ttf', size=24) 
        font_text = ImageFont.truetype('static/plugins/Edu_AU_VIC_WA_NT_Hand/static/EduAUVICWANTHand-Bold.ttf', size=18)  
        font_text1 = ImageFont.truetype('static/plugins/Edu_AU_VIC_WA_NT_Hand/static/EduAUVICWANTHand-Bold.ttf', size=18)  
        font_text2 = ImageFont.truetype('static/plugins/Edu_AU_VIC_WA_NT_Hand/static/EduAUVICWANTHand-Bold.ttf', size=18)  
        font_text3 = ImageFont.truetype('static/plugins/Edu_AU_VIC_WA_NT_Hand/static/EduAUVICWANTHand-Bold.ttf', size=18)  

        d.text((270, 10), f"{coupon.companyid.organizationname}", fill=(255, 255, 255), font=font_title)
        d.text((270, 50), f"Employee ID : {coupon.emp_id.empid}", fill=(255, 255, 255), font=font_text)
        d.text((270, 90), f"Coupon ID : {coupon.coupon_id}", fill=(255, 255, 255), font=font_text1)
        d.text((270, 130), f"Amount : Rs {coupon.amount}", fill=(255, 255, 255), font=font_text2)
        d.text((270, 180), f"Expiry Month :  {coupon.month},{coupon.year}", fill=(255, 255, 255), font=font_text3)

        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        subject = 'Your Coupon Information'
        message = 'Please find attached your coupon information.'
        email_from = settings.EMAIL_HOST_USER
        recipient_list = [coupon.emp_id.web_mail]
        cc = [coupon.emp_id.email]
        email = EmailMessage(subject, message, email_from, recipient_list,cc)
        email.attach(f'coupon_{coupon.coupon_id}.png', buffer.getvalue(), 'image/png')
        email.send()

        buffer.seek(0)

        response = HttpResponse(buffer, content_type='image/png')
        response['Content-Disposition'] = f'attachment; filename="coupon_{coupon.coupon_id}.png"'

        coupon.delete()

        return response
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return HttpResponse(status=500)  



def delete_data11(request,pid):

    project=admin_project_create.objects.filter(project_name=pid)
    project1=Project.objects.filter(p_name=pid)
  
    project1.delete()
    
    project.delete()
   
    messages.success(request, "Project Deleted Successfully")

    return redirect("/Employ_home")

def teamlead_avg(request):
    data = Employs.objects.filter(admin=request.user.id).first()
    
    if not data:
        return render(request, "employ-template/team_avg.html", {
            'error': 'No data found for the current user.',
        })
        
    data1 = data.id

    tsk1_data1 = employperformance.objects.filter(tl_name_id=data1).values(
        'project_name', 'date', 'tl_name_id__first_name'
    ).annotate(avg_performance=Avg('performance'))
    tsk1_data = employperformance.objects.filter(tl_name_id=data1)


    for item in tsk1_data1:
        AveragePerformance.objects.filter(
            project_name=item['project_name'],
            date=item['date'],
            teamlead_id=data1,
        ).delete()  
        
        AveragePerformance.objects.create(
            project_name=item['project_name'],
            date=item['date'],
            avg_performance=item['avg_performance'],
            teamlead_id=data1,
        )

    tasks_with_status = projecttask.objects.filter(l_name=data1)

    tasks_dict = {}
    for task in tasks_with_status:
        key = (task.p_name.p_name, task.task_date)  
        if key not in tasks_dict:
            tasks_dict[key] = []
        tasks_dict[key].append(task)

    avg_perf_with_tasks = []
    for item in tsk1_data1:
        key = (item['project_name'], item['date'])
        avg_perf_with_tasks.append({
            'teamlead_name': item['tl_name_id__first_name'],
            'date': item['date'],
            'avg_performance': item['avg_performance'],
            'tasks': tasks_dict.get(key, [])
        })

    items_per_page = 10
    paginator = Paginator(tasks_with_status, items_per_page)
    page = request.GET.get('page', 1)

    try:
        tasks_with_status = paginator.page(page)
    except PageNotAnInteger:
        tasks_with_status = paginator.page(1)
    except EmptyPage:
        tasks_with_status = paginator.page(paginator.num_pages)

    return render(request, "employ-template/team_avg.html", {
        'avg_perf_with_tasks': avg_perf_with_tasks,
        'data': data,
        'data1': data1,
        'tsk1_data':tsk1_data,
    })


from django.contrib import messages
from django.utils import timezone
from django.db import transaction

def apply_gatepss(request):
    emp = Employs.objects.filter(admin=request.user.id).first()
    if not emp:
        messages.error(request, "Employee not found.")
        return HttpResponseRedirect(reverse("apply_gatepss"))
   
    compid = emp.companyid
    emp_hroption_1 = Employs.objects.filter(companyid=compid, hroptions=1)
    if request.method == "POST":
        reason = request.POST.get("reason")
        starttime = request.POST.get("starttime")
        endtime = request.POST.get("endtime")

        with transaction.atomic():          
            gate_pass_request = Gate_pass_request1.objects.create(
                emp_id=emp,
                companyid=compid,
                status='0',  
                reason=reason,
                starttime=starttime,
                endtime=endtime
            )
            gate_pass_request.save()
            for emp_hroption in emp_hroption_1:
                notify.send(
                    sender=emp.admin,
                    recipient=emp_hroption.admin,
                    verb='gatepass',
                    description=f"{emp.first_name} {emp.last_name} has requested for gate pass."
                )

        return HttpResponseRedirect(reverse("apply_gatepss"))
   
    return render(request, "employ-template/apply_gatepss.html")


def gatepass_data(request):
    data = Employs.objects.filter(admin=request.user.id).first()
    data1 = data.companyid
    k=Gate_pass_request1.objects.filter(companyid=data1)
    page = request.GET.get('page')
    items_per_page = 10
    paginator = Paginator(k, items_per_page)
    try:
        k = paginator.page(page)
    except PageNotAnInteger:
        k = paginator.page(1)
    except EmptyPage:
        k = paginator.page(paginator.num_pages)
    if request.method == "POST":
        srem = request.POST.get("se")
        if srem:
            k =Gate_pass_request1.objects.filter(companyid_id=data1, emp_id__empid__icontains=srem)

    return render(request,"employ-template/gatepss_data.html",{'k':k})

from django.core.mail import EmailMessage
from django.conf import settings
from .models import EmpGatePass1
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from .models import Gate_pass_request1, EmpGatePass1
from django.contrib import messages
from django.utils import timezone
import os

import qrcode
from barcode import Code128
from barcode.writer import ImageWriter
import hashlib

def employee_registration_and_email(request, id):
    k = get_object_or_404(Gate_pass_request1, id=id)
    if request.method == 'POST':
        empid = request.POST.get('empid')
        emp_id = k.emp_id.id
        emp_name = k.emp_id.first_name + ' ' + k.emp_id.last_name
        reason = k.reason
        starttime=k.starttime
        endtime=k.endtime
        date1=k.date1
        email = k.emp_id.email
        contact_no = k.emp_id.contactno
        designation = k.emp_id.designation
        qr_code_dir = os.path.join(settings.MEDIA_ROOT, 'emp_qr_codes')
        bar_code_dir = os.path.join(settings.MEDIA_ROOT, 'emp_bar_codes')
        os.makedirs(qr_code_dir, exist_ok=True)
        os.makedirs(bar_code_dir, exist_ok=True)
        unique_identifier = hashlib.md5(f"{emp_name}_{email}".encode()).hexdigest()
        numeric_identifier = int(unique_identifier[:12], 16)  # Use the first 12 characters as a numeric code
        current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        qr_data = f" {empid}\n {emp_name}\n {email}\n {contact_no}\n {designation}\n {reason}\n {date1}\n {starttime}\n {endtime}\n {current_datetime}\n {numeric_identifier}\n {k.companyid_id}\n {emp_id} "
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white").convert('RGB')
        qr_code_path = os.path.join(qr_code_dir, f"{emp_name}_{email}.png")
        img.save(qr_code_path)
        barcode_data = f"{numeric_identifier}"  # Use a delimiter to separate fields
        bar_code_path = os.path.join(bar_code_dir, f"{emp_name}_{email}.png")
        barcode_instance = Code128(barcode_data, writer=ImageWriter())
        barcode_instance.save(bar_code_path)

        emp = EmpGatePass1(
            empid=empid,
            emp_name=emp_name,
            reason=reason,
            starttime=starttime,
            endtime=endtime,
            date1=date1,
            email=email,
            contact_no=contact_no,
            designation=designation,
            numeric_identifier=numeric_identifier,
            created_at=current_datetime,
            emp_id_id=emp_id,
            companyid_id=k.companyid_id
        )
        emp.qr_code.name = f"emp_qr_codes/{emp_name}_{email}.png"
        emp.bar_code.name = f"emp_bar_codes/{emp_name}_{email}.png"
        emp.save()

        gate_pass_request = Gate_pass_request1.objects.get(id=id)
        
        emp_gate_pass = EmpGatePass1.objects.filter(emp_id=gate_pass_request.emp_id).first()
        
        if emp_gate_pass and emp_gate_pass.qr_code:
            qr_code_path = os.path.join(settings.MEDIA_ROOT, emp_gate_pass.qr_code.name)
            
            if os.path.exists(qr_code_path):
                hr_user_id = Employs.objects.filter(admin=request.user.id).first()
                subject = f'Mr/Miss {emp_gate_pass.emp_id.first_name}.{emp_gate_pass.emp_id.last_name} - Your GatePass Request has been accepted.'
                body=f'Dear {emp_gate_pass.emp_id.first_name}.{emp_gate_pass.emp_id.last_name},\n\n Your GatePass request has been accepted by {hr_user_id.first_name}.{hr_user_id.last_name},\n\n If you require more information, please find the QR code attached below. '
                with open(qr_code_path, 'rb') as qr_file:
                    email = EmailMessage(
                        subject=subject,
                        body=body,
                        from_email=settings.EMAIL_HOST_USER,
                        to=[emp_gate_pass.email],
                        cc=[emp_gate_pass.emp_id.web_mail],
                    )
                    email.attach('qr_code.png', qr_file.read(), 'image/png')
                    email.send()
                    hr_user = Employs.objects.filter(admin=request.user.id).first()
                    gate_pass_request = get_object_or_404(Gate_pass_request1, id=id)
                    gate_pass_request.status = '1'
                    gate_pass_request.save()
                    emp = gate_pass_request.emp_id

                    notify.send(
                        sender=hr_user.admin,  
                        recipient=emp.admin,
                        verb='gatepass_approved',
                        description=f"Your gate pass request has been approved by {hr_user.first_name}.{hr_user.last_name}"
                    )

                    messages.success(request, "Employee Gate Pass Approved Successfully")
                
        return redirect(reverse('gatepass_data'))
    return render(request, 'employ-template/employee_registration.html', {"k": k})
def gatepass_disapprove(request, gate_id):
    
    hr_user = Employs.objects.filter(admin=request.user.id).first()
    gate_pass_request = get_object_or_404(Gate_pass_request1, id=gate_id)
    gate_pass_request.status = '2'
    gate_pass_request.save()
    emp = gate_pass_request.emp_id

    notify.send(
        sender=hr_user.admin,  
        recipient=emp.admin,
        verb='gatepass_approved',
        description=f" Your gate pass request has been disapproved by{hr_user.first_name}.{hr_user.last_name}"
    )
    
    return HttpResponseRedirect(reverse('gatepass_data'))



def save_qr_qt(request):
    return render(request,'employ-template/CapturedQRCodedata.html')

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import CapturedQRCodedata, Companys, Employs
import json
from datetime import datetime

@csrf_exempt
def save_qr_code_rcp(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            qr_data = data.get('data', '')

            data_lines = qr_data.split('\n')
            
            if len(data_lines) < 13:
                return JsonResponse({'status': 'error', 'message': 'Incomplete data'})

            employid = data_lines[0].strip()
            employname = data_lines[1].strip()
            email = data_lines[2].strip()
            contactno = data_lines[3].strip()
            designation = data_lines[4].strip()
            reason = data_lines[5].strip()
            applydate = data_lines[6].strip()
            starttime = data_lines[7].strip()
            endtime = data_lines[8].strip()
            approvedate = data_lines[9].strip()
            numeric_identifier = data_lines[10].strip()
            companyid = data_lines[11].strip()
            emp_id = data_lines[12].strip()

            try:
                company = Companys.objects.get(id=companyid)
            except Companys.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Company does not exist'})

            try:
                employee = Employs.objects.get(id=emp_id)
            except Employs.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Employee does not exist'})
            
            if request.user.employs.companyid != company:
                return JsonResponse({'status': 'error', 'message': 'This user does not match with this company'})

            try:
                start_datetime = datetime.strptime(f"{applydate} {starttime}", "%Y-%m-%d %H:%M:%S")
                end_datetime = datetime.strptime(f"{applydate} {endtime}", "%Y-%m-%d %H:%M:%S")
            except ValueError as e:
                return JsonResponse({'status': 'error', 'message': f"Failed to parse datetime: {str(e)}"})

            now = datetime.now()

            if not (start_datetime <= now <= end_datetime):
                return JsonResponse({'status': 'error', 'message': 'QR code is not valid at this time'})

            CapturedQRCodedata.objects.create(
                employid=employid,
                employname=employname,
                email=email,
                contactno=contactno,
                designation=designation,
                reason=reason,
                applydate=applydate,
                starttime=starttime,
                endtime=endtime,
                approvedate=approvedate,
                numeric_identifier=numeric_identifier,
                companyid=company,
                emp_id=employee
            )
            
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})




def gatepass_approve_data(request):
    data = Employs.objects.filter(admin=request.user.id).first()
    data1 = data.companyid
    
    gate_pass_requests = Gate_pass_request1.objects.filter(companyid=data1, emp_id=data).select_related('emp_id', 'companyid')
    
    data_for_template = []
    for gate_pass_request in gate_pass_requests:
        emp_gate_pass = EmpGatePass1.objects.filter(emp_id=gate_pass_request.emp_id).first()
        if emp_gate_pass:
            data_for_template.append({
                'created_at': emp_gate_pass.created_at,
                'reason': emp_gate_pass.reason,
                'status': gate_pass_request.status,
                'date1': emp_gate_pass.date1,
                'starttime': emp_gate_pass.starttime,
                'endtime': emp_gate_pass.endtime,
            })
    page = request.GET.get('page')
    items_per_page = 10
    paginator = Paginator(data_for_template, items_per_page)
    try:
        data_for_template = paginator.page(page)
    except PageNotAnInteger:
        data_for_template = paginator.page(1)
    except EmptyPage:
        data_for_template = paginator.page(paginator.num_pages)

    return render(request, 'employ-template/gatepass_approve.html', {'data_for_template': data_for_template})

def gatepass_capture_data(request):
    data = Employs.objects.filter(admin=request.user.id).first()
    data1 = data.companyid
    k=CapturedQRCodedata.objects.filter(companyid=data1)
    page = request.GET.get('page')
    items_per_page = 10
    paginator = Paginator(k, items_per_page)
    try:
        k = paginator.page(page)
    except PageNotAnInteger:
        k = paginator.page(1)
    except EmptyPage:
        k = paginator.page(paginator.num_pages)
    if request.method == "POST":
        srem = request.POST.get("se")
        if srem:
            k =Employee_Qrcodes1.objects.filter(companyid_id=data1, emp_id__empid__icontains=srem)

    return render(request,"employ-template/gatepass_capture_data.html",{'k':k})


